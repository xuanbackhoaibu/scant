import os
import tempfile
import pytest
import openpyxl
import pandas as pd
from app.services.data.workbook_scanner import workbook_scanner
from app.services.data.analysis_intent_parser import analysis_intent_parser
from app.services.data.sheet_resolvers import sheet_resolver, column_resolver
from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine
from app.services.data.action_engine import spreadsheet_action_engine
from app.services.data.workbook_chat_service import workbook_chat_service


@pytest.fixture
def sample_multi_sheet_file():
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb = openpyxl.Workbook()

    # Sheet 1: Bang_luong
    ws1 = wb.active
    ws1.title = "Bang_luong"
    ws1.append(["Mã NV", "Họ tên", "Phòng ban", "Lương cơ bản", "Thực lĩnh"])
    ws1.append(["NV001", "Nguyễn Văn A", "Kinh doanh", 12000000, 15500000])
    ws1.append(["NV002", "Trần Thị B", "Kỹ thuật", 18000000, 22000000])
    ws1.append(["NV003", "Lê Văn C", "Kinh doanh", 10000000, 11500000])
    ws1.append(["NV004", "Phạm Thị D", "Nhân sự", 14000000, 16000000])
    ws1.append(["NV005", "Hoàng Văn E", "Kỹ thuật", 18000000, 22000000])  # Duplicate salary

    # Sheet 2: Tong_hop
    ws2 = wb.create_sheet(title="Tong_hop")
    ws2.append(["Phòng ban", "Số nhân sự", "Tổng quỹ lương"])
    ws2.append(["Kinh doanh", 2, 27000000])
    ws2.append(["Kỹ thuật", 2, 44000000])
    ws2.append(["Nhân sự", 1, 16000000])

    wb.save(path)
    wb.close()
    yield path
    if os.path.exists(path):
        os.remove(path)


def test_workbook_scanner_multi_sheet(sample_multi_sheet_file):
    ctx = workbook_scanner.scan_workbook(sample_multi_sheet_file, source_type="excel")
    assert ctx["sheetCount"] == 2
    assert ctx["fileName"].endswith(".xlsx")
    assert ctx["activeSheet"] == "Bang_luong"
    assert len(ctx["sheets"]) == 2

    s1 = ctx["sheets"][0]
    assert s1["name"] == "Bang_luong"
    assert s1["rowCount"] == 5
    assert len(s1["columns"]) == 5
    # Check data types
    col_names = [c["name"] for c in s1["columns"]]
    assert "Thực lĩnh" in col_names
    thuc_linh_col = next(c for c in s1["columns"] if c["name"] == "Thực lĩnh")
    assert thuc_linh_col["dataType"] == "number"
    assert thuc_linh_col["max"] == 22000000


def test_intent_parser_required_prompts():
    # 1. FIND_MAX
    p1 = analysis_intent_parser.parse("nhân viên lương cao nhất")
    assert p1["intent"] == "FIND_MAX"

    # 2. FIND_MAX on specific metric
    p2 = analysis_intent_parser.parse("ai có thực lĩnh cao nhất")
    assert p2["intent"] == "FIND_MAX"
    assert "thực lĩnh" in p2.get("column_mention", "").lower() or p2["intent"] == "FIND_MAX"

    # 3. FIND_MIN
    p3 = analysis_intent_parser.parse("lương thấp nhất")
    assert p3["intent"] == "FIND_MIN"

    # 4. SUM
    p4 = analysis_intent_parser.parse("tổng thực lĩnh")
    assert p4["intent"] == "SUM"

    # 5. AVERAGE
    p5 = analysis_intent_parser.parse("lương trung bình")
    assert p5["intent"] == "AVERAGE"

    # 6. COUNT
    p6 = analysis_intent_parser.parse("có bao nhiêu nhân viên")
    assert p6["intent"] == "COUNT"

    # 7. FIND_DUPLICATES
    p7 = analysis_intent_parser.parse("tìm dữ liệu trùng và tô vàng")
    assert p7["intent"] == "FIND_DUPLICATES"
    assert p7["color"] == "#FEF08A"

    # 8. FIND_BLANKS
    p8 = analysis_intent_parser.parse("tìm ô trống và tô đỏ")
    assert p8["intent"] == "FIND_BLANKS"
    assert p8["color"] == "#FECACA"

    # 9. DETECT_OUTLIERS
    p9 = analysis_intent_parser.parse("phân tích bất thường")
    assert p9["intent"] == "DETECT_OUTLIERS"

    # 10. CROSS_SHEET_COMPARE
    p10 = analysis_intent_parser.parse("kiểm tra bảng tổng hợp có khớp bảng chi tiết không")
    assert p10["intent"] == "CROSS_SHEET_COMPARE"


@pytest.mark.asyncio
async def test_analyze_action_find_max_real_execution(sample_multi_sheet_file):
    res = await workbook_chat_service.analyze_action(
        file_path=sample_multi_sheet_file,
        prompt="nhân viên thực lĩnh cao nhất",
        sheet_name="Bang_luong",
    )
    assert res["mode"] == "analysis_action"
    assert res["result_type"] == "row"
    assert "22.000.000" in res["answer"]
    assert res["evidence"]["operation"] == "MAX"
    assert len(res["actions"]) > 0


@pytest.mark.asyncio
async def test_analyze_action_sum_real_execution(sample_multi_sheet_file):
    res = await workbook_chat_service.analyze_action(
        file_path=sample_multi_sheet_file,
        prompt="tổng thực lĩnh",
        sheet_name="Bang_luong",
    )
    assert res["mode"] == "analysis_action"
    assert res["result_type"] == "scalar"
    assert "87.000.000" in res["answer"]
    assert res["evidence"]["operation"] == "SUM"


@pytest.mark.asyncio
async def test_analyze_action_duplicates_and_outliers(sample_multi_sheet_file):
    # Duplicates on Bang_luong
    res_dup = await workbook_chat_service.analyze_action(
        file_path=sample_multi_sheet_file,
        prompt="tìm dữ liệu trùng",
        sheet_name="Bang_luong",
    )
    assert res_dup["mode"] == "analysis_action"
    assert res_dup["result_type"] == "duplicate"

    # Outliers should ONLY trigger on explicit outlier prompt
    res_outlier = await workbook_chat_service.analyze_action(
        file_path=sample_multi_sheet_file,
        prompt="phân tích bất thường lương cơ bản",
        sheet_name="Bang_luong",
    )
    assert res_outlier["result_type"] == "outlier"
    assert res_outlier["evidence"]["operation"] == "DETECT_OUTLIERS"


@pytest.mark.asyncio
async def test_chat_greeting_and_metadata_isolation(sample_multi_sheet_file):
    # Greeting should not parse sheet or fail
    chat_res = await workbook_chat_service.chat(
        file_path=sample_multi_sheet_file,
        message="Xin chào",
        sheet_name="Bang_luong",
    )
    assert chat_res["intent"] == "GREETING"
    assert "Xin chào" in chat_res["answer"]

    # Sheet count
    sheet_res = await workbook_chat_service.chat(
        file_path=sample_multi_sheet_file,
        message="File này có bao nhiêu sheet?",
        sheet_name="Bang_luong",
    )
    assert "2 sheet" in sheet_res["answer"]


def test_action_engine_and_undo(sample_multi_sheet_file):
    res = spreadsheet_action_engine.highlight_cells(
        file_path=sample_multi_sheet_file,
        sheet_name="Bang_luong",
        cell_addresses=["E2", "E3"],
        color_hex="#FEF08A",
        session_id="test_session_1",
    )
    assert res["ok"] is True
    assert res["highlighted_count"] == 2

    # Undo
    undo_res = spreadsheet_action_engine.undo_last_action(session_id="test_session_1")
    assert undo_res["ok"] is True
    assert undo_res["restored_count"] == 2


@pytest.mark.asyncio
async def test_chat_entity_count_and_exact_number_follow_up(sample_multi_sheet_file):
    conv_id = "test_chat_entity_session_1"
    # 1. Ask: "có bao nhiêu Kinh doanh"
    res1 = await workbook_chat_service.chat(
        file_path=sample_multi_sheet_file,
        message="có bao nhiêu Kinh doanh",
        sheet_name="Bang_luong",
        conversation_id=conv_id,
    )
    assert "Kinh doanh" in res1["answer"]
    assert "2 lượt xuất hiện" in res1["answer"] or "2 dòng" in res1["answer"]
    assert len(res1["actions"]) > 0

    # 2. Ask follow-up: "bạn hãy cho tôi con số cụ thể"
    res2 = await workbook_chat_service.chat(
        file_path=sample_multi_sheet_file,
        message="bạn hãy cho tôi con số cụ thể",
        sheet_name="Bang_luong",
        conversation_id=conv_id,
    )
    assert "Con số cụ thể" in res2["answer"]
    assert "2" in res2["answer"]
    assert "Mục báo cáo" not in res2["answer"]
    assert "đề tài nghiên cứu" not in res2["answer"]


@pytest.mark.asyncio
async def test_toll_station_exact_queries_my_loc():
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "REPORT_XE_QUA_TRAM"
    ws.append(["STT", "Mã giao dịch", "Trạm vào", "Trạm ra", "Biển số xe", "Loại xe", "Giá tiền"])
    ws.append([1, "3248621617", "QL 39", "Đầu Tuyến", "90H04787", "Loại 3", 84.52])
    ws.append([2, "3248112999", "Đầu Tuyến", "QL 39", "90H04787", "Loại 3", 84.52])
    ws.append([3, "3247020547", "Mỹ Lộc", "Mỹ Lộc", "90H04787", "Loại 3", 34])
    ws.append([4, "3246647260", "Mỹ Lộc", "Mỹ Lộc", "90H04787", "Loại 3", 34])
    ws.append([5, "3245392925", "Mỹ Lộc", "Mỹ Lộc", "90H04787", "Loại 3", 34])
    wb.save(path)
    wb.close()

    try:
        conv_id = "test_my_loc_session"
        # 1. Ask: "có bao nhiêu mỹ lộc"
        res1 = await workbook_chat_service.chat(
            file_path=path,
            message="có bao nhiêu mỹ lộc",
            sheet_name="REPORT_XE_QUA_TRAM",
            conversation_id=conv_id,
        )
        assert "mỹ lộc" in res1["answer"].lower() or "my loc" in remove_diacritics(res1["answer"]).lower()
        # In this sheet, Mỹ Lộc appears in 3 rows (6 total cells: 3 in Trạm vào, 3 in Trạm ra)
        assert "6 lượt xuất hiện" in res1["answer"]
        assert "3 dòng" in res1["answer"]
        assert len(res1["actions"]) > 0

        # 2. Ask: "bạn hãy cho tôii con số cụ thể" (with typo in 'tôii')
        res2 = await workbook_chat_service.chat(
            file_path=path,
            message="bạn hãy cho tôii con số cụ thể",
            sheet_name="REPORT_XE_QUA_TRAM",
            conversation_id=conv_id,
        )
        assert "Con số cụ thể" in res2["answer"]
        assert "6" in res2["answer"]
        assert "Mục báo cáo" not in res2["answer"]
    finally:
        if os.path.exists(path):
            os.remove(path)


def test_cross_file_compare():
    # File 1: Toll report
    fd1, path1 = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd1)
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Sheet1"
    ws1.append(["Mã giao dịch", "Biển số", "Giá tiền"])
    ws1.append(["GD001", "90H-1111", 50000])
    ws1.append(["GD002", "90H-2222", 75000])
    ws1.append(["GD003", "90H-3333", 100000])
    wb1.save(path1)
    wb1.close()

    # File 2: Bank statement (GD002 has mismatched price 80000, GD004 only in bank)
    fd2, path2 = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd2)
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Sheet1"
    ws2.append(["Mã giao dịch", "Biển số", "Giá tiền"])
    ws2.append(["GD001", "90H-1111", 50000])
    ws2.append(["GD002", "90H-2222", 80000])  # Mismatch
    ws2.append(["GD004", "90H-4444", 120000]) # Only in File 2
    wb2.save(path2)
    wb2.close()

    try:
        res = spreadsheet_query_engine.cross_file_compare(
            file_path_1=path1,
            file_path_2=path2,
            key_column_1="Mã giao dịch",
            key_column_2="Mã giao dịch",
        )
        assert res["total_keys"] == 4
        assert res["common_count"] == 2
        assert res["in_file1_only_count"] == 1  # GD003
        assert res["in_file2_only_count"] == 1  # GD004
        assert res["value_differences_count"] == 1  # GD002 Giá tiền
    finally:
        if os.path.exists(path1):
            os.remove(path1)
        if os.path.exists(path2):
            os.remove(path2)


