from pathlib import Path

import openpyxl
import pytest


def create_payroll_workbook(path: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang_luong"
    headers = ["Ma NV", "Ho ten", "Phong ban", "Luong co ban", "Thuc linh", "Ngay cong", "Ghi chu", "Thuong"]
    rows = [
        ["NV001", "Nguyen An", "Kinh doanh", 12000000, 14500000, 26, "", 1000000],
        ["NV002", "Tran Binh", "Marketing", 11000000, 13200000, 24, None, 500000],
        ["NV003", "Le Chi", "Ky thuat", 17000000, 22100000, 27, "top", 3000000],
        ["NV004", "Pham Dung", "Kinh doanh", 10000000, 10800000, 21, "", 200000],
        ["NV007", "Vu Quoc Khanh", "Ky thuat", 9000000, 9500000, 18, "can xem", 0],
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)

    ws2 = wb.create_sheet("Doanh_thu")
    ws2.append(["Thang", "Doanh thu"])
    ws2.append(["T1", 100])
    ws2.append(["T2", 90])
    wb.save(path)
    wb.close()
    return str(path)


def create_multi_sheet_workbook(path: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang_luong"
    ws.append(["Ma NV", "Thuc linh", "Ghi chu"])
    ws.append(["NV001", 12000000, ""])
    ws.append(["NV002", 9500000, None])
    ws.append(["NV003", 18000000, "ok"])

    ws2 = wb.create_sheet("Tong_hop")
    ws2.append(["Loai", "Gia tri"])
    ws2.append(["Tong", 39500000])
    ws2.append(["Thieu", None])

    ws3 = wb.create_sheet("Cham_cong")
    ws3.append(["Ma NV", "Ngay cong"])
    ws3.append(["NV001", 26])
    ws3.append(["NV001", 26])
    ws3.append(["NV003", 27])

    wb.save(path)
    wb.close()
    return str(path)


def test_query_engine_resolves_schema_and_computes_grounded_metrics(tmp_path):
    from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    info = spreadsheet_query_engine.get_workbook_info(file_path, active_sheet="bang luong")
    assert info["active_sheet"] == "Bang_luong"
    assert info["sheet_names"] == ["Bang_luong", "Doanh_thu"]
    assert info["evidence"]["operation"] == "GET_WORKBOOK_INFO"

    schema = spreadsheet_query_engine.get_sheet_schema(file_path, "Bảng lương")
    assert schema["sheet"] == "Bang_luong"
    assert [c["name"] for c in schema["columns"]][:3] == ["Ma NV", "Ho ten", "Phong ban"]
    assert schema["columns"][4]["letter"] == "E"

    col = spreadsheet_query_engine.find_column(file_path, "Bang_luong", "thực lĩnh")
    assert col["name"] == "Thuc linh"
    assert col["letter"] == "E"
    assert col["confidence"] >= 0.8

    total = spreadsheet_query_engine.aggregate_column(file_path, "Bang_luong", "Thực lĩnh", "sum")
    assert total["value"] == 70100000
    assert total["evidence"] == {
        "sheet": "Bang_luong",
        "ranges": ["E2:E6"],
        "operation": "SUM",
        "rowCount": 5,
    }

    top = spreadsheet_query_engine.find_top_rows(file_path, "Bang_luong", "lương cơ bản", limit=1)
    assert top["rows"][0]["row_number"] == 4
    assert top["rows"][0]["record"]["Ma NV"] == "NV003"
    assert top["evidence"]["ranges"] == ["D2:D6"]


def test_query_engine_search_outliers_and_group_compare(tmp_path):
    from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    found = spreadsheet_query_engine.search_rows(file_path, "Bang_luong", "NV007")
    assert found["matches"][0]["row_number"] == 6
    assert found["matches"][0]["record"]["Ho ten"] == "Vu Quoc Khanh"
    assert found["evidence"]["operation"] == "SEARCH_ROWS"

    compared = spreadsheet_query_engine.compare_groups(
        file_path,
        "Bang_luong",
        group_column="phong ban",
        value_column="thuc linh",
        op="sum",
    )
    values = {row["group"]: row["value"] for row in compared["groups"]}
    assert values["Kinh doanh"] == 25300000
    assert values["Ky thuat"] == 31600000
    assert compared["evidence"]["ranges"] == ["C2:C6", "E2:E6"]

    outliers = spreadsheet_query_engine.detect_outliers(file_path, "Bang_luong", "ngay cong")
    assert outliers["operation"] == "DETECT_OUTLIERS"
    assert outliers["evidence"]["sheet"] == "Bang_luong"
    assert outliers["evidence"]["ranges"] == ["F2:F6"]


@pytest.mark.asyncio
async def test_workbook_chat_answers_read_only_questions_from_real_workbook(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    total = await workbook_chat_service.chat(
        file_path=file_path,
        message="Tổng thực lĩnh là bao nhiêu?",
        sheet_name="Bang_luong",
        conversation_id="grounded_total",
    )
    assert total["intent"] == "aggregate"
    assert total["answer"].find("70.100.000") >= 0
    assert total["evidence"]["operation"] == "SUM"
    assert total["evidence"]["ranges"] == ["E2:E6"]
    assert "Bạn có thể yêu cầu" not in total["answer"]

    top = await workbook_chat_service.chat(
        file_path=file_path,
        message="Ai có lương cơ bản cao nhất?",
        sheet_name="Bang_luong",
        conversation_id="grounded_followup",
    )
    assert top["intent"] == "sort"
    assert "NV003" in top["answer"]
    assert "Le Chi" in top["answer"]
    assert top["follow_up_context"]["entity"]["row_number"] == 4

    followup = await workbook_chat_service.chat(
        file_path=file_path,
        message="Thực lĩnh của người đó?",
        sheet_name="Bang_luong",
        conversation_id="grounded_followup",
    )
    assert "22.100.000" in followup["answer"]
    assert followup["evidence"]["ranges"] == ["E4"]

    employee = await workbook_chat_service.chat(
        file_path=file_path,
        message="NV007 có thông tin gì?",
        sheet_name="Bang_luong",
        conversation_id="grounded_employee",
    )
    assert "Vu Quoc Khanh" in employee["answer"]
    assert "Ky thuat" in employee["answer"]
    assert employee["evidence"]["operation"] == "SEARCH_ROWS"

    average = await workbook_chat_service.chat(
        file_path=file_path,
        message="Trung bình lương cơ bản?",
        sheet_name="Bang_luong",
        conversation_id="grounded_average",
    )
    assert "11.800.000" in average["answer"]
    assert average["evidence"]["operation"] == "AVERAGE"

    min_salary = await workbook_chat_service.chat(
        file_path=file_path,
        message="Ai có lương cơ bản thấp nhất?",
        sheet_name="Bang_luong",
        conversation_id="grounded_min_salary",
    )
    assert "NV007" in min_salary["answer"]
    assert "Vu Quoc Khanh" in min_salary["answer"]
    assert min_salary["evidence"]["operation"] == "MIN"

    bonus = await workbook_chat_service.chat(
        file_path=file_path,
        message="Trong cột Thưởng ai cao nhất?",
        sheet_name="Bang_luong",
        conversation_id="grounded_bonus_not_sheet",
    )
    assert bonus["context"]["sheet"] == "Bang_luong"
    assert bonus["evidence"]["operation"] == "MAX"
    assert bonus["evidence"]["ranges"] == ["H2:H6"]
    assert "sheet **Thuong**" not in bonus["answer"]


@pytest.mark.asyncio
async def test_workbook_chat_routes_numeric_filters_before_text_search(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    cases = [
        ("tất cả lương thực lĩnh trên 10 triệu", ">", 10000000, 4),
        ("nhân viên có thực lĩnh > 10tr", ">", 10000000, 4),
        ("thực lĩnh dưới 9 triệu", "<", 9000000, 0),
    ]
    for idx, (message, operator, threshold, expected_count) in enumerate(cases):
        result = await workbook_chat_service.chat(
            file_path=file_path,
            message=message,
            sheet_name="Bang_luong",
            conversation_id=f"filter_route_{idx}",
            selected_range="Q2",
        )
        assert result["intent"] == "filter_rows"
        assert result["evidence"]["operation"] == "FILTER_ROWS"
        assert result["result"]["column"]["name"] == "Thuc linh"
        assert result["result"]["operator"] == operator
        assert result["result"]["compare_value"] == threshold
        assert result["result"]["matched_count"] == expected_count
        assert "Tìm thấy" in result["answer"]
        assert "từ khóa" not in result["answer"]
        assert result["context"]["ranges"] == ["E2:E6"]

    highlighted = await workbook_chat_service.chat(
        file_path=file_path,
        message="tô vàng người có thực lĩnh trên 10 triệu",
        sheet_name="Bang_luong",
        conversation_id="filter_route_highlight",
    )
    assert highlighted["intent"] == "filter_rows"
    assert highlighted["result"]["matched_count"] == 4
    assert highlighted["actions"][0]["type"] == "HIGHLIGHT_CELLS"
    assert highlighted["actions"][0]["cells"] == ["E2", "E3", "E4", "E5"]

    ranged = await workbook_chat_service.chat(
        file_path=file_path,
        message="thực lĩnh từ 10 đến 15 triệu",
        sheet_name="Bang_luong",
        conversation_id="filter_route_range",
    )
    assert ranged["intent"] == "filter_rows"
    assert ranged["result"]["operator"] == "between"
    assert ranged["result"]["compare_value"] == [10000000, 15000000]
    assert ranged["result"]["matched_count"] == 3

    search = await workbook_chat_service.chat(
        file_path=file_path,
        message="tìm chữ Linh",
        sheet_name="Bang_luong",
        conversation_id="filter_route_search_text",
    )
    assert search["intent"] == "search_text"
    assert search["evidence"]["operation"] == "SEARCH_AND_COUNT"


@pytest.mark.asyncio
async def test_workbook_chat_does_not_treat_semantic_column_as_sheet(tmp_path):
    from app.services.data.workbook_chat_service import WorkbookChatService, workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")
    available_sheets = ["Bang_luong", "Tong_hop"]

    message = "Ai có giá trị cao nhất trong cột số quan trọng nhất?"
    assert WorkbookChatService.extract_sheet_mention_from_text(message, available_sheets) is None
    assert WorkbookChatService.resolve_target_sheet(message, "Bang_luong", available_sheets) == "Bang_luong"

    result = await workbook_chat_service.chat(
        file_path=file_path,
        message=message,
        sheet_name="Bang_luong",
        conversation_id="semantic_primary_metric",
    )

    assert result["context"]["sheet"] == "Bang_luong"
    assert result["evidence"]["operation"] == "MAX"
    assert result["evidence"]["ranges"] == ["E2:E6"]
    assert result["result"]["column"]["name"] == "Thuc linh"
    assert "Le Chi" in result["answer"]
    assert "NV003" in result["answer"]
    assert "sheet **cột số quan trọng nhất**" not in result["answer"]
    assert "sheet **cot so quan trong nhat**" not in result["answer"]
    assert "Bạn có thể yêu cầu" not in result["answer"]


@pytest.mark.asyncio
async def test_workbook_chat_column_unclear_fallback_lists_real_numeric_columns(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    result = await workbook_chat_service.chat(
        file_path=file_path,
        message="Tính tổng cột nào đó",
        sheet_name="Bang_luong",
        conversation_id="unclear_column_fallback",
    )

    assert result["context"]["sheet"] == "Bang_luong"
    assert result["evidence"]["operation"] == "GET_SHEET_SCHEMA"
    assert "Tôi chưa xác định chắc chắn cột bạn muốn phân tích" in result["answer"]
    assert "Luong co ban" in result["answer"]
    assert "Thuc linh" in result["answer"]
    assert "Bạn có thể yêu cầu tôi kiểm tra trùng lặp" not in result["answer"]


@pytest.mark.asyncio
async def test_workbook_chat_summarizes_and_compares_groups_without_hardcoding(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    summary = await workbook_chat_service.chat(
        file_path=file_path,
        message="Sheet này có vấn đề gì?",
        sheet_name="Bang_luong",
        conversation_id="grounded_summary",
    )
    assert "5 dòng dữ liệu" in summary["answer"]
    assert "ô trống" in summary["answer"]
    assert summary["evidence"]["operation"] == "SHEET_SUMMARY"

    compared = await workbook_chat_service.chat(
        file_path=file_path,
        message="So sánh Kinh doanh với Marketing theo thực lĩnh",
        sheet_name="Bang_luong",
        conversation_id="grounded_compare",
    )
    assert "Kinh doanh" in compared["answer"]
    assert "Marketing" in compared["answer"]
    assert "25.300.000" in compared["answer"]
    assert "13.200.000" in compared["answer"]
    assert compared["evidence"]["operation"] == "SUM_BY_GROUP"


@pytest.mark.asyncio
async def test_workbook_chat_uses_selected_range_and_suggests_near_sheet(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    selected = await workbook_chat_service.chat(
        file_path=file_path,
        message="Có ô trống không?",
        sheet_name="Bang_luong",
        selected_range="G2:G6",
        conversation_id="selected_range_ctx",
    )
    assert selected["context"]["ranges"] == ["G2:G6"]
    assert selected["result"]["missing_count"] == 3
    assert selected["evidence"]["ranges"] == ["G2:G6"]

    missing_sheet = await workbook_chat_service.chat(
        file_path=file_path,
        message="Phân tích sheet doanhthu",
        sheet_name="Bang_luong",
        conversation_id="missing_sheet_ctx",
    )
    assert missing_sheet["error"]["code"] == "SHEET_NOT_FOUND"
    assert missing_sheet["error"]["suggested_sheet"] == "Doanh_thu"
    assert "Doanh_thu" in missing_sheet["answer"]


@pytest.mark.asyncio
async def test_workbook_chat_action_request_returns_pending_confirmation(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    response = await workbook_chat_service.chat(
        file_path=file_path,
        message="Tô vàng các dòng có thực lĩnh dưới 10 triệu",
        sheet_name="Bang_luong",
        conversation_id="pending_action_ctx",
    )
    assert response["pending_actions"]
    assert response["pending_actions"][0]["type"] == "HIGHLIGHT_ROWS"
    assert response["pending_actions"][0]["requires_confirmation"] is True
    assert response["pending_actions"][0]["rows"] == [6]
    assert response["actions"] == []


@pytest.mark.asyncio
async def test_workbook_analysis_action_runs_structured_jobs(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_payroll_workbook(tmp_path / "payroll.xlsx")

    duplicate = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Tìm dữ liệu trùng trong A2:A6 và tô vàng",
        sheet_name="Bang_luong",
    )
    assert duplicate["mode"] == "analysis_action"
    assert duplicate["result"]["duplicate_count"] == 0
    assert duplicate["evidence"]["operation"].startswith("FIND_DUPLICATES")
    assert duplicate["analysis_history_item"]["prompt"] == "Tìm dữ liệu trùng trong A2:A6 và tô vàng"

    blank = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Tìm ô trống",
        sheet_name="Bang_luong",
        selected_range="G2:G6",
    )
    assert blank["context"]["ranges"] == ["G2:G6"]
    assert blank["result"]["missing_count"] == 3
    assert blank["actions"][0]["type"] == "HIGHLIGHT_CELLS"

    outlier = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Phân tích bất thường ngày công",
        sheet_name="Bang_luong",
    )
    assert outlier["result"]["operation"] == "DETECT_OUTLIERS"
    assert outlier["evidence"]["ranges"] == ["F2:F6"]


@pytest.mark.asyncio
async def test_workbook_analysis_action_honors_structured_scope(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_multi_sheet_workbook(tmp_path / "multi.xlsx")

    workbook_scope = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Tóm tắt workbook và tìm ô trống",
        sheet_name="Bang_luong",
        scope={"type": "workbook"},
    )
    assert workbook_scope["mode"] == "analysis_action"
    assert workbook_scope["evidence"]["operation"] == "WORKBOOK_ANALYSIS"
    assert workbook_scope["context"]["sheet"] == "workbook"
    assert {item["sheet"] for item in workbook_scope["result"]["sheet_results"]} == {"Bang_luong", "Tong_hop", "Cham_cong"}
    assert "Tong_hop" in workbook_scope["answer"]

    selected_sheets = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Tìm ô trống",
        sheet_name="Bang_luong",
        scope={"type": "sheets", "sheets": ["Bang_luong", "Tong_hop"]},
    )
    assert selected_sheets["evidence"]["operation"] == "MULTI_SHEET_ANALYSIS"
    assert [item["sheet"] for item in selected_sheets["result"]["sheet_results"]] == ["Bang_luong", "Tong_hop"]

    ranged = await workbook_chat_service.analyze_action(
        file_path=file_path,
        prompt="Tìm ô trống",
        sheet_name="Bang_luong",
        scope={"type": "range", "sheet": "Bang_luong", "range": "C2:C4"},
    )
    assert ranged["context"]["sheet"] == "Bang_luong"
    assert ranged["context"]["ranges"] == ["C2:C4"]
    assert ranged["result"]["missing_count"] == 2


@pytest.mark.asyncio
async def test_workbook_chat_handles_conversational_and_metadata_intents(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service

    file_path = create_multi_sheet_workbook(tmp_path / "multi_chat.xlsx")

    # 1. Greeting
    greeting = await workbook_chat_service.chat(
        file_path=file_path,
        message="Xin chào",
        sheet_name="Bang_luong",
    )
    assert greeting["intent"] == "GREETING"
    assert "Xin chào" in greeting["answer"]
    assert "error" not in greeting

    # 2. Help
    help_res = await workbook_chat_service.chat(
        file_path=file_path,
        message="Bạn làm được gì?",
        sheet_name="Bang_luong",
    )
    assert help_res["intent"] == "HELP"
    assert "Tra cứu & Tổng hợp" in help_res["answer"]

    # 3. Small talk
    thanks = await workbook_chat_service.chat(
        file_path=file_path,
        message="Cảm ơn bạn nhiều",
        sheet_name="Bang_luong",
    )
    assert thanks["intent"] == "SMALL_TALK"
    assert "hỗ trợ" in thanks["answer"]

    # 4. Sheet count metadata
    sheet_count = await workbook_chat_service.chat(
        file_path=file_path,
        message="File này có bao nhiêu sheet?",
        sheet_name="Bang_luong",
    )
    assert sheet_count["intent"] == "WORKBOOK_QUESTION"
    assert "3 sheet" in sheet_count["answer"]
    assert "Bang_luong" in sheet_count["answer"]
    assert "Tong_hop" in sheet_count["answer"]

    # 5. Current sheet
    cur_sheet = await workbook_chat_service.chat(
        file_path=file_path,
        message="Sheet hiện tại tên gì?",
        sheet_name="Bang_luong",
    )
    assert "Bang_luong" in cur_sheet["answer"]

    # 6. Max sheet & follow-up
    max_sheet = await workbook_chat_service.chat(
        file_path=file_path,
        message="Sheet nào lớn nhất?",
        sheet_name="Bang_luong",
        conversation_id="conv_followup_test",
    )
    assert "Bang_luong" in max_sheet["answer"]

    followup = await workbook_chat_service.chat(
        file_path=file_path,
        message="Sheet đó có bao nhiêu dòng?",
        conversation_id="conv_followup_test",
    )
    assert "Bang_luong" in followup["answer"]
    assert "dòng dữ liệu" in followup["answer"]
