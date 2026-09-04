import pytest
import pytest_asyncio
import pandas as pd
import json
from pathlib import Path
from httpx import AsyncClient, ASGITransport
from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession, async_sessionmaker
from app.main import app
from app.core.database import Base, get_db
from app.core.config import settings
from app.services.data.data_engine import data_engine

TEST_DB_URL = "sqlite+aiosqlite:///:memory:"
test_engine = create_async_engine(TEST_DB_URL, connect_args={"check_same_thread": False})
TestAsyncSession = async_sessionmaker(bind=test_engine, class_=AsyncSession, expire_on_commit=False)


@pytest_asyncio.fixture(scope="function")
async def db_session():
    async with test_engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    async with TestAsyncSession() as session:
        yield session
    async with test_engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)


@pytest_asyncio.fixture(scope="function")
async def client(db_session: AsyncSession):
    async def override_get_db():
        yield db_session
    app.dependency_overrides[get_db] = override_get_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as c:
        yield c
    app.dependency_overrides.clear()


def test_data_engine_profiling_and_aggregation():
    # Create sample CSV
    data = {
        "Region": ["North", "South", "North", "Central", "South", "North"],
        "Product": ["EV Sedan", "EV SUV", "EV Sedan", "EV Truck", "EV SUV", "EV Sedan"],
        "Revenue": [1200, 1800, 1100, 950, 2200, 1350],
        "UnitsSold": [12, 15, 11, 8, 20, 14],
    }
    df = pd.DataFrame(data)
    csv_path = settings.UPLOAD_DIR / "sample_sales.csv"
    df.to_csv(str(csv_path), index=False)

    # 1. Test profile
    profile = data_engine.profile_dataset(str(csv_path))
    assert profile["total_rows"] == 6
    assert profile["total_columns"] == 4
    assert len(profile["columns"]) == 4

    # 2. Test aggregate by Region sum Revenue
    agg = data_engine.aggregate_data(str(csv_path), group_by="Region", metric_column="Revenue", aggregation="sum")
    assert "North" in agg["labels"]
    assert "South" in agg["labels"]
    assert len(agg["values"]) == 3

    # 3. Test chart spec
    chart = data_engine.build_chart_specification(str(csv_path), chart_type="bar", group_by="Region", metric_column="Revenue")
    assert chart["chart_type"] == "bar"
    assert len(chart["datasets"][0]["data"]) == 3


def test_data_engine_profiles_selected_excel_sheet_and_range(tmp_path):
    file_path = tmp_path / "multi_sheet.xlsx"
    with pd.ExcelWriter(file_path) as writer:
        pd.DataFrame({"Ignore": [1, 2]}).to_excel(writer, sheet_name="Other", index=False)
        pd.DataFrame({
            "Nhan vien": ["A", "B", "C"],
            "Phong ban": ["Ke toan", "Kinh doanh", "Nhan su"],
            "Luong": [100, 200, 300],
            "Ghi chu": ["x", "y", "z"],
        }).to_excel(writer, sheet_name="BangLuong", index=False)

    profile = data_engine.profile_dataset(str(file_path), sheet_range="BangLuong!A1:C3")

    assert profile["selection"]["sheet_range"] == "BangLuong!A1:C3"
    assert profile["sheet_count"] == 1
    assert profile["sheets"][0]["name"] == "BangLuong"
    assert profile["total_rows"] == 2
    assert profile["total_columns"] == 3
    assert [col["name"] for col in profile["columns"]] == ["Nhan vien", "Phong ban", "Luong"]
    assert profile["preview_rows"][1]["Luong"] == 200


@pytest.mark.asyncio
async def test_data_api_flow(client: AsyncClient):
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "data_lead@corp.com",
        "password": "Password123!",
        "name": "Data Lead"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    proj_res = await client.post("/api/v1/projects", json={
        "name": "Dự án Dữ liệu Doanh thu",
        "type": "data_analysis"
    }, headers=headers)
    project_id = proj_res.json()["id"]

    # Upload CSV file
    csv_content = b"Department,Budget,Staff\nSales,50000,20\nMarketing,30000,10\nEngineering,80000,35\n"
    files = {"file": ("departments.csv", csv_content, "text/csv")}
    data = {"project_id": project_id, "document_type": "dataset"}
    upload_res = await client.post("/api/v1/files/upload", data=data, files=files, headers=headers)
    file_id = upload_res.json()["id"]

    # Profile dataset
    profile_res = await client.get(f"/api/v1/data/profile/{file_id}", headers=headers)
    assert profile_res.status_code == 200
    assert profile_res.json()["total_rows"] == 3

    # Aggregate
    agg_res = await client.post("/api/v1/data/aggregate", json={
        "file_id": file_id,
        "group_by": "Department",
        "metric_column": "Budget",
        "aggregation": "sum"
    }, headers=headers)
    assert agg_res.status_code == 200
    assert "Engineering" in agg_res.json()["labels"]


@pytest.mark.asyncio
async def test_data_preview_from_link_with_sheet_range_and_request(client: AsyncClient, monkeypatch):
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "linked_data@corp.com",
        "password": "Password123!",
        "name": "Linked Data"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    csv_content = b"Nhan vien,Phong ban,Luong\nA,Ke toan,100\nB,Kinh doanh,200\nC,Nhan su,300\n"
    data_url = "https://example.com/bang-luong.csv"

    async def fake_load(url: str, sheet_range: str = None):
        assert url == data_url
        return csv_content, "bang-luong.csv", "text/csv"

    monkeypatch.setattr("app.api.v1.data.url_dataset_loader.load", fake_load)
    preview_res = await client.post("/api/v1/data/preview-upload", data={
        "data_source_url": data_url,
        "sheet_range": "A1:C3",
        "analysis_request": "Phân tích lương theo phòng ban",
    }, headers=headers)

    assert preview_res.status_code == 200
    payload = preview_res.json()
    assert payload["source_mode"] == "url"
    assert payload["sheet_range"] == "A1:C3"
    assert payload["analysis_request"] == "Phân tích lương theo phòng ban"
    assert payload["total_rows"] == 2
    assert payload["total_columns"] == 3
    assert payload["sheets"][0]["records"][1]["Luong"] == 200


@pytest.mark.asyncio
async def test_dataset_upload_deduplicates_exact_same_file(client: AsyncClient):
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "dedupe_data@corp.com",
        "password": "Password123!",
        "name": "Dedupe Data"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    proj_res = await client.post("/api/v1/projects", json={
        "name": "Không gian dữ liệu",
        "type": "data_analysis"
    }, headers=headers)
    project_id = proj_res.json()["id"]

    csv_content = b"Nhan vien,Phong ban,Luong co ban\nA,Ke toan,100\nB,Kinh doanh,200\n"
    upload_payload = {"project_id": project_id, "document_type": "dataset"}
    files = {"file": ("bang_luong.csv", csv_content, "text/csv")}
    first = await client.post("/api/v1/files/upload", data=upload_payload, files=files, headers=headers)

    files = {"file": ("bang_luong_copy.csv", csv_content, "text/csv")}
    second = await client.post("/api/v1/files/upload", data=upload_payload, files=files, headers=headers)
    listed = await client.get("/api/v1/files", headers=headers)

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json()["id"] == first.json()["id"]
    assert len([item for item in listed.json() if item["file_type"] == "excel"]) == 1
    comparison = second.json()["metadata_json"]["dataset_comparison"]
    assert comparison["last_duplicate_upload_ignored"] == "bang_luong_copy.csv"


@pytest.mark.asyncio
async def test_dataset_upload_groups_highly_similar_file(client: AsyncClient):
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "similar_data@corp.com",
        "password": "Password123!",
        "name": "Similar Data"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    proj_res = await client.post("/api/v1/projects", json={
        "name": "Không gian dữ liệu",
        "type": "data_analysis"
    }, headers=headers)
    project_id = proj_res.json()["id"]
    upload_payload = {"project_id": project_id, "document_type": "dataset"}

    first_csv = b"Nhan vien,Phong ban,Luong co ban\nA,Ke toan,100\nB,Kinh doanh,200\n"
    second_csv = b"Nhan vien,Phong ban,Luong co ban\nA,Ke toan,100\nB,Kinh doanh,201\n"
    first = await client.post(
        "/api/v1/files/upload",
        data=upload_payload,
        files={"file": ("bang_luong.csv", first_csv, "text/csv")},
        headers=headers,
    )
    second = await client.post(
        "/api/v1/files/upload",
        data=upload_payload,
        files={"file": ("bang_luong_update.csv", second_csv, "text/csv")},
        headers=headers,
    )
    listed = await client.get("/api/v1/files", headers=headers)

    assert first.status_code == 200
    assert second.status_code == 200
    assert second.json()["id"] != first.json()["id"]
    assert len([item for item in listed.json() if item["file_type"] == "excel"]) == 2
    first_group = first.json()["metadata_json"]["dataset_comparison"]["dataset_group_id"]
    second_comparison = second.json()["metadata_json"]["dataset_comparison"]
    assert second_comparison["comparison_status"] == "similar"
    assert second_comparison["primary_file_id"] == first.json()["id"]
    assert second_comparison["dataset_group_id"] == first_group


def test_google_sheets_url_normalizer_and_range():
    from app.services.data.url_dataset_loader import UrlDatasetLoader

    # Standard Google Sheet URL (prioritizes full XLSX workbook export for visual preview)
    url1 = "https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms/edit?usp=sharing"
    norm1 = UrlDatasetLoader.normalize_url(url1)
    assert "export?format=xlsx" in norm1
    assert "1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms" in norm1

    # With gid in fragment
    url2 = "https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms/edit#gid=987654321"
    candidates2 = UrlDatasetLoader.get_google_sheet_candidates(url2)
    assert any("gid=987654321" in c for c in candidates2)

    # With sheet_range (sheet name)
    url3 = "https://docs.google.com/spreadsheets/d/1BxiMVs0XRA5nFMdKvBdBZjgmUUqptlbs74OgvE2upms/edit"
    norm3 = UrlDatasetLoader.normalize_url(url3, sheet_range="Class Data")
    assert "gviz/tq?tqx=out:csv&sheet=Class%20Data" in norm3

    # With sheet_range (sheet name + range)
    norm4 = UrlDatasetLoader.normalize_url(url3, sheet_range="DS1!A1:H50")
    assert "gviz/tq?tqx=out:csv&sheet=DS1&range=A1%3AH50" in norm4


def test_url_dataset_loader_ssrf_blocking():
    from app.services.data.url_dataset_loader import UrlDatasetLoader

    with pytest.raises(ValueError, match="Chỉ hỗ trợ liên kết http/https công khai"):
        UrlDatasetLoader._validate_public_host("file:///etc/passwd")

    with pytest.raises(ValueError, match="máy chủ nội bộ"):
        UrlDatasetLoader._validate_public_host("http://localhost:8080/data.csv")

    with pytest.raises(ValueError, match="máy chủ nội bộ"):
        UrlDatasetLoader._validate_public_host("http://127.0.0.1:8000/data.csv")

    with pytest.raises(ValueError, match="mạng nội bộ/private"):
        UrlDatasetLoader._validate_public_host("http://192.168.1.100/data.csv")

    with pytest.raises(ValueError, match="mạng nội bộ/private"):
        UrlDatasetLoader._validate_public_host("http://10.0.0.1/data.csv")


@pytest.mark.asyncio
async def test_url_dataset_loader_rejects_data_urls_in_public_flow():
    from app.services.data.url_dataset_loader import UrlDatasetLoader

    with pytest.raises(ValueError, match="http/https công khai"):
        await UrlDatasetLoader.load("data:text/csv,Nhan%20vien,Luong")


@pytest.mark.asyncio
async def test_preview_upload_requires_either_file_or_url(client: AsyncClient):
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "validate_data@corp.com",
        "password": "Password123!",
        "name": "Validate Data"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Neither file nor url provided
    res = await client.post("/api/v1/data/preview-upload", data={}, headers=headers)
    assert res.status_code == 400
    assert "Vui lòng tải tệp dữ liệu từ máy hoặc dán link dữ liệu" in res.json()["detail"]


@pytest.mark.asyncio
async def test_preview_upload_with_xlsx_file(client: AsyncClient):
    import io
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "xlsx_data@corp.com",
        "password": "Password123!",
        "name": "Xlsx Data"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Create dummy Excel in memory
    df = pd.DataFrame({
        "San pham": ["Laptop", "Dien thoai", "May tinh bang"],
        "So luong": [10, 25, 15],
        "Don gia": [15000000, 8000000, 6000000]
    })
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    files = {"file": ("bao_cao_ban_hang.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    res = await client.post("/api/v1/data/preview-upload", files=files, headers=headers)
    assert res.status_code == 200
    data = res.json()
    assert data["ok"] is True
    assert data["source_mode"] == "file"
    assert data["total_rows"] == 3
    assert data["total_columns"] == 3
    assert len(data["columns"]) == 3
    assert data["columns"][0]["name"] == "San pham"
    assert "visual_workbook" in data
    assert len(data["visual_workbook"]["sheets"]) >= 1


def test_spreadsheet_visual_engine_full_fidelity():
    import io
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from app.services.data.spreadsheet_visual_engine import spreadsheet_visual_engine

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đơn Giá"

    # Merged Title across B1:I1
    ws.merge_cells("B1:I1")
    title_cell = ws["B1"]
    title_cell.value = "BẢNG GIÁ THUÊ THẦU TẠI MIỀN NAM"
    title_cell.font = Font(name="Arial", size=14, bold=True, color="000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
    ws.row_dimensions[1].height = 35

    # Headers in row 2 with bilingual text
    ws["A2"].value = "STT"
    ws["B2"].value = "始发地点\nNơi xuất phát"
    ws["B2"].font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    ws["B2"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data row 3 with formatted currency and empty cell
    ws["A3"].value = 1
    ws["B3"].value = "胡志明 Hồ Chí Minh"
    ws["C3"].value = None  # empty cell
    ws["D3"].value = 8500000
    ws["D3"].number_format = "#,##0"

    # Yellow highlighted cell in row 4
    ws["A4"].value = 2
    ws["B4"].value = "多乐 Đắk Lắk"
    ws["D4"].value = 10000000
    ws["D4"].number_format = "#,##0"
    ws["D4"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    tmp_file = "/tmp/test_full_fidelity.xlsx"
    with open(tmp_file, "wb") as f:
        f.write(buf.getvalue())

    visual = spreadsheet_visual_engine.extract_visual_workbook(tmp_file)
    assert visual["sheet_count"] == 1
    sheet = visual["sheets"][0]
    assert sheet["name"] == "Đơn Giá"
    assert len(sheet["merged_cells"]) == 1
    assert sheet["merged_cells"][0]["range"] == "B1:I1"
    assert sheet["merged_cells"][0]["col_span"] == 8

    # Title master cell B1
    b1 = sheet["cells"][0][1]
    assert b1["display_value"] == "BẢNG GIÁ THUÊ THẦU TẠI MIỀN NAM"
    assert b1["col_span"] == 8
    assert b1["font"]["bold"] is True
    assert b1["fill"]["color"] == "#E2EFDA"

    # Slave cell C1
    c1 = sheet["cells"][0][2]
    assert c1["is_merged_slave"] is True

    # Bilingual header B2
    b2 = sheet["cells"][1][1]
    assert "始发地点" in b2["display_value"]
    assert "Nơi xuất phát" in b2["display_value"]
    assert b2["font"]["color"] == "#FFFFFF"
    assert b2["fill"]["color"] == "#1F4E78"
    assert b2["alignment"]["wrap_text"] is True

    # Formatted price D3
    d3 = sheet["cells"][2][3]
    assert d3["display_value"] == "8,500,000"

    # Blank cell C3
    c3 = sheet["cells"][2][2]
    assert c3["display_value"] == ""

    # Yellow cell D4
    d4 = sheet["cells"][3][3]
    assert d4["display_value"] == "10,000,000"
    assert d4["fill"]["color"] == "#FFF2CC"


@pytest.mark.asyncio
async def test_sheet_analysis_service_multi_sheet_and_stats():
    import io
    import openpyxl
    from app.services.data.sheet_analysis_service import sheet_analysis_service

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Đơn Giá"
    # Row 1: Merged Title
    ws1.merge_cells("B1:E1")
    ws1["B1"].value = "BẢNG GIÁ THUÊ THẦU TẠI MIỀN NAM"
    # Row 2: Headers
    ws1["A2"].value = "STT"
    ws1["B2"].value = "Nơi xuất phát"
    ws1["C2"].value = "Nơi đến"
    ws1["D2"].value = "Giá (VND)"
    ws1["E2"].value = "Hình Thức"

    routes = [
        (1, "Hồ Chí Minh", "Cần Thơ", 8500000, "Cố định"),
        (2, "Hồ Chí Minh", "Đắk Lắk", 10000000, "Cố định"),
        (3, "Hồ Chí Minh", "Nha Trang", 15000000, "Event"),
        (4, "Bình Dương", "Cần Thơ", 9000000, "Cố định"),
        (5, "Đồng Nai", "Vũng Tàu", 3500000, "Khác"),
    ]
    for r_idx, (stt, start, end, price, form) in enumerate(routes, start=3):
        ws1[f"A{r_idx}"].value = stt
        ws1[f"B{r_idx}"].value = start
        ws1[f"C{r_idx}"].value = end
        ws1[f"D{r_idx}"].value = price
        ws1[f"E{r_idx}"].value = form

    # Sheet 2: "HCM T8"
    ws2 = wb.create_sheet(title="HCM T8")
    ws2["A1"].value = "Mã xe"
    ws2["B1"].value = "Tài xế"
    ws2["C1"].value = "Doanh thu"
    ws2["A2"].value = "51A-12345"
    ws2["B2"].value = "Nguyễn Văn A"
    ws2["C2"].value = 25000000
    ws2["A3"].value = "51B-67890"
    ws2["B3"].value = "Trần Văn B"
    ws2["C3"].value = 30000000

    tmp_file = "/tmp/test_sheet_analysis.xlsx"
    wb.save(tmp_file)

    # 1. Test Sheet 1 Analysis
    res1 = await sheet_analysis_service.analyze_sheet(tmp_file, sheet_name="Đơn Giá")
    assert res1["sheet_name"] == "Đơn Giá"
    assert res1["all_sheets"] == ["Đơn Giá", "HCM T8"]
    assert res1["overview"]["total_rows"] == 5
    assert res1["overview"]["total_columns"] == 5
    assert res1["overview"]["numeric_columns_count"] >= 1

    col_names = [c["name"] for c in res1["columns"]]
    assert "Nơi xuất phát" in col_names
    assert "Giá (VND)" in col_names
    
    # Check Price column statistics
    price_col = next(c for c in res1["columns"] if c["name"] == "Giá (VND)")
    assert price_col["type"] == "currency"
    assert price_col["min"] == 3500000.0
    assert price_col["max"] == 15000000.0
    assert price_col["sum"] == 46000000.0

    # Check charts
    assert len(res1["charts"]) >= 1
    assert any(c["type"] == "bar" for c in res1["charts"])

    # 2. Test Sheet 2 Analysis
    res2 = await sheet_analysis_service.analyze_sheet(tmp_file, sheet_name="HCM T8")
    assert res2["sheet_name"] == "HCM T8"
    assert res2["overview"]["total_rows"] == 2
    assert [c["name"] for c in res2["columns"]] == ["Mã xe", "Tài xế", "Doanh thu"]


@pytest.mark.asyncio
async def test_api_analyze_sheet_endpoint(client):
    import io
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "sheet_api_test@corp.com",
        "password": "Password123!",
        "name": "Sheet Api Test"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Prepare XLSX with 2 sheets
    df1 = pd.DataFrame({"Khu vuc": ["Mien Nam", "Mien Bac"], "Doanh thu": [50000000, 45000000]})
    df2 = pd.DataFrame({"Thang": ["T8", "T9"], "Chi phi": [12000000, 14000000]})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df1.to_excel(writer, sheet_name="DoanhThu", index=False)
        df2.to_excel(writer, sheet_name="ChiPhi", index=False)
    buf.seek(0)

    # Analyze sheet DoanhThu
    files = {"file": ("test_two_sheets.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {"sheet_name": "DoanhThu"}
    res = await client.post("/api/v1/data/analyze-sheet", files=files, data=data, headers=headers)
    assert res.status_code == 200
    res_data = res.json()
    assert res_data["ok"] is True
    analysis = res_data["analysis"]
    assert analysis["sheet_name"] == "DoanhThu"
    assert analysis["overview"]["total_rows"] == 2
    assert len(analysis["columns"]) == 2
    assert "ai_insights" in analysis


@pytest.mark.asyncio
async def test_docx_report_workflow_preservation(client):
    """
    Verifies that the existing automated report generation workflow remains 100% functional.
    """
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "docx_preservation@corp.com",
        "password": "Password123!",
        "name": "Docx Preservation"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Ensure /reports list works
    res_reports = await client.get("/api/v1/reports", headers=headers)
    assert res_reports.status_code == 200


def test_spreadsheet_query_engine_range_parsing():
    from app.services.data.spreadsheet_query_engine import parse_excel_range, extract_excel_ranges_from_text

    # 1. Standard range
    p1 = parse_excel_range("H6:H137")
    assert p1["valid"] is True
    assert p1["start_col"] == "H"
    assert p1["start_col_idx"] == 8
    assert p1["start_row"] == 6
    assert p1["end_col"] == "H"
    assert p1["end_col_idx"] == 8
    assert p1["end_row"] == 137
    assert p1["row_count"] == 132

    # 2. Multi-column range
    p2 = parse_excel_range("A1:C20")
    assert p2["valid"] is True
    assert p2["start_col"] == "A"
    assert p2["start_col_idx"] == 1
    assert p2["end_col"] == "C"
    assert p2["end_col_idx"] == 3
    assert p2["row_count"] == 20
    assert p2["col_count"] == 3

    # 3. Sheet with quotes
    p3 = parse_excel_range("'HN Chính T8'!H6:I137")
    assert p3["valid"] is True
    assert p3["sheet_name"] == "HN Chính T8"
    assert p3["start_col"] == "H"
    assert p3["end_col"] == "I"

    # 4. Extract ranges from natural language
    text = "Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không"
    extracted = extract_excel_ranges_from_text(text)
    assert "H6:H137" in extracted
    assert "I6:I137" in extracted


def test_spreadsheet_query_engine_duplicate_detection_and_blanks(tmp_path):
    from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HN Chính T8"

    # Set up data in H and I columns
    # Row 6: H6 = "VGL 102", I6 = "VGL 102" (Intersection match)
    ws["H6"] = " VGL 102 "
    ws["I6"] = "vgl 102"

    # Row 10: H10 = "29A-123.45", I10 = "Other"
    ws["H10"] = "29A-123.45"
    ws["I10"] = "Other"

    # Row 20: I20 = "29A-123.45" (Intersection match across rows)
    ws["I20"] = "29A-123.45"

    # Row 30: Blank cells (must be ignored)
    ws["H30"] = ""
    ws["I30"] = None
    ws["H31"] = "   "

    file_path = str(tmp_path / "test_duplicates.xlsx")
    wb.save(file_path)
    wb.close()

    # Find duplicates intersection between H6:H35 and I6:I35
    res = spreadsheet_query_engine.find_duplicates(
        file_path=file_path,
        sheet_name="HN Chính T8",
        ranges=["H6:H35", "I6:I35"],
        normalize=True,
        ignore_blank=True,
    )

    assert res["duplicate_count"] == 2  # "VGL 102" and "29A-123.45"
    matched_addrs = [c["address"] for c in res["matched_cells"]]
    assert "H6" in matched_addrs
    assert "I6" in matched_addrs
    assert "H10" in matched_addrs
    assert "I20" in matched_addrs
    # Blanks must NOT be in matched cells
    assert "H30" not in matched_addrs
    assert "I30" not in matched_addrs
    assert "H31" not in matched_addrs

    # Test apply highlights
    out_path = str(tmp_path / "highlighted_out.xlsx")
    spreadsheet_query_engine.apply_highlights_to_workbook(
        file_path=file_path,
        sheet_name="HN Chính T8",
        cell_addresses=["H6", "I6"],
        color_hex="FFFF00",
        output_path=out_path,
    )
    assert Path(out_path).exists()


@pytest.mark.asyncio
async def test_workbook_chat_service_duplicate_query_and_actions(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HN Chính T8"
    ws["H6"] = "ABC"
    ws["I15"] = "abc"
    file_path = str(tmp_path / "chat_test.xlsx")
    wb.save(file_path)
    wb.close()

    # 1. Main user prompt
    prompt = "Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không và bôi vàng luôn."
    res = await workbook_chat_service.chat(
        file_path=file_path,
        message=prompt,
        sheet_name="HN Chính T8",
        conversation_id="conv_test_123",
    )

    assert "HN Chính T8" in res["context"]["sheet"]
    assert res["result"]["duplicate_count"] == 1
    assert len(res["actions"]) == 1
    action = res["actions"][0]
    assert action["type"] == "HIGHLIGHT_CELLS"
    assert "H6" in action["cells"]
    assert "I15" in action["cells"]
    assert action["autoScrollTo"] == "H6"

    # 2. Multi-turn follow-up test: "Xóa màu"
    res_clear = await workbook_chat_service.chat(
        file_path=file_path,
        message="Xóa màu đánh dấu",
        sheet_name="HN Chính T8",
        conversation_id="conv_test_123",
    )
    assert len(res_clear["actions"]) == 1
    assert res_clear["actions"][0]["type"] == "CLEAR_HIGHLIGHTS"


@pytest.mark.asyncio
async def test_api_workbook_chat_and_apply_modifications(client: AsyncClient, tmp_path):
    import io
    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "chat_api_user@corp.com",
        "password": "Password123!",
        "name": "Chat API User"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    # Prepare XLSX file
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HN Chính T8"
    ws["H6"] = "BKS-01"
    ws["I10"] = "bks-01"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 1. Test POST /workbook-chat
    files = {"file": ("google_sheet.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    data = {
        "sheet_name": "HN Chính T8",
        "message": "Kiểm tra H6:H20 và I6:I20 xem có trùng không và bôi vàng",
    }
    chat_res = await client.post("/api/v1/data/workbook-chat", files=files, data=data, headers=headers)
    assert chat_res.status_code == 200
    chat_data = chat_res.json()
    assert chat_data["ok"] is True
    assert chat_data["result"]["duplicate_count"] == 1
    assert len(chat_data["actions"]) >= 1

    # 2. Test POST /apply-modifications
    buf.seek(0)
    files2 = {"file": ("google_sheet.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    mod_data = {
        "sheet_name": "HN Chính T8",
        "cells": '["H6", "I10"]',
        "color_hex": "FFFF00",
    }
    mod_res = await client.post("/api/v1/data/apply-modifications", files=files2, data=mod_data, headers=headers)
    assert mod_res.status_code == 200
    mod_json = mod_res.json()
    assert mod_json["ok"] is True
    assert mod_json["highlighted_count"] == 2
    assert "download_url" in mod_json

    download_res = await client.get(mod_json["download_url"], headers=headers)
    assert download_res.status_code == 200
    highlighted_wb = openpyxl.load_workbook(io.BytesIO(download_res.content))
    try:
        highlighted_ws = highlighted_wb["HN Chính T8"]
        assert highlighted_ws["H6"].fill.fill_type == "solid"
        assert highlighted_ws["I10"].fill.fill_type == "solid"
        assert highlighted_ws["H6"].fill.start_color.rgb == "FFFFFF00"
        assert highlighted_ws["I10"].fill.start_color.rgb == "FFFFFF00"
    finally:
        highlighted_wb.close()


@pytest.mark.asyncio
async def test_api_workbook_analysis_action_endpoint(client: AsyncClient):
    import io
    import openpyxl

    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "analysis_action_user@corp.com",
        "password": "Password123!",
        "name": "Analysis Action User"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bang_luong"
    ws.append(["Ma NV", "Thuc linh", "Ghi chu"])
    ws.append(["NV001", 12000000, ""])
    ws.append(["NV002", 9500000, None])
    ws.append(["NV003", 18000000, "ok"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    res = await client.post(
        "/api/v1/data/workbook-analysis-action",
        files={"file": ("bang_luong.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={
            "sheet_name": "Bang_luong",
            "prompt": "Tìm ô trống",
            "selected_range": "C2:C4",
            "scope": json.dumps({"type": "range", "sheet": "Bang_luong", "range": "C2:C4"}),
        },
        headers=headers,
    )

    assert res.status_code == 200
    payload = res.json()
    assert payload["ok"] is True
    assert payload["mode"] == "analysis_action"
    assert payload["context"]["ranges"] == ["C2:C4"]
    assert payload["result"]["missing_count"] == 2
    assert payload["evidence"]["operation"] == "FIND_MISSING"
    assert payload["actions"][0]["type"] == "HIGHLIGHT_CELLS"


@pytest.mark.asyncio
async def test_api_workbook_chat_intents_and_conversational_flows(client: AsyncClient):
    import io
    import openpyxl

    reg_res = await client.post("/api/v1/auth/register", json={
        "email": "chat_e2e_user@corp.com",
        "password": "Password123!",
        "name": "Chat E2E User"
    })
    token = reg_res.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Bang_luong"
    ws1.append(["Ho ten", "Ma NV", "Thuc linh"])
    ws1.append(["Nguyễn Văn A", "NV001", 12000000])
    ws1.append(["Trần Thị B", "NV002", 9500000])
    ws1.append(["Lê Văn C", "NV003", 28000000])

    ws2 = wb.create_sheet("Tong_hop")
    ws2.append(["Khoản mục", "Số tiền"])
    ws2.append(["Chi phí", 50000000])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_bytes = buf.getvalue()

    # 1. Test Greeting ("Xin chào")
    res_greet = await client.post(
        "/api/v1/data/workbook-chat",
        files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Bang_luong", "message": "Xin chào"},
        headers=headers,
    )
    assert res_greet.status_code == 200
    greet_json = res_greet.json()
    assert greet_json["ok"] is True
    assert greet_json["intent"] == "GREETING"
    assert "Xin chào" in greet_json["answer"]

    # 2. Test Help ("Bạn làm được gì?")
    res_help = await client.post(
        "/api/v1/data/workbook-chat",
        files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Bang_luong", "message": "Bạn làm được gì?"},
        headers=headers,
    )
    assert res_help.status_code == 200
    help_json = res_help.json()
    assert help_json["ok"] is True
    assert help_json["intent"] == "HELP"
    assert "Tra cứu" in help_json["answer"]

    # 3. Test Metadata ("File này có bao nhiêu sheet?")
    res_sheets = await client.post(
        "/api/v1/data/workbook-chat",
        files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Bang_luong", "message": "File này có bao nhiêu sheet?"},
        headers=headers,
    )
    assert res_sheets.status_code == 200
    sheets_json = res_sheets.json()
    assert sheets_json["ok"] is True
    assert "2 sheet" in sheets_json["answer"]
    assert "Bang_luong" in sheets_json["answer"]
    assert "Tong_hop" in sheets_json["answer"]

    # 4. Test Data Query ("Ai có Thực lĩnh cao nhất?")
    res_top = await client.post(
        "/api/v1/data/workbook-chat",
        files={"file": ("test.xlsx", file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"sheet_name": "Bang_luong", "message": "Ai có Thực lĩnh cao nhất?"},
        headers=headers,
    )
    assert res_top.status_code == 200
    top_json = res_top.json()
    assert top_json["ok"] is True
    assert "Lê Văn C" in top_json["answer"] or "NV003" in top_json["answer"]
    assert "28.000.000" in top_json["answer"]



def test_semantic_type_inference_upgrades():
    from app.services.data.sheet_analysis_service import sheet_analysis_service

    # Tải Trọng with 1T9, 5T -> category or numeric, NOT date
    s_payload = pd.Series(["1T9", "5T", "8T", "15T", "1T9"])
    t_payload = sheet_analysis_service.infer_column_type(s_payload, "Tải Trọng")
    assert t_payload in ["category", "text", "numeric"]
    assert t_payload != "date"

    # Transport / Flight Schedule (STA, STD) -> date/time, NOT currency
    s_sta = pd.Series(["08:30", "14:00", "18:45"])
    t_sta = sheet_analysis_service.infer_column_type(s_sta, "STA")
    assert t_sta in ["date", "time"]
    assert t_sta != "currency"

    # CCCD / Biển số -> id_code
    s_id = pd.Series(["001200001234", "001200005678"])
    t_id = sheet_analysis_service.infer_column_type(s_id, "Số CCCD")
    assert t_id == "id_code"


def test_content_aware_column_width_extraction(tmp_path):
    from app.services.data.spreadsheet_visual_engine import spreadsheet_visual_engine
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dslx"

    # Header with long text
    ws["A1"] = "BIỂN KIỂM SOÁT"
    ws["B1"] = "TRỌNG TẢI"
    ws["C1"] = "DỰ ÁN"
    ws["D1"] = "TUYẾN"
    ws["E1"] = "SỐ ĐIỆN THOẠI"

    ws["A2"] = "29C-888.99"
    ws["B2"] = "5T"
    ws["C2"] = "Dự án Vận tải Miền Bắc"
    ws["D2"] = "Hà Nội - Hải Phòng"
    ws["E2"] = "0987654321"

    file_path = str(tmp_path / "test_dslx.xlsx")
    wb.save(file_path)
    wb.close()

    result = spreadsheet_visual_engine.extract_visual_workbook(file_path)
    assert result["sheet_count"] == 1
    sheet = result["sheets"][0]
    assert sheet["name"] == "dslx"

    # Column widths must be wide enough to prevent character-by-character wrap
    col_widths = sheet["column_widths"]
    assert col_widths["1"] >= 100  # BIỂN KIỂM SOÁT
    assert col_widths["2"] >= 85   # TRỌNG TẢI
    assert col_widths["3"] >= 100  # DỰ ÁN
    assert col_widths["4"] >= 100  # TUYẾN
    assert col_widths["5"] >= 100  # SỐ ĐIỆN THOẠI


def test_voice_transcript_range_extraction_normalization():
    from app.services.data.spreadsheet_query_engine import extract_excel_ranges_from_text

    # 1. Standard speech transcript with spaces: "H 6 đến H 137"
    t1 = "Kiểm tra từ dòng H 6 đến H 137 và I 6 đến I 137 xem có bị trùng lặp không"
    r1 = extract_excel_ranges_from_text(t1)
    assert "H6:H137" in r1
    assert "I6:I137" in r1

    # 2. Speech with "tới" / "sang"
    t2 = "Xem từ H6 tới H137 và bôi vàng"
    r2 = extract_excel_ranges_from_text(t2)
    assert "H6:H137" in r2

    # 3. Shorthand "H6 đến 137"
    t3 = "So sánh H6 đến 137 với I6 đến 137"
    r3 = extract_excel_ranges_from_text(t3)
    assert "H6:H137" in r3
    assert "I6:I137" in r3


@pytest.mark.asyncio
async def test_workbook_chat_voice_prompt_and_dslx_sheet(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "dslx"
    ws["H6"] = "29C-111.22"
    ws["I10"] = "29C-111.22"

    file_path = str(tmp_path / "voice_chat_test.xlsx")
    wb.save(file_path)
    wb.close()

    # User speaks: "Kiểm tra H 6 đến H 137 và I 6 đến I 137 xem có bị trùng không và bôi vàng"
    prompt = "Kiểm tra H 6 đến H 137 và I 6 đến I 137 xem có bị trùng không và bôi vàng"
    res = await workbook_chat_service.chat(
        file_path=file_path,
        message=prompt,
        sheet_name="dslx",
        conversation_id="voice_session_dslx",
    )

    assert res["context"]["sheet"] == "dslx"
    assert res["result"]["duplicate_count"] == 1
    assert len(res["actions"]) >= 1
    assert res["actions"][0]["type"] == "HIGHLIGHT_CELLS"
    assert "H6" in res["actions"][0]["cells"]
    assert "I10" in res["actions"][0]["cells"]


def test_sheet_resolution_exact_and_diacritic_rules():
    from app.services.data.spreadsheet_query_engine import resolve_sheet_name_in_wb

    available = ["dslx", "HN Chinh T8", "HN nhánh T8", "HCM t8"]

    # 1. "HN Chính T8" must resolve to "HN Chinh T8" (without diacritics)
    assert resolve_sheet_name_in_wb("HN Chính T8", available) == "HN Chinh T8"
    assert resolve_sheet_name_in_wb("hn chính t8", available) == "HN Chinh T8"
    assert resolve_sheet_name_in_wb("HN CHINH T8", available) == "HN Chinh T8"
    assert resolve_sheet_name_in_wb("HN Chinh T8", available) == "HN Chinh T8"

    # 2. "HN nhánh T8" must resolve to "HN nhánh T8" (NEVER confused with HN Chính T8)
    assert resolve_sheet_name_in_wb("HN nhánh T8", available) == "HN nhánh T8"
    assert resolve_sheet_name_in_wb("hn nhanh t8", available) == "HN nhánh T8"

    # 3. "HCM t8"
    assert resolve_sheet_name_in_wb("HCM t8", available) == "HCM t8"
    assert resolve_sheet_name_in_wb("hcm t8", available) == "HCM t8"

    # 4. "dslx"
    assert resolve_sheet_name_in_wb("dslx", available) == "dslx"


def test_range_parser_does_not_parse_t8_as_range():
    from app.services.data.spreadsheet_query_engine import extract_excel_ranges_from_text

    # Case 1: Exact user prompt
    prompt1 = "Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không (HN Chính T8)"
    r1 = extract_excel_ranges_from_text(prompt1)
    assert r1 == ["H6:H137", "I6:I137"]
    assert "T8" not in r1

    # Case 2: Parentheses with spaces
    prompt2 = "Kiểm tra H6:H137 với I6:I137 ở ( HN Chính T8 )"
    r2 = extract_excel_ranges_from_text(prompt2)
    assert r2 == ["H6:H137", "I6:I137"]
    assert "T8" not in r2

    # Case 3: "ở sheet HN Chính T8"
    prompt3 = "Kiểm tra H6:H137 với I6:I137 ở sheet HN Chính T8"
    r3 = extract_excel_ranges_from_text(prompt3)
    assert r3 == ["H6:H137", "I6:I137"]
    assert "T8" not in r3


def test_mock_data_3_way_duplicates_algorithm(tmp_path):
    """
    Mock worksheet:
    H6 = ABC, H7 = DEF, H8 = ABC
    I6 = XYZ, I7 = ABC, I8 = QQQ
    Expected:
    duplicates_in_first_range (in H): ABC -> H6, H8
    duplicates_in_second_range (in I): none
    cross_range_duplicates (intersection): ABC -> H6, H8, I7
    """
    from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HN Chinh T8"

    ws["H6"] = "ABC"
    ws["H7"] = "DEF"
    ws["H8"] = "ABC"

    ws["I6"] = "XYZ"
    ws["I7"] = "ABC"
    ws["I8"] = "QQQ"

    file_path = str(tmp_path / "mock_dup.xlsx")
    wb.save(file_path)
    wb.close()

    result = spreadsheet_query_engine.find_duplicates(
        file_path=file_path,
        sheet_name="HN Chính T8",
        ranges=["H6:H8", "I6:I8"],
        normalize=True,
        ignore_blank=True,
    )

    assert result["duplicate_count"] >= 1
    assert result["cross_range_count"] == 1
    assert result["cross_range_duplicates"][0]["value"] == "ABC"
    assert "H6" in result["cross_range_duplicates"][0]["first_range_cells"]
    assert "H8" in result["cross_range_duplicates"][0]["first_range_cells"]
    assert "I7" in result["cross_range_duplicates"][0]["second_range_cells"]

    # In first range H: ABC is duplicate
    assert result["within_first_range_count"] == 1
    assert result["duplicates_in_first_range"][0]["value"] == "ABC"

    # In second range I: No duplicate
    assert result["within_second_range_count"] == 0

    # Matched cells contains H6, H8, I7
    matched_addrs = [c["address"] for c in result["matched_cells"]]
    assert "H6" in matched_addrs
    assert "H8" in matched_addrs
    assert "I7" in matched_addrs


def test_row_column_offset_raw_cell_access(tmp_path):
    """
    Ensures raw Excel coordinates are queried directly without DataFrame header offset.
    Row 1: Title
    Row 2: Blank
    Row 3: Header
    Row 4..10: Data
    """
    from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine, parse_excel_range
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "BÁO CÁO DOANH THU"
    ws["A3"] = "MÃ ĐƠN"
    ws["B3"] = "KHÁCH HÀNG"

    ws["B4"] = "KH_001"
    ws["B5"] = "KH_002"
    ws["B6"] = "KH_001"

    file_path = str(tmp_path / "offset_test.xlsx")
    wb.save(file_path)
    wb.close()

    range_spec = parse_excel_range("B4:B6")
    cells = spreadsheet_query_engine.read_range_cells(file_path, "Sheet1", range_spec)

    assert len(cells) == 3
    assert cells[0]["address"] == "B4"
    assert cells[0]["value"] == "KH_001"
    assert cells[1]["address"] == "B5"
    assert cells[1]["value"] == "KH_002"
    assert cells[2]["address"] == "B6"
    assert cells[2]["value"] == "KH_001"


@pytest.mark.asyncio
async def test_full_workbook_chat_resolves_hn_chinh_t8_and_finds_duplicates(tmp_path):
    """
    Full End-to-End Chat test:
    User asks:
    'Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không (HN Chính T8)'
    Workbook contains sheets: ['dslx', 'HN Chinh T8', 'HN nhánh T8']
    """
    from app.services.data.workbook_chat_service import workbook_chat_service
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "dslx"
    # Put no duplicates on dslx
    ws1["H6"] = "VAL_1"
    ws1["I6"] = "VAL_2"

    ws2 = wb.create_sheet(title="HN Chinh T8")
    # Put real duplicates on HN Chinh T8
    ws2["H11"] = "BKS-29C-999.88"
    ws2["H30"] = "BKS-29C-999.88"
    ws2["I28"] = "BKS-29C-999.88"
    ws2["H50"] = "29H-123.45"
    ws2["I50"] = "29H-123.45"

    ws3 = wb.create_sheet(title="HN nhánh T8")
    ws3["H6"] = "BRANCH_1"

    file_path = str(tmp_path / "multi_sheet_chat.xlsx")
    wb.save(file_path)
    wb.close()

    user_query = "Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không (HN Chính T8)"
    res = await workbook_chat_service.chat(
        file_path=file_path,
        message=user_query,
        sheet_name="dslx",  # Active sheet in UI was dslx, but user specified (HN Chính T8)
        conversation_id="e2e_hn_chinh_t8_test",
    )

    # 1. Target sheet must be resolved as 'HN Chinh T8'
    assert res["context"]["sheet"] == "HN Chinh T8"

    # 2. Ranges must be exactly H6:H137 and I6:I137 without 'T8'
    assert res["context"]["ranges"] == ["H6:H137", "I6:I137"]

    # 3. Duplicate count must be > 0 (found BKS-29C-999.88 and 29H-123.45)
    assert res["result"]["duplicate_count"] >= 2
    assert res["result"]["cross_range_count"] >= 2

    # 4. Answer must contain matched values and details
    assert "HN Chinh T8" in res["answer"]
    assert "29C-999.88" in res["answer"] or "123.45" in res["answer"]
    assert "Không phát hiện giá trị nào bị trùng lặp" not in res["answer"]

    # 5. Matched cells must include H11, H30, I28, H50, I50
    matched_addrs = [c["address"] for c in res["result"]["matched_cells"]]
    assert "H11" in matched_addrs
    assert "H30" in matched_addrs
    assert "I28" in matched_addrs
    assert "H50" in matched_addrs
    assert "I50" in matched_addrs


@pytest.mark.asyncio
async def test_workbook_chat_resolves_sheet_without_parentheses_and_matches_numeric_floats(tmp_path):
    from app.services.data.workbook_chat_service import workbook_chat_service
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "dslx"

    ws2 = wb.create_sheet(title="HN Chinh T8")
    # Test float 123.0 vs int 123
    ws2["H10"] = 123.0
    ws2["I10"] = 123
    ws2["H20"] = "  MÃ_A  "
    ws2["I25"] = "mã_a"

    file_path = str(tmp_path / "float_test.xlsx")
    wb.save(file_path)
    wb.close()

    user_query = "Kiểm tra H6:H30 và I6:I30 trong HN Chính T8 xem có trùng lặp không"
    res = await workbook_chat_service.chat(
        file_path=file_path,
        message=user_query,
        sheet_name="dslx",
        conversation_id="e2e_no_paren_test",
    )

    assert res["context"]["sheet"] == "HN Chinh T8"
    assert res["context"]["ranges"] == ["H6:H30", "I6:I30"]
    assert res["result"]["cross_range_count"] >= 2
    matched_addrs = [c["address"] for c in res["result"]["matched_cells"]]
    assert "H10" in matched_addrs
    assert "I10" in matched_addrs
    assert "H20" in matched_addrs
    assert "I25" in matched_addrs
