import hashlib
import json
import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
import numpy as np


class DataAnalysisEngine:
    """
    Python-powered Data Analysis & Statistical Aggregation Engine.
    Enforces strict mathematical accuracy (AI never invents totals or averages).
    """

    FACT_LIMIT = 220
    RECORD_PREVIEW_LIMIT = 100
    FORMULA_LIMIT = 150

    @classmethod
    def _json_safe(cls, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (np.integer,)):
            return int(value)
        if isinstance(value, (np.floating,)):
            if math.isnan(float(value)) or math.isinf(float(value)):
                return None
            return round(float(value), 4)
        if isinstance(value, float):
            if math.isnan(value) or math.isinf(value):
                return None
            return round(value, 4)
        if isinstance(value, (pd.Timestamp,)):
            return value.isoformat()
        if pd.isna(value):
            return None
        return value

    @classmethod
    def _safe_records(cls, df: pd.DataFrame, limit: int = RECORD_PREVIEW_LIMIT) -> List[Dict[str, Any]]:
        records = df.head(limit).replace({np.nan: None}).to_dict(orient="records")
        return [
            {str(k): cls._json_safe(v) for k, v in row.items()}
            for row in records
        ]

    @classmethod
    def _cell_range(cls, sheet_name: str, col_idx: int, row_count: int, header_row: int = 1) -> str:
        try:
            from openpyxl.utils import get_column_letter
            col = get_column_letter(col_idx)
        except Exception:
            col = str(col_idx)
        start = header_row + 1
        end = header_row + max(row_count, 1)
        return f"{sheet_name}!{col}{start}:{col}{end}"

    @classmethod
    def _source(cls, file_path: str, sheet_name: str, cell_range: str) -> Dict[str, str]:
        return {"file": Path(file_path).name, "sheet": sheet_name, "range": cell_range}

    @classmethod
    def parse_sheet_range(cls, sheet_range: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
        value = (sheet_range or "").strip()
        if not value:
            return None, None
        if "!" in value:
            sheet_name, range_ref = value.split("!", 1)
            return sheet_name.strip().strip("'\"") or None, range_ref.strip().upper() or None
        if re.fullmatch(r"[A-Za-z]{1,3}\d+:[A-Za-z]{1,3}\d+", value):
            return None, value.upper()
        return value.strip().strip("'\""), None

    @classmethod
    def _apply_a1_range(cls, df: pd.DataFrame, range_ref: Optional[str]) -> pd.DataFrame:
        if not range_ref:
            return df
        from openpyxl.utils.cell import column_index_from_string

        match = re.fullmatch(r"([A-Z]{1,3})(\d+):([A-Z]{1,3})(\d+)", range_ref.strip().upper())
        if not match:
            raise ValueError(f"Vùng dữ liệu không hợp lệ: {range_ref}")
        start_col, start_row, end_col, end_row = match.groups()
        start_col_idx = column_index_from_string(start_col) - 1
        end_col_idx = column_index_from_string(end_col)
        start_row_idx = max(int(start_row) - 1, 0)
        end_row_idx = int(end_row)

        raw = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx].copy()
        if raw.empty:
            raise ValueError(f"Không tìm thấy dữ liệu trong vùng {range_ref}")
        header = [str(value).strip() if str(value).strip() else f"Column_{idx + 1}" for idx, value in enumerate(raw.iloc[0].tolist())]
        body = raw.iloc[1:].reset_index(drop=True)
        body.columns = header
        for column in body.columns:
            converted = pd.to_numeric(body[column], errors="coerce")
            original_non_null = body[column].notna()
            if original_non_null.any() and converted[original_non_null].notna().all():
                body[column] = converted
        return body

    @classmethod
    def _detect_column_type(cls, series: pd.Series) -> str:
        if pd.api.types.is_numeric_dtype(series):
            return "numeric"
        if pd.api.types.is_datetime64_any_dtype(series):
            return "datetime"
        non_null = series.dropna()
        if len(non_null) > 0:
            sample_text = " ".join(str(x) for x in non_null.head(20).tolist())
            if not re.search(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}", sample_text):
                return "text"
            parsed_dates = pd.to_datetime(non_null.head(50), errors="coerce")
            if parsed_dates.notna().mean() >= 0.8:
                return "datetime"
        if pd.api.types.is_bool_dtype(series):
            return "boolean"
        return "text"

    @classmethod
    def _infer_fact_type(cls, column_name: str, operation: str) -> str:
        text = f"{column_name} {operation}".lower()
        if any(k in text for k in ["thực lĩnh", "thu nhap", "thu nhập", "lương", "luong", "salary", "income"]):
            base = "salary"
        elif any(k in text for k in ["thuế", "tax", "tncn"]):
            base = "tax"
        elif any(k in text for k in ["bhxh", "bhyt", "bhtn", "bảo hiểm", "bao hiem", "insurance"]):
            base = "insurance"
        elif any(k in text for k in ["ngày công", "ngay cong", "working"]):
            base = "working_days"
        elif any(k in text for k in ["thưởng", "thuong", "bonus"]):
            base = "bonus"
        elif any(k in text for k in ["phụ cấp", "phu cap", "allowance"]):
            base = "allowance"
        elif any(k in text for k in ["khấu trừ", "khau tru", "deduction"]):
            base = "deduction"
        elif any(k in text for k in ["phòng", "phong", "department"]):
            base = "department"
        elif any(k in text for k in ["chức", "chuc", "role", "position"]):
            base = "role"
        elif any(k in text for k in ["nhân viên", "nhan vien", "employee"]):
            base = "employee"
        else:
            base = "general"

        if "tổng" in text or "sum" in text:
            suffix = "total"
        elif "trung bình" in text or "average" in text or "mean" in text:
            suffix = "avg"
        elif "thấp nhất" in text or "min" in text:
            suffix = "min"
        elif "cao nhất" in text or "max" in text:
            suffix = "max"
        elif "grouped" in text or "theo" in text:
            suffix = "grouped"
        elif "category" in text:
            suffix = "category"
        else:
            suffix = "value"
        return f"{base}_{suffix}"

    @classmethod
    def load_dataframe(cls, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        ext = Path(file_path).suffix.lower()
        if ext in [".xlsx", ".xls", ".xlsm"]:
            return pd.read_excel(file_path, sheet_name=sheet_name or 0)
        elif ext == ".csv":
            return pd.read_csv(file_path)
        else:
            raise ValueError(f"Unsupported tabular format: {ext}")

    @classmethod
    def _extract_formulas(cls, file_path: str) -> Dict[str, Dict[str, str]]:
        if Path(file_path).suffix.lower() not in [".xlsx", ".xlsm"]:
            return {}
        formulas: Dict[str, Dict[str, str]] = {}
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, data_only=False, read_only=False)
            for sheet in workbook.worksheets:
                sheet_formulas: Dict[str, str] = {}
                for row in sheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            sheet_formulas[cell.coordinate] = value
                            if len(sheet_formulas) >= cls.FORMULA_LIMIT:
                                break
                    if len(sheet_formulas) >= cls.FORMULA_LIMIT:
                        break
                if sheet_formulas:
                    formulas[sheet.title] = sheet_formulas
        except Exception:
            return {}
        return formulas

    @classmethod
    def _profile_sheet(cls, df: pd.DataFrame, sheet_name: str, file_path: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        total_rows, total_cols = df.shape
        missing_count = int(df.isnull().sum().sum())
        duplicate_rows = int(df.duplicated().sum())

        columns_profile: List[Dict[str, Any]] = []
        verified_facts: List[Dict[str, Any]] = [
            {
                "name": "row_count",
                "fact": f"Số dòng dữ liệu của sheet {sheet_name}",
                "value": int(total_rows),
                "source": cls._source(file_path, sheet_name, "used_range"),
                "fact_type": "row_count",
                "type": "FACT",
            },
            {
                "name": "column_count",
                "fact": f"Số cột dữ liệu của sheet {sheet_name}",
                "value": int(total_cols),
                "source": cls._source(file_path, sheet_name, "used_range"),
                "fact_type": "column_count",
                "type": "FACT",
            },
        ]
        numeric_columns: List[str] = []
        categorical_columns: List[str] = []
        datetime_columns: List[str] = []

        for idx, col in enumerate(df.columns, 1):
            series = df[col]
            col_name = str(col)
            col_type = cls._detect_column_type(series)
            is_numeric = col_type == "numeric"
            null_pct = round(float(series.isnull().mean() * 100), 1)
            source_range = cls._cell_range(sheet_name, idx, total_rows)

            col_info: Dict[str, Any] = {
                "name": col_name,
                "type": col_type,
                "source_range": source_range,
                "non_null_count": int(series.notna().sum()),
                "null_count": int(series.isna().sum()),
                "null_percentage": null_pct,
                "unique_count": int(series.nunique()),
                "sample_values": [str(x) for x in series.dropna().head(5).tolist()],
            }

            if is_numeric:
                numeric_columns.append(col_name)
                clean_s = pd.to_numeric(series, errors="coerce").dropna()
                if len(clean_s) > 0:
                    stats = {
                        "count": int(clean_s.count()),
                        "sum": cls._json_safe(clean_s.sum()),
                        "min": cls._json_safe(clean_s.min()),
                        "max": cls._json_safe(clean_s.max()),
                        "average": cls._json_safe(clean_s.mean()),
                        "median": cls._json_safe(clean_s.median()),
                    }
                    col_info.update(stats)
                    for label, value in [
                        ("Tổng", stats["sum"]),
                        ("Thấp nhất", stats["min"]),
                        ("Cao nhất", stats["max"]),
                        ("Trung bình", stats["average"]),
                        ("Trung vị", stats["median"]),
                    ]:
                        verified_facts.append({
                            "name": f"{label.lower().replace(' ', '_')}_{col_name}",
                            "fact": f"{label} của cột {col_name}",
                            "value": value,
                            "source": cls._source(file_path, sheet_name, source_range),
                            "fact_type": cls._infer_fact_type(col_name, label),
                            "type": "DERIVED_FACT",
                        })
            elif col_type == "datetime":
                datetime_columns.append(col_name)
                parsed = pd.to_datetime(series, errors="coerce").dropna()
                if len(parsed) > 0:
                    col_info["min_date"] = parsed.min().date().isoformat()
                    col_info["max_date"] = parsed.max().date().isoformat()
                    verified_facts.append({
                        "name": f"date_range_{col_name}",
                        "fact": f"Khoảng thời gian của cột {col_name}",
                        "value": f"{col_info['min_date']} đến {col_info['max_date']}",
                        "source": cls._source(file_path, sheet_name, source_range),
                        "fact_type": "date",
                        "type": "DERIVED_FACT",
                    })
            else:
                categorical_columns.append(col_name)
                top_values = series.dropna().astype(str).value_counts().head(10)
                col_info["top_values"] = [
                    {"value": str(key), "count": int(value)}
                    for key, value in top_values.items()
                ]
                if 0 < int(series.nunique()) <= 50:
                    verified_facts.append({
                        "name": f"categories_{col_name}",
                        "fact": f"Giá trị phân loại trong cột {col_name}",
                        "value": [str(x) for x in top_values.index.tolist()],
                        "source": cls._source(file_path, sheet_name, source_range),
                        "fact_type": cls._infer_fact_type(col_name, "category"),
                        "type": "FACT",
                    })

            columns_profile.append(col_info)

        categorical_summary: Dict[str, Any] = {}
        grouped_statistics: List[Dict[str, Any]] = []
        for cat_col in categorical_columns[:4]:
            if df[cat_col].nunique(dropna=True) > 30:
                continue
            categorical_summary[cat_col] = columns_profile[[c["name"] for c in columns_profile].index(cat_col)].get("top_values", [])
            for metric_col in numeric_columns[:6]:
                grouped = (
                    df.assign(**{metric_col: pd.to_numeric(df[metric_col], errors="coerce")})
                    .groupby(cat_col, dropna=True)[metric_col]
                    .agg(["count", "sum", "mean", "min", "max"])
                    .reset_index()
                    .sort_values("sum", ascending=False)
                    .head(12)
                )
                records = [
                    {
                        "group": str(row[cat_col]),
                        "count": int(row["count"]),
                        "sum": cls._json_safe(row["sum"]),
                        "average": cls._json_safe(row["mean"]),
                        "min": cls._json_safe(row["min"]),
                        "max": cls._json_safe(row["max"]),
                    }
                    for _, row in grouped.iterrows()
                ]
                grouped_statistics.append({
                    "group_by": cat_col,
                    "metric": metric_col,
                    "statistics": records,
                })
                for rec in records[:6]:
                    verified_facts.append({
                        "name": f"grouped_{metric_col}_by_{cat_col}_{rec['group']}",
                        "fact": f"{metric_col} theo {cat_col} = {rec['group']}",
                        "value": {"sum": rec["sum"], "average": rec["average"], "count": rec["count"]},
                        "source": cls._source(file_path, sheet_name, f"{cat_col}+{metric_col}"),
                        "fact_type": f"{cls._infer_fact_type(metric_col, 'grouped')}_by_{cls._infer_fact_type(cat_col, 'category')}",
                        "type": "DERIVED_FACT",
                    })

        warnings = []
        if missing_count:
            warnings.append(f"Sheet {sheet_name} có {missing_count} ô trống.")
        if duplicate_rows:
            warnings.append(f"Sheet {sheet_name} có {duplicate_rows} dòng trùng lặp.")

        sheet_profile = {
            "name": sheet_name,
            "row_count": int(total_rows),
            "column_count": int(total_cols),
            "columns": columns_profile,
            "records": cls._safe_records(df),
            "statistics": {
                "missing_values_count": missing_count,
                "duplicate_rows_count": duplicate_rows,
                "numeric_columns": numeric_columns,
                "categorical_columns": categorical_columns,
                "datetime_columns": datetime_columns,
            },
            "categorical_summary": categorical_summary,
            "grouped_statistics": grouped_statistics[:24],
            "warnings": warnings,
        }
        return sheet_profile, verified_facts

    @classmethod
    def profile_dataset(cls, file_path: str, sheet_range: Optional[str] = None) -> Dict[str, Any]:
        path = Path(file_path)
        ext = path.suffix.lower()
        formulas = cls._extract_formulas(file_path)
        sheets: List[Dict[str, Any]] = []
        verified_facts: List[Dict[str, Any]] = []
        selected_sheet_name, selected_range = cls.parse_sheet_range(sheet_range)

        if ext in [".xlsx", ".xls", ".xlsm"]:
            excel = pd.ExcelFile(file_path)
            sheet_names = [selected_sheet_name] if selected_sheet_name else excel.sheet_names
            for sheet_name in sheet_names:
                if sheet_name not in excel.sheet_names:
                    raise ValueError(f"Không tìm thấy sheet '{sheet_name}'. Các sheet hiện có: {', '.join(excel.sheet_names)}")
                if selected_range:
                    raw_df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    df = cls._apply_a1_range(raw_df, selected_range)
                else:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_profile, facts = cls._profile_sheet(df, sheet_name, file_path)
                if formulas.get(sheet_name):
                    sheet_profile["formulas"] = formulas[sheet_name]
                    for cell, formula in list(formulas[sheet_name].items())[:30]:
                        facts.append({
                            "name": f"formula_{sheet_name}_{cell}",
                            "fact": f"Công thức tại {sheet_name}!{cell}",
                            "value": formula,
                            "source": cls._source(file_path, sheet_name, cell),
                            "fact_type": "formula",
                            "type": "FACT",
                        })
                else:
                    sheet_profile["formulas"] = {}
                sheets.append(sheet_profile)
                verified_facts.extend(facts)
        else:
            if selected_range:
                df = pd.read_csv(file_path, header=None)
                df = cls._apply_a1_range(df, selected_range)
            else:
                df = cls.load_dataframe(file_path)
            sheet_profile, facts = cls._profile_sheet(df, "CSV", file_path)
            sheet_profile["formulas"] = {}
            sheets.append(sheet_profile)
            verified_facts.extend(facts)

        for idx, fact in enumerate(verified_facts[: cls.FACT_LIMIT], 1):
            fact["id"] = f"FACT_{idx:03d}"

        warnings = [warning for sheet in sheets for warning in sheet.get("warnings", [])]
        total_rows = sum(sheet.get("row_count", 0) for sheet in sheets)
        total_columns = max([sheet.get("column_count", 0) for sheet in sheets] or [0])
        primary_sheet = sheets[0] if sheets else {}
        missing_count = sum((sheet.get("statistics") or {}).get("missing_values_count", 0) for sheet in sheets)
        duplicate_rows = sum((sheet.get("statistics") or {}).get("duplicate_rows_count", 0) for sheet in sheets)

        return {
            "file_name": path.name,
            "source_type": "excel" if ext in [".xlsx", ".xls", ".xlsm"] else "csv",
            "sheet_count": len(sheets),
            "total_rows": int(total_rows),
            "total_columns": int(total_columns),
            "sheets": sheets,
            "missing_values_count": int(missing_count),
            "duplicate_rows_count": int(duplicate_rows),
            "columns": primary_sheet.get("columns", []),
            "preview_rows": primary_sheet.get("records", []),
            "verified_facts": verified_facts[: cls.FACT_LIMIT],
            "warnings": warnings,
            "grounding_rules": cls.grounding_rules(),
            "selection": {
                "sheet_range": sheet_range or "",
                "sheet_name": selected_sheet_name or "",
                "range": selected_range or "",
            },
        }

    @classmethod
    def grounding_rules(cls) -> List[str]:
        return [
            "Excel/CSV Dataset Profile và VERIFIED_FACTS là nguồn sự thật duy nhất cho mọi số liệu.",
            "Không tự tạo, đoán, suy luận hoặc thay đổi KPI, tên người, phòng ban, chức vụ, ngày, tiền, tỷ lệ hoặc tên sheet.",
            "Nếu dữ liệu không tồn tại trong profile, phải ghi rõ: Dữ liệu nguồn không cung cấp thông tin này.",
            "Word template chỉ dùng để lấy cấu trúc, định dạng, bố cục, bảng mẫu và vị trí ảnh; không dùng số liệu hoặc nội dung mẫu làm dữ liệu thật.",
            "Nếu template Word và Excel/CSV mâu thuẫn, Excel/CSV luôn thắng.",
        ]

    @classmethod
    def infer_report_title(cls, profile: Dict[str, Any]) -> str:
        text_parts = [str(profile.get("file_name") or "")]
        for sheet in profile.get("sheets", [])[:3]:
            text_parts.append(str(sheet.get("name") or ""))
            text_parts.extend(str(col.get("name") or "") for col in sheet.get("columns", [])[:40])
        text = " ".join(text_parts).lower()

        has_salary = any(k in text for k in ["lương", "luong", "salary", "thực lĩnh", "thu nhập", "income"])
        has_people = any(k in text for k in ["nhân viên", "nhan vien", "employee", "phòng ban", "phong ban", "department"])
        if has_salary and has_people:
            return "Báo cáo phân tích dữ liệu bảng lương nhân viên"
        if has_salary:
            return "Báo cáo phân tích dữ liệu lương"
        if any(k in text for k in ["doanh thu", "revenue", "chi phí", "cost", "profit", "lợi nhuận"]):
            return "Báo cáo phân tích dữ liệu tài chính"
        return "Báo cáo phân tích dữ liệu từ file Excel/CSV"

    @classmethod
    def dataset_schema_signature(cls, profile: Dict[str, Any]) -> str:
        sheets = []
        for sheet in profile.get("sheets", []):
            columns = [
                {
                    "name": str(col.get("name") or "").strip().lower(),
                    "type": str(col.get("type") or "").strip().lower(),
                }
                for col in sheet.get("columns", [])
            ]
            sheets.append({
                "name": str(sheet.get("name") or "").strip().lower(),
                "columns": columns,
            })
        payload = json.dumps(sheets, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @classmethod
    def dataset_row_signature(cls, profile: Dict[str, Any], limit: int = 100) -> str:
        values = []
        for sheet in profile.get("sheets", [])[:5]:
            rows = sheet.get("records", []) or []
            columns = [str(col.get("name") or "") for col in sheet.get("columns", [])]
            for row in rows[:limit]:
                values.append([
                    cls._normalize_compare_value(row.get(column))
                    for column in columns
                ])
        payload = json.dumps(values, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    @classmethod
    def compare_dataset_profiles(cls, first: Dict[str, Any], second: Dict[str, Any]) -> Dict[str, Any]:
        first_schema = cls.dataset_schema_signature(first)
        second_schema = cls.dataset_schema_signature(second)
        schema_match = first_schema == second_schema

        first_rows = cls._profile_compare_values(first)
        second_rows = cls._profile_compare_values(second)
        max_rows = max(len(first_rows), len(second_rows), 1)
        matched_rows = sum(1 for left, right in zip(first_rows, second_rows) if left == right)
        row_score = matched_rows / max_rows
        compared_cells = 0
        matched_cells = 0
        for left, right in zip(first_rows, second_rows):
            width = max(len(left), len(right), 1)
            compared_cells += width
            matched_cells += sum(1 for left_cell, right_cell in zip(left, right) if left_cell == right_cell)
        cell_score = matched_cells / max(compared_cells, 1)

        if schema_match:
            similarity_score = 0.7 + (max(row_score, cell_score) * 0.3)
        else:
            similarity_score = row_score * 0.5

        if schema_match and similarity_score >= 0.99:
            status = "duplicate"
        elif schema_match and similarity_score >= 0.9:
            status = "similar"
        else:
            status = "different"

        return {
            "status": status,
            "similarity_score": round(similarity_score, 4),
            "schema_match": schema_match,
            "schema_signature": first_schema,
            "row_signature": cls.dataset_row_signature(first),
            "matched_rows": matched_rows,
            "compared_rows": max_rows,
        }

    @classmethod
    def _profile_compare_rows(cls, profile: Dict[str, Any], limit: int = 100) -> List[str]:
        return [
            json.dumps(row, ensure_ascii=False, sort_keys=True)
            for row in cls._profile_compare_values(profile, limit)
        ]

    @classmethod
    def _profile_compare_values(cls, profile: Dict[str, Any], limit: int = 100) -> List[List[str]]:
        rows = []
        for sheet in profile.get("sheets", [])[:5]:
            columns = [str(col.get("name") or "") for col in sheet.get("columns", [])]
            for row in (sheet.get("records", []) or [])[:limit]:
                normalized = [cls._normalize_compare_value(row.get(column)) for column in columns]
                rows.append(normalized)
        return rows

    @staticmethod
    def _normalize_compare_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            return str(round(value, 6))
        return re.sub(r"\s+", " ", str(value).strip().lower())

    @classmethod
    def format_profile_for_prompt(cls, profile: Dict[str, Any], max_chars: int = 18000) -> str:
        lines = [
            "CRITICAL DATA GROUNDING RULES",
            *[f"- {rule}" for rule in profile.get("grounding_rules", cls.grounding_rules())],
            "",
            f"DATASET: {profile.get('file_name')} ({profile.get('source_type')})",
            f"Số sheet: {profile.get('sheet_count')} | Tổng dòng: {profile.get('total_rows')}",
            "",
            "VERIFIED_FACTS:",
        ]
        for fact in profile.get("verified_facts", [])[:120]:
            source = fact.get("source") or {}
            if isinstance(source, dict):
                source_text = f"{source.get('file', '')}/{source.get('sheet', '')}!{source.get('range', '')}"
            else:
                source_text = str(source)
            lines.append(f"- {fact.get('id')}: {fact.get('name') or fact.get('fact')} = {fact.get('value')} [source: {source_text}]")
        for sheet in profile.get("sheets", [])[:6]:
            lines.extend([
                "",
                f"SHEET {sheet.get('name')}: {sheet.get('row_count')} dòng, {sheet.get('column_count')} cột",
                "Cột dữ liệu:",
            ])
            for col in sheet.get("columns", [])[:30]:
                details = [col.get("type"), f"unique={col.get('unique_count')}", f"missing={col.get('null_count')}"]
                if col.get("type") == "numeric":
                    details.extend([
                        f"sum={col.get('sum')}",
                        f"min={col.get('min')}",
                        f"max={col.get('max')}",
                        f"avg={col.get('average')}",
                        f"median={col.get('median')}",
                    ])
                if col.get("top_values"):
                    top = ", ".join(f"{x['value']}({x['count']})" for x in col["top_values"][:5])
                    details.append(f"top={top}")
                lines.append(f"- {col.get('name')}: " + "; ".join(str(x) for x in details if x is not None))
            for grouped in sheet.get("grouped_statistics", [])[:8]:
                label = grouped.get("group_by")
                metric = grouped.get("metric")
                rows = grouped.get("statistics", [])[:5]
                summary = "; ".join(f"{r['group']}: sum={r['sum']}, avg={r['average']}, count={r['count']}" for r in rows)
                lines.append(f"Nhóm {metric} theo {label}: {summary}")
            formulas = sheet.get("formulas", {})
            if formulas:
                lines.append("Công thức Excel phát hiện:")
                for cell, formula in list(formulas.items())[:20]:
                    lines.append(f"- {sheet.get('name')}!{cell}: {formula}")
            if sheet.get("warnings"):
                lines.append("Cảnh báo dữ liệu: " + "; ".join(sheet["warnings"]))
        text = "\n".join(lines)
        if len(text) > max_chars:
            return text[:max_chars] + "\n...[Dataset profile đã rút gọn để vừa ngữ cảnh]..."
        return text

    @classmethod
    def aggregate_data(
        cls,
        file_path: str,
        group_by: Optional[str] = None,
        metric_column: Optional[str] = None,
        aggregation: str = "sum",  # sum, mean, count, min, max
        top_n: int = 10
    ) -> Dict[str, Any]:
        df = cls.load_dataframe(file_path)

        if not group_by or not metric_column:
            # Overall metric
            if metric_column and metric_column in df.columns:
                s = pd.to_numeric(df[metric_column], errors="coerce").dropna()
                val = float(s.sum() if aggregation == "sum" else s.mean() if aggregation == "mean" else len(s))
                return {"metric": metric_column, "aggregation": aggregation, "value": round(val, 2)}
            return {"error": "Invalid columns"}

        # Perform pandas groupby
        df[metric_column] = pd.to_numeric(df[metric_column], errors="coerce")
        grouped = df.groupby(group_by)[metric_column].agg(aggregation).reset_index()
        grouped = grouped.sort_values(by=metric_column, ascending=False).head(top_n)

        labels = [str(x) for x in grouped[group_by].tolist()]
        values = [round(float(x), 2) if not pd.isna(x) else 0.0 for x in grouped[metric_column].tolist()]

        return {
            "group_by": group_by,
            "metric": metric_column,
            "aggregation": aggregation,
            "labels": labels,
            "values": values,
            "table_data": grouped.replace({np.nan: None}).to_dict(orient="records"),
        }

    @classmethod
    def build_chart_specification(
        cls,
        file_path: str,
        chart_type: str,  # bar, line, pie, donut, horizontal_bar, area
        group_by: str,
        metric_column: str,
        aggregation: str = "sum",
        title: Optional[str] = None
    ) -> Dict[str, Any]:
        agg_result = cls.aggregate_data(file_path, group_by=group_by, metric_column=metric_column, aggregation=aggregation)

        chart_spec = {
            "chart_type": chart_type,
            "title": title or f"{aggregation.upper()} of {metric_column} by {group_by}",
            "labels": agg_result.get("labels", []),
            "datasets": [
                {
                    "name": metric_column,
                    "data": agg_result.get("values", []),
                }
            ],
            "metadata": {
                "source_file": Path(file_path).name,
                "group_by": group_by,
                "metric": metric_column,
                "aggregation": aggregation,
            }
        }
        return chart_spec


data_engine = DataAnalysisEngine()
