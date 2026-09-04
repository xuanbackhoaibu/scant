import csv
import datetime
import math
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class SpreadsheetVisualEngine:
    """
    High-fidelity Spreadsheet Visual Engine.
    Extracts complete visual layout (merged cells, column widths, row heights,
    cell background colors, borders, font styles, text alignments, number formats,
    multilingual Unicode content) from Excel workbooks (.xlsx, .xlsm, .xls) and CSV files.
    """

    DEFAULT_COL_WIDTH_PX = 85
    DEFAULT_ROW_HEIGHT_PX = 26
    MAX_PREVIEW_ROWS = 500
    MAX_PREVIEW_COLS = 100

    THEME_COLORS = {
        0: "#FFFFFF",  # Light 1
        1: "#000000",  # Dark 1
        2: "#EEECE1",  # Light 2
        3: "#1F497D",  # Dark 2
        4: "#4F81BD",  # Accent 1
        5: "#C0504D",  # Accent 2
        6: "#9BBB59",  # Accent 3
        7: "#8064A2",  # Accent 4
        8: "#4BACC6",  # Accent 5
        9: "#F79646",  # Accent 6
    }

    @classmethod
    def _extract_color(cls, color_obj: Any) -> Optional[str]:
        if not color_obj:
            return None
        try:
            color_type = getattr(color_obj, "type", None)
            if color_type == "rgb":
                rgb = getattr(color_obj, "rgb", None)
                if isinstance(rgb, str) and rgb:
                    # Clean ARGB / RGB
                    clean_hex = rgb.strip()
                    if len(clean_hex) == 8:
                        # Format: AARRGGBB -> take last 6 chars for #RRGGBB
                        return f"#{clean_hex[2:]}"
                    elif len(clean_hex) == 6:
                        return f"#{clean_hex}"
            elif color_type == "theme":
                theme_idx = getattr(color_obj, "theme", None)
                if theme_idx is not None and isinstance(theme_idx, int):
                    return cls.THEME_COLORS.get(theme_idx, None)
            elif color_type == "indexed":
                # Indexed colors (fallback to None or standard index)
                return None
        except Exception:
            return None
        return None

    @classmethod
    def _extract_border_side(cls, side_obj: Any) -> Optional[Dict[str, Any]]:
        if not side_obj or not getattr(side_obj, "style", None):
            return None
        style = getattr(side_obj, "style", "thin")
        color = cls._extract_color(getattr(side_obj, "color", None)) or "#000000"
        return {
            "style": style,
            "color": color,
        }

    @classmethod
    def _format_cell_display_value(cls, val: Any, num_fmt: Optional[str], data_type: Optional[str]) -> str:
        if val is None:
            return ""
        if isinstance(val, bool):
            return "TRUE" if val else "FALSE"
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val.strftime("%Y-%m-%d")
        if isinstance(val, (int, float)):
            fmt = (num_fmt or "").lower()
            if "%" in fmt:
                return f"{val * 100:.1f}%"
            if "#,##0" in fmt or "0,000" in fmt:
                if ".00" in fmt:
                    return f"{val:,.2f}"
                return f"{val:,.0f}"
            if isinstance(val, float):
                if math.isnan(val) or math.isinf(val):
                    return ""
                if val.is_integer():
                    return f"{int(val)}"
                return f"{val:.4g}"
            return str(val)
        return str(val)

    @classmethod
    def _extract_xlsx_sheet(
        cls,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        ws_data_only: Optional[openpyxl.worksheet.worksheet.Worksheet] = None,
        max_rows: int = MAX_PREVIEW_ROWS,
        max_cols: int = MAX_PREVIEW_COLS,
    ) -> Dict[str, Any]:
        sheet_name = ws.title
        max_r = min(max(ws.max_row or 1, 1), max_rows)
        max_c = min(max(ws.max_column or 1, 1), max_cols)

        # 1. Process merged cells
        merged_cells_list: List[Dict[str, Any]] = []
        merged_cell_map: Dict[Tuple[int, int], Dict[str, Any]] = {}

        for rng in ws.merged_cells.ranges:
            min_col, min_row, max_c_rng, max_r_rng = rng.min_col, rng.min_row, rng.max_col, rng.max_row
            row_span = max_r_rng - min_row + 1
            col_span = max_c_rng - min_col + 1
            range_str = str(rng)

            merged_cells_list.append({
                "range": range_str,
                "start_row": min_row,
                "start_col": min_col,
                "end_row": max_r_rng,
                "end_col": max_c_rng,
                "row_span": row_span,
                "col_span": col_span,
            })

            for r in range(min_row, max_r_rng + 1):
                for c in range(min_col, max_c_rng + 1):
                    is_master = (r == min_row and c == min_col)
                    merged_cell_map[(r, c)] = {
                        "is_master": is_master,
                        "master_coord": f"{get_column_letter(min_col)}{min_row}",
                        "row_span": row_span if is_master else 1,
                        "col_span": col_span if is_master else 1,
                        "range": range_str,
                    }

        # 2. Process column widths with content-aware minimum fallback
        col_widths_px: Dict[str, int] = {}
        for c_idx in range(1, max_c + 1):
            col_letter = get_column_letter(c_idx)
            
            # Sample first 40 rows to estimate content length
            max_line_len = 0
            for r in range(1, min(max_r + 1, 40)):
                val_raw = ws.cell(row=r, column=c_idx).value
                if val_raw is not None:
                    val_str = str(val_raw).strip()
                    for line in val_str.split("\n"):
                        max_line_len = max(max_line_len, len(line.strip()))
            
            content_est_px = max(round(max_line_len * 8.5 + 24), 85)

            dim = ws.column_dimensions.get(col_letter)
            if dim and dim.width is not None:
                try:
                    w = float(dim.width)
                    if w > 0:
                        # Convert Excel character width to pixels (1 char ~ 7.5-8px + padding)
                        px = max(round(w * 8.0 + 14), 55)
                        col_widths_px[str(c_idx)] = px
                    else:
                        col_widths_px[str(c_idx)] = min(content_est_px, 350)
                except Exception:
                    col_widths_px[str(c_idx)] = min(content_est_px, 350)
            else:
                col_widths_px[str(c_idx)] = min(content_est_px, 350)

        # 3. Process row heights
        row_heights_px: Dict[str, int] = {}
        for r_idx in range(1, max_r + 1):
            dim = ws.row_dimensions.get(r_idx)
            if dim and dim.height is not None:
                try:
                    h = float(dim.height)
                    # Convert Excel point height to pixels (1 pt ~ 1.333 px)
                    px = max(round(h * 1.333), 22)
                    row_heights_px[str(r_idx)] = px
                except Exception:
                    row_heights_px[str(r_idx)] = 28 if r_idx == 1 else cls.DEFAULT_ROW_HEIGHT_PX
            else:
                row_heights_px[str(r_idx)] = 28 if r_idx == 1 else cls.DEFAULT_ROW_HEIGHT_PX

        # 4. Process all cells in grid
        cells_matrix: List[List[Dict[str, Any]]] = []

        for r in range(1, max_r + 1):
            row_cells: List[Dict[str, Any]] = []
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                cached_cell = ws_data_only.cell(row=r, column=c) if ws_data_only else None

                raw_val = cell.value
                cached_val = cached_cell.value if cached_cell else raw_val
                
                # Check formula
                formula = None
                if isinstance(raw_val, str) and raw_val.startswith("="):
                    formula = raw_val
                    val_for_display = cached_val
                else:
                    val_for_display = raw_val

                display_val = cls._format_cell_display_value(val_for_display, cell.number_format, cell.data_type)

                # Merge info
                m_info = merged_cell_map.get((r, c))
                is_merged = bool(m_info)
                is_merged_slave = bool(m_info and not m_info["is_master"])
                row_span = m_info["row_span"] if m_info else 1
                col_span = m_info["col_span"] if m_info else 1
                merged_range = m_info["range"] if m_info else None

                # Font
                font_info = None
                if cell.font:
                    font_color = cls._extract_color(cell.font.color)
                    font_info = {
                        "name": cell.font.name or "Arial",
                        "size": int(cell.font.size) if cell.font.size else 11,
                        "bold": bool(cell.font.bold),
                        "italic": bool(cell.font.italic),
                        "underline": bool(cell.font.underline),
                        "color": font_color,
                    }

                # Fill
                fill_info = None
                if cell.fill and cell.fill.fill_type:
                    fill_color = cls._extract_color(cell.fill.start_color or cell.fill.fgColor)
                    if fill_color:
                        fill_info = {
                            "type": cell.fill.fill_type,
                            "color": fill_color,
                        }

                # Border
                border_info = None
                if cell.border:
                    top_b = cls._extract_border_side(cell.border.top)
                    bot_b = cls._extract_border_side(cell.border.bottom)
                    left_b = cls._extract_border_side(cell.border.left)
                    right_b = cls._extract_border_side(cell.border.right)
                    if any([top_b, bot_b, left_b, right_b]):
                        border_info = {
                            "top": top_b,
                            "bottom": bot_b,
                            "left": left_b,
                            "right": right_b,
                        }

                # Alignment
                align_info = None
                if cell.alignment:
                    horiz = cell.alignment.horizontal
                    vert = cell.alignment.vertical
                    wrap = bool(cell.alignment.wrap_text)
                    if horiz or vert or wrap:
                        align_info = {
                            "horizontal": horiz,
                            "vertical": vert,
                            "wrap_text": wrap,
                            "text_rotation": getattr(cell.alignment, "text_rotation", 0) or 0,
                        }

                cell_dict: Dict[str, Any] = {
                    "row": r,
                    "col": c,
                    "coordinate": cell.coordinate,
                    "value": str(raw_val) if (raw_val is not None and not isinstance(raw_val, (int, float, bool))) else raw_val,
                    "display_value": display_val,
                    "formula": formula,
                    "number_format": cell.number_format or "General",
                    "font": font_info,
                    "fill": fill_info,
                    "border": border_info,
                    "alignment": align_info,
                    "is_merged": is_merged,
                    "is_merged_slave": is_merged_slave,
                    "row_span": row_span,
                    "col_span": col_span,
                    "merged_range": merged_range,
                }
                row_cells.append(cell_dict)
            cells_matrix.append(row_cells)

        return {
            "name": sheet_name,
            "max_row": max_r,
            "max_column": max_c,
            "merged_cells": merged_cells_list,
            "column_widths": col_widths_px,
            "row_heights": row_heights_px,
            "hidden_rows": [],
            "hidden_columns": [],
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "cells": cells_matrix,
        }

    @classmethod
    def _extract_csv_sheet(cls, file_path: str, max_rows: int = MAX_PREVIEW_ROWS, max_cols: int = MAX_PREVIEW_COLS) -> Dict[str, Any]:
        rows: List[List[str]] = []
        try:
            # Try utf-8-sig first to strip BOM
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for r in reader:
                    rows.append(r)
                    if len(rows) >= max_rows:
                        break
        except UnicodeDecodeError:
            rows = []
            with open(file_path, "r", encoding="latin-1") as f:
                reader = csv.reader(f)
                for r in reader:
                    rows.append(r)
                    if len(rows) >= max_rows:
                        break

        max_r = max(len(rows), 1)
        max_c = min(max([len(r) for r in rows] or [1]), max_cols)

        # Estimate column widths from CSV content
        col_widths: Dict[str, int] = {}
        for c in range(1, max_c + 1):
            max_line_len = 0
            for r in rows[:40]:
                v = r[c - 1] if c - 1 < len(r) else ""
                for line in str(v).split("\n"):
                    max_line_len = max(max_line_len, len(line.strip()))
            col_widths[str(c)] = max(min(round(max_line_len * 8.5 + 24), 320), 85)

        row_heights: Dict[str, int] = {}
        cells_matrix: List[List[Dict[str, Any]]] = []

        for r_idx in range(1, max_r + 1):
            row_heights[str(r_idx)] = 28 if r_idx == 1 else cls.DEFAULT_ROW_HEIGHT_PX
            raw_row = rows[r_idx - 1] if r_idx - 1 < len(rows) else []
            row_cells: List[Dict[str, Any]] = []

            for c_idx in range(1, max_c + 1):
                val = raw_row[c_idx - 1] if c_idx - 1 < len(raw_row) else ""
                col_letter = get_column_letter(c_idx)
                coord = f"{col_letter}{r_idx}"

                is_header = (r_idx == 1)
                has_newline = "\n" in str(val)
                row_cells.append({
                    "row": r_idx,
                    "col": c_idx,
                    "coordinate": coord,
                    "value": val,
                    "display_value": val,
                    "formula": None,
                    "number_format": "General",
                    "font": {
                        "name": "Arial",
                        "size": 11,
                        "bold": is_header,
                        "italic": False,
                        "underline": False,
                        "color": "#000000",
                    },
                    "fill": {
                        "type": "solid",
                        "color": "#F1F5F9" if is_header else None,
                    } if is_header else None,
                    "border": {
                        "top": {"style": "thin", "color": "#CBD5E1"},
                        "bottom": {"style": "thin", "color": "#CBD5E1"},
                        "left": {"style": "thin", "color": "#CBD5E1"},
                        "right": {"style": "thin", "color": "#CBD5E1"},
                    },
                    "alignment": {
                        "horizontal": "center" if is_header else "left",
                        "vertical": "center",
                        "wrap_text": has_newline,
                        "text_rotation": 0,
                    },
                    "is_merged": False,
                    "is_merged_slave": False,
                    "row_span": 1,
                    "col_span": 1,
                    "merged_range": None,
                })
            cells_matrix.append(row_cells)

        return {
            "name": "CSV",
            "max_row": max_r,
            "max_column": max_c,
            "merged_cells": [],
            "column_widths": col_widths,
            "row_heights": row_heights,
            "hidden_rows": [],
            "hidden_columns": [],
            "freeze_panes": None,
            "cells": cells_matrix,
        }

    @classmethod
    def extract_visual_workbook(
        cls,
        file_path: str,
        max_rows: int = MAX_PREVIEW_ROWS,
        max_cols: int = MAX_PREVIEW_COLS,
    ) -> Dict[str, Any]:
        """
        Extract complete visual workbook representation for frontend SpreadsheetPreview.
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext in [".xlsx", ".xlsm"]:
            try:
                # Load with formulas/styles
                wb = openpyxl.load_workbook(file_path, data_only=False, read_only=False)
                # Load data only for cached values
                try:
                    wb_data = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
                except Exception:
                    wb_data = None

                sheets: List[Dict[str, Any]] = []
                for ws in wb.worksheets:
                    ws_data = wb_data[ws.title] if (wb_data and ws.title in wb_data.sheetnames) else None
                    sheet_dict = cls._extract_xlsx_sheet(ws, ws_data_only=ws_data, max_rows=max_rows, max_cols=max_cols)
                    sheets.append(sheet_dict)

                return {
                    "source_type": "excel",
                    "file_name": path.name,
                    "sheet_count": len(sheets),
                    "active_sheet_index": 0,
                    "sheets": sheets,
                }
            except Exception as ex:
                # Fallback to CSV extraction if openpyxl fails
                pass

        if ext == ".csv":
            sheet_dict = cls._extract_csv_sheet(file_path, max_rows=max_rows, max_cols=max_cols)
            return {
                "source_type": "csv",
                "file_name": path.name,
                "sheet_count": 1,
                "active_sheet_index": 0,
                "sheets": [sheet_dict],
            }

        # For .xls or other tabular files, fallback to pandas read and convert to visual grid
        try:
            import pandas as pd
            if ext in [".xls", ".xlsx", ".xlsm"]:
                xls_file = pd.ExcelFile(file_path)
                sheets = []
                for s_name in xls_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=s_name, header=None)
                    # Convert DataFrame to basic visual sheet
                    max_r = min(len(df), max_rows)
                    max_c = min(len(df.columns) if not df.empty else 1, max_cols)
                    cells_matrix = []
                    col_widths = {str(c): cls.DEFAULT_COL_WIDTH_PX for c in range(1, max_c + 1)}
                    row_heights = {str(r): cls.DEFAULT_ROW_HEIGHT_PX for r in range(1, max_r + 1)}

                    for r in range(max_r):
                        row_cells = []
                        for c in range(max_c):
                            val = df.iloc[r, c] if r < len(df) and c < len(df.columns) else None
                            val_str = "" if pd.isna(val) else str(val)
                            coord = f"{get_column_letter(c + 1)}{r + 1}"
                            row_cells.append({
                                "row": r + 1,
                                "col": c + 1,
                                "coordinate": coord,
                                "value": val_str,
                                "display_value": val_str,
                                "formula": None,
                                "number_format": "General",
                                "font": {"name": "Arial", "size": 11, "bold": (r == 0), "italic": False, "color": "#000000"},
                                "fill": {"type": "solid", "color": "#F8FAFC" if r == 0 else None} if r == 0 else None,
                                "border": {"top": {"style": "thin", "color": "#E2E8F0"}, "bottom": {"style": "thin", "color": "#E2E8F0"}, "left": {"style": "thin", "color": "#E2E8F0"}, "right": {"style": "thin", "color": "#E2E8F0"}},
                                "alignment": {"horizontal": "left", "vertical": "center", "wrap_text": True, "text_rotation": 0},
                                "is_merged": False,
                                "is_merged_slave": False,
                                "row_span": 1,
                                "col_span": 1,
                                "merged_range": None,
                            })
                        cells_matrix.append(row_cells)

                    sheets.append({
                        "name": s_name,
                        "max_row": max_r,
                        "max_column": max_c,
                        "merged_cells": [],
                        "column_widths": col_widths,
                        "row_heights": row_heights,
                        "hidden_rows": [],
                        "hidden_columns": [],
                        "freeze_panes": None,
                        "cells": cells_matrix,
                    })

                return {
                    "source_type": "excel",
                    "file_name": path.name,
                    "sheet_count": len(sheets),
                    "active_sheet_index": 0,
                    "sheets": sheets,
                }
        except Exception:
            pass

        # Final fallback
        return {
            "source_type": "unknown",
            "file_name": path.name,
            "sheet_count": 0,
            "active_sheet_index": 0,
            "sheets": [],
        }


spreadsheet_visual_engine = SpreadsheetVisualEngine()
