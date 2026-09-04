import hashlib
import math
import os
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd


def col_letter_to_index(col_letter: str) -> int:
    """Converts Excel column letters (e.g. 'A', 'H', 'AA') to 1-based index (e.g. 1, 8, 27)."""
    col_letter = col_letter.upper().strip()
    idx = 0
    for char in col_letter:
        if 'A' <= char <= 'Z':
            idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx


def col_index_to_letter(col_idx: int) -> str:
    """Converts 1-based column index to Excel column letter."""
    letter = ""
    temp = col_idx
    while temp > 0:
        mod = (temp - 1) % 26
        letter = chr(65 + mod) + letter
        temp = (temp - mod) // 26
    return letter or "A"


def parse_excel_range(range_str: str) -> Dict[str, Any]:
    """
    Parses standard Excel ranges into coordinates:
    e.g. 'H6:H137', 'I6:I137', 'H11', 'A1:C20', 'H:I', '6:137', "'HN Chính T8'!H6:H137"
    """
    clean_str = range_str.strip()
    sheet_name = None

    # Check for sheet prefix, e.g. 'Sheet1'!A1:B10 or Sheet1!A1:B10
    if "!" in clean_str:
        parts = clean_str.split("!", 1)
        sheet_name = parts[0].strip().strip("'").strip('"')
        clean_str = parts[1].strip()

    # Pattern 1: Standard cell or range like H6:H137 or H6 or A1:C20
    match_standard = re.match(r"^([A-Za-z]+)(\d+)(?::([A-Za-z]+)(\d+))?$", clean_str)
    if match_standard:
        start_col = match_standard.group(1).upper()
        start_row = int(match_standard.group(2))
        end_col = (match_standard.group(3) or start_col).upper()
        end_row = int(match_standard.group(4) or start_row)

        start_col_idx = col_letter_to_index(start_col)
        end_col_idx = col_letter_to_index(end_col)

        # Normalize bounding box
        min_col_idx, max_col_idx = min(start_col_idx, end_col_idx), max(start_col_idx, end_col_idx)
        min_row, max_row = min(start_row, end_row), max(start_row, end_row)

        return {
            "valid": True,
            "raw": range_str,
            "sheet_name": sheet_name,
            "start_col": col_index_to_letter(min_col_idx),
            "start_col_idx": min_col_idx,
            "start_row": min_row,
            "end_col": col_index_to_letter(max_col_idx),
            "end_col_idx": max_col_idx,
            "end_row": max_row,
            "is_single_cell": (min_col_idx == max_col_idx and min_row == max_row),
            "col_count": max_col_idx - min_col_idx + 1,
            "row_count": max_row - min_row + 1,
        }

    # Pattern 2: Whole columns e.g. H:I or H:H
    match_cols = re.match(r"^([A-Za-z]+):([A-Za-z]+)$", clean_str)
    if match_cols:
        start_col = match_cols.group(1).upper()
        end_col = match_cols.group(2).upper()
        start_col_idx = col_letter_to_index(start_col)
        end_col_idx = col_letter_to_index(end_col)
        min_col_idx, max_col_idx = min(start_col_idx, end_col_idx), max(start_col_idx, end_col_idx)
        return {
            "valid": True,
            "raw": range_str,
            "sheet_name": sheet_name,
            "start_col": col_index_to_letter(min_col_idx),
            "start_col_idx": min_col_idx,
            "start_row": 1,
            "end_col": col_index_to_letter(max_col_idx),
            "end_col_idx": max_col_idx,
            "end_row": 1000000,
            "is_whole_columns": True,
            "col_count": max_col_idx - min_col_idx + 1,
            "row_count": 1000000,
        }

    # Pattern 3: Whole rows e.g. 6:137
    match_rows = re.match(r"^(\d+):(\d+)$", clean_str)
    if match_rows:
        min_row = min(int(match_rows.group(1)), int(match_rows.group(2)))
        max_row = max(int(match_rows.group(1)), int(match_rows.group(2)))
        return {
            "valid": True,
            "raw": range_str,
            "sheet_name": sheet_name,
            "start_col": "A",
            "start_col_idx": 1,
            "start_row": min_row,
            "end_col": "ZZ",
            "end_col_idx": 702,
            "end_row": max_row,
            "is_whole_rows": True,
            "col_count": 702,
            "row_count": max_row - min_row + 1,
        }

    return {"valid": False, "raw": range_str, "sheet_name": sheet_name}


def remove_diacritics(text: str) -> str:
    """Removes Vietnamese diacritics / accents for robust matching ('HN Chính T8' -> 'hn chinh t8')."""
    if not text:
        return ""
    text = text.replace("đ", "d").replace("Đ", "D")
    nfkd = unicodedata.normalize("NFKD", text)
    cleaned = "".join([c for c in nfkd if not unicodedata.combining(c)])
    return re.sub(r"\s+", " ", cleaned).strip().lower()


def resolve_sheet_name_in_wb(requested_sheet: Optional[str], available_sheets: List[str]) -> str:
    """
    Deterministically resolves a requested sheet name against the workbook's actual sheets.
    Handles:
    1. Exact match ('HN Chinh T8' == 'HN Chinh T8')
    2. Case-insensitive match ('hn chinh t8' == 'HN Chinh T8')
    3. Diacritic-insensitive match ('HN Chính T8' == 'HN Chinh T8')
    4. Token-level subset match
    5. Substring containment match
    Never matches unrelated sheets ('HN nhánh T8' != 'HN Chính T8').
    """
    if not available_sheets:
        return requested_sheet or "Sheet1"

    if not requested_sheet:
        return available_sheets[0]

    req_clean = requested_sheet.strip()

    # 1. Exact match
    if req_clean in available_sheets:
        return req_clean

    # 2. Case-insensitive match
    req_lower = req_clean.lower()
    for s in available_sheets:
        if s.strip().lower() == req_lower:
            return s

    # 3. Diacritics-insensitive match ('HN Chính T8' == 'HN Chinh T8')
    req_no_accent = remove_diacritics(req_clean)
    for s in available_sheets:
        if remove_diacritics(s) == req_no_accent:
            return s

    # 4. Token-based exact set match (ignores word order and accents)
    req_tokens = set(req_no_accent.split())
    if req_tokens:
        for s in available_sheets:
            s_tokens = set(remove_diacritics(s).split())
            if req_tokens == s_tokens:
                return s

        best_match = None
        best_score = 0
        for s in available_sheets:
            s_tokens = set(remove_diacritics(s).split())
            if req_tokens.issubset(s_tokens):
                score = len(req_tokens) / max(len(s_tokens), 1)
                if score > best_score:
                    best_score = score
                    best_match = s
            elif s_tokens.issubset(req_tokens):
                score = len(s_tokens) / max(len(req_tokens), 1)
                if score > best_score:
                    best_score = score
                    best_match = s

        if best_match and best_score >= 0.6:
            return best_match

    # 5. Substring match
    for s in available_sheets:
        s_no_accent = remove_diacritics(s)
        if req_no_accent in s_no_accent or s_no_accent in req_no_accent:
            return s

    # Fallback to first sheet
    return available_sheets[0]


def extract_excel_ranges_from_text(
    text: str,
    sheet_name_hint: Optional[str] = None,
    available_sheets: Optional[List[str]] = None,
) -> List[str]:
    """
    Extracts all Excel range references from user natural language / voice prompt.
    Correctly ignores sheet names (e.g. 'HN Chính T8') to prevent 'T8' from being parsed as range.
    e.g. 'Tôi cần bạn xem từ dòng H6 đến H137 và I6 đến I137 xem có bị trùng lặp không (HN Chính T8)' -> ['H6:H137', 'I6:I137']
    """
    ranges: List[str] = []

    # 1. Strip explicit sheet mentions in parentheses / quotes to avoid token collision like 'T8'
    cleaned = re.sub(r"\([^)]+\)", " ", text)
    cleaned = re.sub(r"['\"][^'\"]+['\"]", " ", cleaned)

    if sheet_name_hint:
        cleaned = cleaned.replace(sheet_name_hint, " ")
        cleaned = cleaned.replace(remove_diacritics(sheet_name_hint), " ")

    if available_sheets:
        for s in sorted(available_sheets, key=lambda x: len(x), reverse=True):
            cleaned = cleaned.replace(s, " ")
            cleaned = cleaned.replace(remove_diacritics(s), " ")

    # 2. Normalize voice tokens like "H 6" -> "H6", "I 6" -> "I6"
    pre_normalized = re.sub(r"\b([A-Za-z])\s+(\d+)\b", r"\1\2", cleaned)

    # 3. Handle "cột H từ 6 đến 137" -> "H6:H137" or "cột H từ dòng 6 đến dòng 137" -> "H6:H137"
    pre_normalized = re.sub(
        r"(?:cột|cot)\s+([A-Za-z]+)\s*(?:từ|tu)?\s*(?:dòng|dong)?\s*(\d+)\s*(?:đến|tới|sang|qua|to|-)\s*(?:dòng|dong)?\s*(\d+)",
        r"\1\2:\1\3",
        pre_normalized,
        flags=re.IGNORECASE,
    )

    # 4. Replace "(từ )?(dòng |cột |ô )?H6 đến H137" -> "H6:H137"
    normalized = re.sub(
        r"(?:từ\s+|tu\s+)?(?:dòng\s+|dong\s+|cột\s+|cot\s+|ô\s+|cell\s+)?([A-Za-z]+\d+)\s*(?:đến|tới|sang|qua|to|-|->)\s*(?:dòng\s+|dong\s+|cột\s+|cot\s+|ô\s+|cell\s+)?([A-Za-z]+\d+)",
        r"\1:\2",
        pre_normalized,
        flags=re.IGNORECASE,
    )

    # 5. Handle shorthand "từ H6 đến 137" -> "H6:H137"
    normalized = re.sub(
        r"(?:từ\s+|tu\s+)?(?:dòng\s+|dong\s+|cột\s+|cot\s+|ô\s+|cell\s+)?([A-Za-z]+)(\d+)\s*(?:đến|tới|sang|qua|to|-)\s*(\d+)",
        r"\1\2:\1\3",
        normalized,
        flags=re.IGNORECASE,
    )

    # 6. Extract multi-cell ranges (H6:H137, A1:B10, H:I, etc.)
    multi_range_pattern = r"(?:(?:'[^']+'|[A-Za-z0-9_]+)!)?[A-Za-z]+\d+:[A-Za-z]+\d+|[A-Za-z]+:[A-Za-z]+"
    for m in re.finditer(multi_range_pattern, normalized):
        r_str = m.group(0).strip()
        parsed = parse_excel_range(r_str)
        if parsed.get("valid") and r_str not in ranges:
            ranges.append(r_str)

    # 7. Extract standalone single cells ONLY if explicitly preceded by cell markers ('ô H11', 'cell I25')
    remaining_text = re.sub(multi_range_pattern, " ", normalized)
    single_cell_pattern = r"\b(?:ô|cell|tại)\s+([A-Za-z]+\d+)\b"
    for m in re.finditer(single_cell_pattern, remaining_text, flags=re.IGNORECASE):
        cell_str = m.group(1).upper().strip()
        if cell_str not in ranges:
            parsed = parse_excel_range(cell_str)
            if parsed.get("valid"):
                ranges.append(cell_str)

    return ranges


class SpreadsheetQueryEngine:
    """
    High-performance, deterministic engine for querying real spreadsheet cells,
    finding duplicates across/within ranges, comparing columns, detecting missing/outliers,
    and generating structured cell references for direct UI highlight.
    """

    @classmethod
    def normalize_value(cls, val: Any, mode: str = "normalized") -> str:
        """
        Normalizes a cell value for robust comparison.
        - Trims whitespace
        - Handles integer floats (123.0 -> "123")
        - Decomposes and recomposes Unicode (NFC)
        - Lowercases if mode == 'normalized'
        """
        if val is None or (isinstance(val, float) and math.isnan(val)):
            return ""

        # Convert date / datetime to standard string
        if isinstance(val, (pd.Timestamp, pd.DatetimeIndex)):
            return val.strftime("%Y-%m-%d %H:%M:%S").strip()

        if isinstance(val, float) and val.is_integer():
            val_str = str(int(val))
        else:
            val_str = str(val).strip()

        # Remove trailing .0 from numeric strings
        if re.match(r"^-?\d+\.0$", val_str):
            val_str = val_str[:-2]

        if not val_str:
            return ""

        # Unicode normalization
        val_str = unicodedata.normalize("NFC", val_str)
        # Collapse multiple inner spaces
        val_str = re.sub(r"\s+", " ", val_str)

        if mode == "normalized":
            val_str = val_str.lower()

        return val_str

    @classmethod
    def is_blank(cls, val: Any) -> bool:
        """Checks whether cell is blank, None, NaN or whitespace-only."""
        if val is None:
            return True
        if isinstance(val, float) and math.isnan(val):
            return True
        s = str(val).strip()
        return len(s) == 0 or s.lower() in ["nan", "null", "none"]

    @classmethod
    def read_range_cells(
        cls,
        file_path: str,
        sheet_name: str,
        range_spec: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """
        Reads cells within range_spec from real workbook file using exact openpyxl cell coordinates.
        Never relies on pandas DataFrame index offset.
        """
        path = Path(file_path)
        ext = path.suffix.lower()

        cells: List[Dict[str, Any]] = []

        if ext in [".xlsx", ".xls", ".xlsm"]:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            target_sheet_name = resolve_sheet_name_in_wb(sheet_name, wb.sheetnames)
            ws = wb[target_sheet_name]

            start_row = max(1, range_spec["start_row"])
            end_row = min(range_spec["end_row"], ws.max_row if ws.max_row else range_spec["end_row"])
            start_col = max(1, range_spec["start_col_idx"])
            end_col = min(range_spec["end_col_idx"], ws.max_column if ws.max_column else range_spec["end_col_idx"])

            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    col_letter = col_index_to_letter(c)
                    addr = f"{col_letter}{r}"
                    cell = ws.cell(row=r, column=c)
                    raw_val = cell.value
                    display_val = str(raw_val) if raw_val is not None else ""
                    norm_val = cls.normalize_value(raw_val)
                    is_empty = cls.is_blank(raw_val)

                    cells.append({
                        "address": addr,
                        "row": r,
                        "column": col_letter,
                        "column_index": c,
                        "value": raw_val,
                        "display_value": display_val,
                        "normalized_value": norm_val,
                        "is_blank": is_empty,
                    })
            wb.close()
            return cells

        # CSV fallback via pandas (row 1 is row 1)
        try:
            df = pd.read_csv(file_path, header=None)
        except Exception:
            df = pd.read_csv(file_path, header=None, encoding="latin-1")

        start_row = max(1, range_spec["start_row"])
        end_row = min(range_spec["end_row"], len(df))
        start_col = max(1, range_spec["start_col_idx"])
        end_col = min(range_spec["end_col_idx"], len(df.columns))

        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                col_letter = col_index_to_letter(c)
                addr = f"{col_letter}{r}"
                raw_val = df.iloc[r - 1, c - 1]
                display_val = str(raw_val) if (raw_val is not None and not pd.isna(raw_val)) else ""
                norm_val = cls.normalize_value(raw_val)
                is_empty = cls.is_blank(raw_val)

                cells.append({
                    "address": addr,
                    "row": r,
                    "column": col_letter,
                    "column_index": c,
                    "value": raw_val,
                    "display_value": display_val,
                    "normalized_value": norm_val,
                    "is_blank": is_empty,
                })

        return cells

    @classmethod
    def _load_sheet(cls, file_path: str, sheet_name: Optional[str] = None):
        path = Path(file_path)
        if path.suffix.lower() not in [".xlsx", ".xls", ".xlsm"]:
            raise ValueError("Workbook query tools hiện hỗ trợ tốt nhất cho file Excel.")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        resolved_sheet = resolve_sheet_name_in_wb(sheet_name, wb.sheetnames)
        return wb, wb[resolved_sheet], resolved_sheet

    @classmethod
    def _norm_key(cls, text: Any) -> str:
        base = remove_diacritics(str(text or ""))
        base = base.replace("_", " ").replace("-", " ")
        return re.sub(r"\s+", " ", base).strip().lower()

    @classmethod
    def _header_row(cls, ws) -> int:
        best_row = 1
        best_count = -1
        max_scan = min(ws.max_row or 1, 20)
        for row_idx in range(1, max_scan + 1):
            values = [ws.cell(row=row_idx, column=c).value for c in range(1, (ws.max_column or 1) + 1)]
            count = sum(1 for value in values if not cls.is_blank(value))
            if count > best_count:
                best_count = count
                best_row = row_idx
        return best_row

    @classmethod
    def _sheet_records(cls, file_path: str, sheet_name: str) -> Tuple[str, int, List[Dict[str, Any]], List[Dict[str, Any]]]:
        wb, ws, resolved_sheet = cls._load_sheet(file_path, sheet_name)
        try:
            header_row = cls._header_row(ws)
            columns: List[Dict[str, Any]] = []
            seen: Dict[str, int] = {}
            for col_idx in range(1, (ws.max_column or 1) + 1):
                raw_name = ws.cell(row=header_row, column=col_idx).value
                name = str(raw_name).strip() if not cls.is_blank(raw_name) else f"Column {col_idx}"
                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 1
                letter = col_index_to_letter(col_idx)
                sample_values = []
                numeric_count = 0
                non_empty_count = 0
                for row_idx in range(header_row + 1, min(ws.max_row or header_row, header_row + 10) + 1):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    if not cls.is_blank(value):
                        non_empty_count += 1
                        sample_values.append(value)
                        if isinstance(value, (int, float)) and not isinstance(value, bool):
                            numeric_count += 1
                columns.append({
                    "name": name,
                    "letter": letter,
                    "index": col_idx,
                    "header_row": header_row,
                    "sample_values": sample_values[:5],
                    "type": "numeric" if non_empty_count and numeric_count >= max(1, non_empty_count * 0.6) else "text",
                })

            records: List[Dict[str, Any]] = []
            for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                record = {"_row_number": row_idx}
                has_value = False
                for col in columns:
                    value = ws.cell(row=row_idx, column=col["index"]).value
                    if not cls.is_blank(value):
                        has_value = True
                    record[col["name"]] = value
                if has_value:
                    records.append(record)
            return resolved_sheet, header_row, columns, records
        finally:
            wb.close()

    @classmethod
    def _coerce_number(cls, value: Any) -> Optional[float]:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value)):
            return float(value)
        text = str(value).strip()
        if not text:
            return None
        text = re.sub(r"[^\d,\.\-]", "", text)
        if not text:
            return None
        if "," in text and "." in text:
            text = text.replace(".", "").replace(",", ".")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    @classmethod
    def get_workbook_info(cls, file_path: str, active_sheet: Optional[str] = None) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            resolved_active = resolve_sheet_name_in_wb(active_sheet, sheet_names)
            sheets = []
            for name in sheet_names:
                ws = wb[name]
                sheets.append({"name": name, "max_row": ws.max_row, "max_column": ws.max_column})
            return {
                "sheet_names": sheet_names,
                "active_sheet": resolved_active,
                "sheets": sheets,
                "evidence": {
                    "sheet": resolved_active,
                    "ranges": [],
                    "operation": "GET_WORKBOOK_INFO",
                    "rowCount": len(sheets),
                },
            }
        finally:
            wb.close()

    @classmethod
    def get_sheet_schema(cls, file_path: str, sheet_name: str) -> Dict[str, Any]:
        resolved_sheet, header_row, columns, records = cls._sheet_records(file_path, sheet_name)
        return {
            "sheet": resolved_sheet,
            "header_row": header_row,
            "row_count": len(records),
            "columns": columns,
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"A{header_row}:{col_index_to_letter(len(columns))}{header_row}"],
                "operation": "GET_SHEET_SCHEMA",
                "rowCount": 1,
            },
        }

    @classmethod
    def find_column(cls, file_path: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        schema = cls.get_sheet_schema(file_path, sheet_name)
        requested = cls._norm_key(column_name)
        requested_tokens = set(requested.split())
        best_col = None
        best_score = 0.0
        for col in schema["columns"]:
            candidate = cls._norm_key(col["name"])
            candidate_tokens = set(candidate.split())
            if candidate == requested:
                score = 1.0
            elif requested and (requested in candidate or candidate in requested):
                score = min(len(requested), len(candidate)) / max(len(requested), len(candidate))
            else:
                overlap = len(requested_tokens & candidate_tokens)
                score = overlap / max(len(requested_tokens | candidate_tokens), 1)
            if score > best_score:
                best_score = score
                best_col = col
        if not best_col or best_score < 0.35:
            return {
                "found": False,
                "requested": column_name,
                "sheet": schema["sheet"],
                "confidence": 0,
                "candidates": [c["name"] for c in schema["columns"][:12]],
                "evidence": {"sheet": schema["sheet"], "ranges": [], "operation": "FIND_COLUMN", "rowCount": 0},
            }
        return {
            "found": True,
            "requested": column_name,
            "sheet": schema["sheet"],
            "name": best_col["name"],
            "letter": best_col["letter"],
            "index": best_col["index"],
            "header_row": best_col["header_row"],
            "confidence": round(best_score, 3),
            "evidence": {
                "sheet": schema["sheet"],
                "ranges": [f"{best_col['letter']}{best_col['header_row']}"],
                "operation": "FIND_COLUMN",
                "rowCount": 1,
            },
        }

    @classmethod
    def read_range(cls, file_path: str, sheet_name: str, range_text: str) -> Dict[str, Any]:
        parsed = parse_excel_range(range_text)
        if not parsed.get("valid"):
            return {"operation": "READ_RANGE", "sheet": sheet_name, "ranges": [range_text], "cells": [], "error": "Range không hợp lệ."}
        wb, _ws, resolved_sheet = cls._load_sheet(file_path, parsed.get("sheet_name") or sheet_name)
        wb.close()
        cells = cls.read_range_cells(file_path, resolved_sheet, parsed)
        return {
            "operation": "READ_RANGE",
            "sheet": resolved_sheet,
            "ranges": [range_text],
            "cells": cells,
            "evidence": {"sheet": resolved_sheet, "ranges": [range_text], "operation": "READ_RANGE", "rowCount": len(cells)},
        }

    @classmethod
    def aggregate_column(cls, file_path: str, sheet_name: str, column_name: str, op: str = "sum") -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        col = cls.find_column(file_path, resolved_sheet, column_name)
        if not col.get("found"):
            return {"operation": "AGGREGATE", "sheet": resolved_sheet, "error": f"Không tìm thấy cột '{column_name}'.", "evidence": col["evidence"]}
        values = [cls._coerce_number(row.get(col["name"])) for row in records]
        numbers = [v for v in values if v is not None]
        op_key = op.lower()
        if op_key in ["avg", "average", "mean", "trung bình", "trung binh"]:
            value = sum(numbers) / len(numbers) if numbers else None
            operation = "AVERAGE"
        elif op_key in ["min", "nhỏ nhất", "nho nhat"]:
            value = min(numbers) if numbers else None
            operation = "MIN"
        elif op_key in ["max", "cao nhất", "cao nhat", "lớn nhất", "lon nhat"]:
            value = max(numbers) if numbers else None
            operation = "MAX"
        elif op_key in ["count", "đếm", "dem"]:
            value = len([row for row in records if not cls.is_blank(row.get(col["name"]))])
            operation = "COUNT"
        else:
            value = sum(numbers)
            operation = "SUM"
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        data_start = col["header_row"] + 1
        data_end = data_start + len(records) - 1
        return {
            "operation": "AGGREGATE",
            "sheet": resolved_sheet,
            "column": col,
            "value": value,
            "numericCount": len(numbers),
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"{col['letter']}{data_start}:{col['letter']}{data_end}"],
                "operation": operation,
                "rowCount": len(numbers) if operation != "COUNT" else len(records),
            },
        }

    @classmethod
    def find_top_rows(cls, file_path: str, sheet_name: str, column_name: str, limit: int = 1, descending: bool = True) -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        col = cls.find_column(file_path, resolved_sheet, column_name)
        if not col.get("found"):
            return {"operation": "FIND_TOP_ROWS", "sheet": resolved_sheet, "rows": [], "error": f"Không tìm thấy cột '{column_name}'.", "evidence": col["evidence"]}
        ranked = []
        for row in records:
            num = cls._coerce_number(row.get(col["name"]))
            if num is not None:
                ranked.append({"row_number": row["_row_number"], "value": int(num) if num.is_integer() else num, "record": {k: v for k, v in row.items() if k != "_row_number"}})
        ranked.sort(key=lambda item: item["value"], reverse=descending)
        data_start = col["header_row"] + 1
        data_end = data_start + len(records) - 1
        return {
            "operation": "FIND_TOP_ROWS",
            "sheet": resolved_sheet,
            "column": col,
            "rows": ranked[:limit],
            "evidence": {"sheet": resolved_sheet, "ranges": [f"{col['letter']}{data_start}:{col['letter']}{data_end}"], "operation": "MAX" if descending else "MIN", "rowCount": len(ranked)},
        }

    @classmethod
    def search_rows(cls, file_path: str, sheet_name: str, query: str, limit: int = 20) -> Dict[str, Any]:
        resolved_sheet, header_row, columns, records = cls._sheet_records(file_path, sheet_name)
        needle = cls._norm_key(query)
        matches = []
        for row in records:
            haystack = " ".join(cls._norm_key(row.get(col["name"])) for col in columns)
            if needle and needle in haystack:
                matches.append({"row_number": row["_row_number"], "record": {k: v for k, v in row.items() if k != "_row_number"}})
            if len(matches) >= limit:
                break
        return {
            "operation": "SEARCH_ROWS",
            "sheet": resolved_sheet,
            "query": query,
            "matches": matches,
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"A{header_row + 1}:{col_index_to_letter(len(columns))}{header_row + len(records)}"],
                "operation": "SEARCH_ROWS",
                "rowCount": len(records),
            },
        }

    @classmethod
    def search_and_count_entity(cls, file_path: str, sheet_name: str, query: str, limit: int = 500) -> Dict[str, Any]:
        resolved_sheet, header_row, columns, records = cls._sheet_records(file_path, sheet_name)
        needle = cls._norm_key(query)
        if not needle:
            return {
                "operation": "SEARCH_AND_COUNT",
                "sheet": resolved_sheet,
                "query": query,
                "total_occurrences": 0,
                "unique_rows_count": 0,
                "breakdown_by_column": {},
                "matched_cells": [],
                "matched_rows": [],
                "evidence": {"sheet": resolved_sheet, "ranges": [], "operation": "SEARCH_AND_COUNT", "rowCount": 0},
            }

        matched_cells = []
        matched_row_numbers = set()
        breakdown_by_column = {}
        matched_rows_samples = []

        for row in records:
            row_num = row["_row_number"]
            row_matched = False
            for col in columns:
                col_name = col["name"]
                val = row.get(col_name)
                val_norm = cls._norm_key(val)
                if needle in val_norm:
                    cell_addr = f"{col['letter']}{row_num}"
                    matched_cells.append({
                        "address": cell_addr,
                        "row": row_num,
                        "column": col_name,
                        "value": val,
                    })
                    matched_row_numbers.add(row_num)
                    row_matched = True
                    breakdown_by_column[col_name] = breakdown_by_column.get(col_name, 0) + 1
            if row_matched and len(matched_rows_samples) < 10:
                matched_rows_samples.append({k: v for k, v in row.items() if k != "_row_number"})

        cell_addrs = [c["address"] for c in matched_cells]
        return {
            "operation": "SEARCH_AND_COUNT",
            "sheet": resolved_sheet,
            "query": query,
            "total_occurrences": len(matched_cells),
            "unique_rows_count": len(matched_row_numbers),
            "breakdown_by_column": breakdown_by_column,
            "matched_cells": matched_cells[:limit],
            "matched_rows": list(sorted(matched_row_numbers)),
            "sample_rows": matched_rows_samples,
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": cell_addrs[:20],
                "operation": "SEARCH_AND_COUNT",
                "rowCount": len(matched_row_numbers),
            },
        }

    @classmethod
    def get_cell(cls, file_path: str, sheet_name: str, address: str) -> Dict[str, Any]:
        parsed = parse_excel_range(address)
        if not parsed.get("valid") or not parsed.get("is_single_cell"):
            return {"operation": "GET_CELL", "sheet": sheet_name, "address": address, "error": "Địa chỉ ô không hợp lệ."}
        wb, ws, resolved_sheet = cls._load_sheet(file_path, parsed.get("sheet_name") or sheet_name)
        try:
            cell = ws[address.upper()]
            return {
                "operation": "GET_CELL",
                "sheet": resolved_sheet,
                "address": address.upper(),
                "value": cell.value,
                "display_value": "" if cell.value is None else str(cell.value),
                "evidence": {"sheet": resolved_sheet, "ranges": [address.upper()], "operation": "GET_CELL", "rowCount": 1},
            }
        finally:
            wb.close()

    @classmethod
    def filter_rows(
        cls,
        file_path: str,
        sheet_name: str,
        column_name: str,
        operator: str,
        compare_value: Any,
        limit: int = 200,
    ) -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        col = cls.find_column(file_path, resolved_sheet, column_name)
        if not col.get("found"):
            return {"operation": "FILTER_ROWS", "sheet": resolved_sheet, "rows": [], "error": f"Không tìm thấy cột '{column_name}'.", "evidence": col["evidence"]}

        compare_values = compare_value if isinstance(compare_value, (list, tuple)) else [compare_value]
        compare_nums = [cls._coerce_number(value) for value in compare_values]
        compare_num = compare_nums[0] if compare_nums else None
        compare_text = cls._norm_key(compare_value)
        rows = []
        for row in records:
            value = row.get(col["name"])
            value_num = cls._coerce_number(value)
            matched = False
            if operator == "between" and len(compare_nums) >= 2 and compare_nums[0] is not None and compare_nums[1] is not None and value_num is not None:
                low, high = sorted([compare_nums[0], compare_nums[1]])
                matched = low <= value_num <= high
            elif operator in ["<", "<=", ">", ">=", "="] and compare_num is not None and value_num is not None:
                if operator == "<":
                    matched = value_num < compare_num
                elif operator == "<=":
                    matched = value_num <= compare_num
                elif operator == ">":
                    matched = value_num > compare_num
                elif operator == ">=":
                    matched = value_num >= compare_num
                elif operator == "=":
                    matched = value_num == compare_num
            elif operator in ["contains", "="]:
                norm_value = cls._norm_key(value)
                matched = norm_value == compare_text if operator == "=" else compare_text in norm_value
            if matched:
                rows.append({"row_number": row["_row_number"], "value": value, "record": {k: v for k, v in row.items() if k != "_row_number"}})
            if len(rows) >= limit:
                break

        data_start = col["header_row"] + 1
        data_end = data_start + len(records) - 1
        matched_cells = [
            {
                "address": f"{col['letter']}{row['row_number']}",
                "sheet": resolved_sheet,
                "row": row["row_number"],
                "column": col["name"],
                "cell": f"{col['letter']}{row['row_number']}",
                "value": row["value"],
                "record": row["record"],
            }
            for row in rows
        ]
        return {
            "operation": "FILTER_ROWS",
            "sheet": resolved_sheet,
            "column": col,
            "operator": operator,
            "compare_value": compare_value,
            "matched_count": len(rows),
            "matched_cells": matched_cells,
            "rows": rows,
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"{col['letter']}{data_start}:{col['letter']}{data_end}"],
                "operation": "FILTER_ROWS",
                "rowCount": len(records),
            },
        }

    @classmethod
    def compare_groups(cls, file_path: str, sheet_name: str, group_column: str, value_column: str, op: str = "sum") -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        group_col = cls.find_column(file_path, resolved_sheet, group_column)
        value_col = cls.find_column(file_path, resolved_sheet, value_column)
        if not group_col.get("found") or not value_col.get("found"):
            return {"operation": "COMPARE_GROUPS", "sheet": resolved_sheet, "groups": [], "error": "Không tìm thấy cột để so sánh.", "evidence": {"sheet": resolved_sheet, "ranges": [], "operation": "COMPARE_GROUPS", "rowCount": 0}}
        buckets: Dict[str, List[float]] = {}
        for row in records:
            group = str(row.get(group_col["name"]) or "(Trống)").strip()
            num = cls._coerce_number(row.get(value_col["name"]))
            if num is not None:
                buckets.setdefault(group, []).append(num)
        groups = []
        for group, nums in sorted(buckets.items(), key=lambda item: item[0]):
            if op.lower() in ["avg", "average", "mean", "trung bình", "trung binh"]:
                value = sum(nums) / len(nums)
                operation = "AVERAGE_BY_GROUP"
            else:
                value = sum(nums)
                operation = "SUM_BY_GROUP"
            groups.append({"group": group, "value": int(value) if float(value).is_integer() else value, "count": len(nums)})
        data_start = value_col["header_row"] + 1
        data_end = data_start + len(records) - 1
        return {
            "operation": "COMPARE_GROUPS",
            "sheet": resolved_sheet,
            "group_column": group_col,
            "value_column": value_col,
            "groups": groups,
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"{group_col['letter']}{data_start}:{group_col['letter']}{data_end}", f"{value_col['letter']}{data_start}:{value_col['letter']}{data_end}"],
                "operation": operation,
                "rowCount": len(records),
            },
        }

    @classmethod
    def detect_outliers(cls, file_path: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        col = cls.find_column(file_path, resolved_sheet, column_name)
        if not col.get("found"):
            return {"operation": "DETECT_OUTLIERS", "sheet": resolved_sheet, "outliers": [], "error": f"Không tìm thấy cột '{column_name}'.", "evidence": col["evidence"]}
        values = []
        for row in records:
            num = cls._coerce_number(row.get(col["name"]))
            if num is not None:
                values.append((row, num))
        nums = sorted(num for _row, num in values)
        outliers = []
        if len(nums) >= 4:
            q1 = nums[len(nums) // 4]
            q3 = nums[(len(nums) * 3) // 4]
            iqr = q3 - q1
            low = q1 - 1.5 * iqr
            high = q3 + 1.5 * iqr
            for row, num in values:
                if num < low or num > high:
                    outliers.append({"row_number": row["_row_number"], "value": int(num) if num.is_integer() else num, "record": {k: v for k, v in row.items() if k != "_row_number"}})
        data_start = col["header_row"] + 1
        data_end = data_start + len(records) - 1
        return {
            "operation": "DETECT_OUTLIERS",
            "sheet": resolved_sheet,
            "column": col,
            "outliers": outliers,
            "evidence": {"sheet": resolved_sheet, "ranges": [f"{col['letter']}{data_start}:{col['letter']}{data_end}"], "operation": "DETECT_OUTLIERS", "rowCount": len(values)},
        }

    @classmethod
    def find_max(cls, file_path: str, sheet_name: str, column_name: str, limit: int = 1) -> Dict[str, Any]:
        return cls.find_top_rows(file_path, sheet_name, column_name, limit=limit, descending=True)

    @classmethod
    def find_min(cls, file_path: str, sheet_name: str, column_name: str, limit: int = 1) -> Dict[str, Any]:
        return cls.find_top_rows(file_path, sheet_name, column_name, limit=limit, descending=False)

    @classmethod
    def sum(cls, file_path: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        return cls.aggregate_column(file_path, sheet_name, column_name, op="sum")

    @classmethod
    def average(cls, file_path: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        return cls.aggregate_column(file_path, sheet_name, column_name, op="average")

    @classmethod
    def count(cls, file_path: str, sheet_name: str, column_name: Optional[str] = None) -> Dict[str, Any]:
        if column_name:
            return cls.aggregate_column(file_path, sheet_name, column_name, op="count")
        schema = cls.get_sheet_schema(file_path, sheet_name)
        last_col = schema["columns"][-1]["letter"] if schema.get("columns") else "A"
        return {
            "operation": "COUNT",
            "sheet": schema["sheet"],
            "value": schema["row_count"],
            "evidence": {
                "sheet": schema["sheet"],
                "ranges": [f"A{schema['header_row'] + 1}:{last_col}{schema['header_row'] + schema['row_count']}"],
                "operation": "COUNT",
                "rowCount": schema["row_count"],
            },
        }

    @classmethod
    def count_distinct(cls, file_path: str, sheet_name: str, column_name: str) -> Dict[str, Any]:
        resolved_sheet, _header_row, _columns, records = cls._sheet_records(file_path, sheet_name)
        col = cls.find_column(file_path, resolved_sheet, column_name)
        if not col.get("found"):
            return {"operation": "COUNT_DISTINCT", "sheet": resolved_sheet, "error": f"Không tìm thấy cột '{column_name}'."}
        distinct_values = set()
        for row in records:
            val = row.get(col["name"])
            if not cls.is_blank(val):
                distinct_values.add(str(val).strip())
        data_start = col["header_row"] + 1
        data_end = data_start + len(records) - 1
        return {
            "operation": "COUNT_DISTINCT",
            "sheet": resolved_sheet,
            "column": col,
            "value": len(distinct_values),
            "distinct_values": sorted(list(distinct_values))[:20],
            "evidence": {
                "sheet": resolved_sheet,
                "ranges": [f"{col['letter']}{data_start}:{col['letter']}{data_end}"],
                "operation": "COUNT_DISTINCT",
                "rowCount": len(records),
            },
        }

    @classmethod
    def cross_sheet_compare(
        cls,
        file_path: str,
        sheet1_name: str,
        sheet2_name: str,
        key_column_name: Optional[str] = None,
        metric_column_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Compares records and metrics across 2 sheets in the same workbook.
        """
        s1, h1, cols1, recs1 = cls._sheet_records(file_path, sheet1_name)
        s2, h2, cols2, recs2 = cls._sheet_records(file_path, sheet2_name)

        # Auto-detect common key column if not specified
        key1 = cls.find_column(file_path, s1, key_column_name)["name"] if key_column_name else (cols1[0]["name"] if cols1 else None)
        key2 = cls.find_column(file_path, s2, key_column_name)["name"] if key_column_name else (cols2[0]["name"] if cols2 else None)

        # Build maps
        map1 = {}
        for r in recs1:
            k = str(r.get(key1) or "").strip()
            if k:
                map1[k] = r

        map2 = {}
        for r in recs2:
            k = str(r.get(key2) or "").strip()
            if k:
                map2[k] = r

        all_keys = set(map1.keys()).union(set(map2.keys()))
        in_s1_only = sorted(list(set(map1.keys()) - set(map2.keys())))
        in_s2_only = sorted(list(set(map2.keys()) - set(map1.keys())))
        common_keys = sorted(list(set(map1.keys()).intersection(set(map2.keys()))))

        discrepancies = []
        if in_s1_only:
            discrepancies.append(f"Có {len(in_s1_only)} khóa chỉ xuất hiện ở '{s1}' (ví dụ: {', '.join(in_s1_only[:4])})")
        if in_s2_only:
            discrepancies.append(f"Có {len(in_s2_only)} khóa chỉ xuất hiện ở '{s2}' (ví dụ: {', '.join(in_s2_only[:4])})")

        evidence = {
            "sheet": f"{s1} vs {s2}",
            "ranges": [f"{s1}!A{h1}:{col_index_to_letter(len(cols1))}{h1 + len(recs1)}", f"{s2}!A{h2}:{col_index_to_letter(len(cols2))}{h2 + len(recs2)}"],
            "operation": "CROSS_SHEET_COMPARE",
            "rowCount": len(all_keys),
        }

        return {
            "operation": "CROSS_SHEET_COMPARE",
            "sheet1": s1,
            "sheet2": s2,
            "key1": key1,
            "key2": key2,
            "total_keys": len(all_keys),
            "common_count": len(common_keys),
            "in_sheet1_only_count": len(in_s1_only),
            "in_sheet2_only_count": len(in_s2_only),
            "in_sheet1_only": in_s1_only[:20],
            "in_sheet2_only": in_s2_only[:20],
            "discrepancies": discrepancies,
            "evidence": evidence,
        }

    @classmethod
    def cross_file_compare(
        cls,
        file_path_1: str,
        sheet1_name: Optional[str] = None,
        file_path_2: str = "",
        sheet2_name: Optional[str] = None,
        key_column_1: Optional[str] = None,
        key_column_2: Optional[str] = None,
        compare_columns: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Cross-compares 2 independent files (Excel or CSV) with key-based reconciliation,
        finding missing rows and value discrepancies.
        """
        s1, h1, cols1, recs1 = cls._sheet_records(file_path_1, sheet1_name)
        s2, h2, cols2, recs2 = cls._sheet_records(file_path_2, sheet2_name)

        # Resolve keys
        k1_name = key_column_1 or (cols1[0]["name"] if cols1 else "Key")
        k2_name = key_column_2 or (cols2[0]["name"] if cols2 else "Key")

        col1_dict = {cls._norm_key(c["name"]): c for c in cols1}
        col2_dict = {cls._norm_key(c["name"]): c for c in cols2}

        # Index maps
        map1: Dict[str, Dict[str, Any]] = {}
        for r in recs1:
            raw_k = r.get(k1_name)
            k = cls._norm_key(raw_k)
            if k:
                map1[k] = r

        map2: Dict[str, Dict[str, Any]] = {}
        for r in recs2:
            raw_k = r.get(k2_name)
            k = cls._norm_key(raw_k)
            if k:
                map2[k] = r

        all_keys = set(map1.keys()).union(set(map2.keys()))
        in_f1_only = sorted(list(set(map1.keys()) - set(map2.keys())))
        in_f2_only = sorted(list(set(map2.keys()) - set(map1.keys())))
        common_keys = sorted(list(set(map1.keys()).intersection(set(map2.keys()))))

        # Check shared columns for value differences
        shared_cols = []
        for c1_norm, c1_meta in col1_dict.items():
            if c1_norm in col2_dict and c1_norm != cls._norm_key(k1_name):
                shared_cols.append((c1_meta["name"], col2_dict[c1_norm]["name"]))

        value_differences = []
        for k in common_keys:
            r1 = map1[k]
            r2 = map2[k]
            for col_a, col_b in shared_cols:
                v1 = r1.get(col_a)
                v2 = r2.get(col_b)
                if cls._norm_key(v1) != cls._norm_key(v2):
                    value_differences.append({
                        "key": str(r1.get(k1_name) or k),
                        "row1": r1.get("_row_number"),
                        "row2": r2.get("_row_number"),
                        "column": col_a,
                        "file1_value": v1,
                        "file2_value": v2,
                    })

        summary_lines = []
        if in_f1_only:
            summary_lines.append(f"• Có **{len(in_f1_only)} bản ghi** chỉ có ở File 1 (thiếu ở File 2).")
        if in_f2_only:
            summary_lines.append(f"• Có **{len(in_f2_only)} bản ghi** chỉ có ở File 2 (thiếu ở File 1).")
        if value_differences:
            summary_lines.append(f"• Phát hiện **{len(value_differences)} điểm lệch số liệu** giữa 2 file cho các bản ghi khớp nhau.")
        if not summary_lines:
            summary_lines.append("✅ Dữ liệu giữa 2 file khớp nhau 100% không có sai lệch.")

        f1_label = Path(file_path_1).name if file_path_1 else "File 1"
        f2_label = Path(file_path_2).name if file_path_2 else "File 2"

        return {
            "operation": "CROSS_FILE_COMPARE",
            "file1_sheet": s1,
            "file2_sheet": s2,
            "file1_name": f1_label,
            "file2_name": f2_label,
            "key1": k1_name,
            "key2": k2_name,
            "total_keys": len(all_keys),
            "common_count": len(common_keys),
            "in_file1_only_count": len(in_f1_only),
            "in_file2_only_count": len(in_f2_only),
            "in_file1_only": in_f1_only[:30],
            "in_file2_only": in_f2_only[:30],
            "value_differences_count": len(value_differences),
            "value_differences": value_differences[:50],
            "summary": "\n".join(summary_lines),
            "evidence": {
                "sheet": f"{f1_label} vs {f2_label}",
                "ranges": [],
                "operation": "CROSS_FILE_COMPARE",
                "rowCount": len(all_keys),
            },
        }

    @classmethod
    def find_duplicates(
        cls,
        file_path: str,
        sheet_name: str,
        ranges: List[str],
        mode: str = "intersection",
        normalize: bool = True,
        ignore_blank: bool = True,
    ) -> Dict[str, Any]:
        """
        Executes deterministic duplicate detection:
        Computes 3 categories simultaneously:
        1. cross_range_duplicates (values appearing in both Range A and Range B)
        2. duplicates_in_first_range (internal duplicates within Range A)
        3. duplicates_in_second_range (internal duplicates within Range B)
        """
        norm_mode = "normalized" if normalize else "exact"
        parsed_ranges = [parse_excel_range(r) for r in ranges if parse_excel_range(r).get("valid")]

        # Resolve exact sheet name from workbook
        path = Path(file_path)
        ext = path.suffix.lower()
        if ext in [".xlsx", ".xls", ".xlsm"]:
            wb_check = openpyxl.load_workbook(file_path, read_only=True)
            resolved_sheet = resolve_sheet_name_in_wb(sheet_name, wb_check.sheetnames)
            wb_check.close()
        else:
            resolved_sheet = sheet_name or "Sheet1"

        if not parsed_ranges:
            return {
                "operation": "FIND_DUPLICATES",
                "sheet": resolved_sheet,
                "ranges": ranges,
                "error": "Không có vùng Excel (range) hợp lệ để kiểm tra.",
                "duplicate_count": 0,
                "matched_cells": [],
            }

        # Case 1: Multiple ranges (Comparison between 2 ranges, e.g. H6:H137 and I6:I137)
        if len(parsed_ranges) >= 2:
            range_a = parsed_ranges[0]
            range_b = parsed_ranges[1] if len(parsed_ranges) > 1 else parsed_ranges[0]

            cells_a = cls.read_range_cells(file_path, resolved_sheet, range_a)
            cells_b = cls.read_range_cells(file_path, resolved_sheet, range_b)

            all_cells_a_count = len(cells_a)
            all_cells_b_count = len(cells_b)

            # Filter blanks if requested
            non_empty_a = [c for c in cells_a if not c["is_blank"]] if ignore_blank else cells_a
            non_empty_b = [c for c in cells_b if not c["is_blank"]] if ignore_blank else cells_b

            # Map normalized value -> cells in A and cells in B
            map_a: Dict[str, List[Dict[str, Any]]] = {}
            for c in non_empty_a:
                map_a.setdefault(c["normalized_value"], []).append(c)

            map_b: Dict[str, List[Dict[str, Any]]] = {}
            for c in non_empty_b:
                map_b.setdefault(c["normalized_value"], []).append(c)

            # 1. Intersection values (Cross-range duplicates)
            intersection_keys = set(map_a.keys()).intersection(set(map_b.keys()))
            cross_range_duplicates = []
            matched_cells_dict: Dict[str, Dict[str, Any]] = {}

            for key in sorted(intersection_keys):
                in_a = map_a[key]
                in_b = map_b[key]
                sample_display = in_a[0]["display_value"] or in_b[0]["display_value"]

                for c in in_a + in_b:
                    matched_cells_dict[c["address"]] = {
                        "address": c["address"],
                        "row": c["row"],
                        "column": c["column"],
                        "value": c["value"],
                        "display_value": c["display_value"],
                        "reason": "duplicate_intersection",
                        "group_value": sample_display,
                    }

                cross_range_duplicates.append({
                    "value": sample_display,
                    "normalized_value": key,
                    "count_in_first_range": len(in_a),
                    "count_in_second_range": len(in_b),
                    "total_occurrences": len(in_a) + len(in_b),
                    "first_range_cells": [c["address"] for c in in_a],
                    "second_range_cells": [c["address"] for c in in_b],
                })

            # 2. Internal duplicates within Range A
            duplicates_in_first_range = []
            for key, group_cells in sorted(map_a.items()):
                if len(group_cells) > 1:
                    sample_display = group_cells[0]["display_value"]
                    for c in group_cells:
                        if c["address"] not in matched_cells_dict:
                            matched_cells_dict[c["address"]] = {
                                "address": c["address"],
                                "row": c["row"],
                                "column": c["column"],
                                "value": c["value"],
                                "display_value": c["display_value"],
                                "reason": "duplicate_within_first_range",
                                "group_value": sample_display,
                            }
                    duplicates_in_first_range.append({
                        "value": sample_display,
                        "normalized_value": key,
                        "count": len(group_cells),
                        "cells": [c["address"] for c in group_cells],
                    })

            # 3. Internal duplicates within Range B
            duplicates_in_second_range = []
            for key, group_cells in sorted(map_b.items()):
                if len(group_cells) > 1:
                    sample_display = group_cells[0]["display_value"]
                    for c in group_cells:
                        if c["address"] not in matched_cells_dict:
                            matched_cells_dict[c["address"]] = {
                                "address": c["address"],
                                "row": c["row"],
                                "column": c["column"],
                                "value": c["value"],
                                "display_value": c["display_value"],
                                "reason": "duplicate_within_second_range",
                                "group_value": sample_display,
                            }
                    duplicates_in_second_range.append({
                        "value": sample_display,
                        "normalized_value": key,
                        "count": len(group_cells),
                        "cells": [c["address"] for c in group_cells],
                    })

            # 4. Same-row matches (H_i == I_i)
            same_row_matches = []
            row_map_a = {c["row"]: c for c in non_empty_a}
            row_map_b = {c["row"]: c for c in non_empty_b}
            common_rows = set(row_map_a.keys()).intersection(set(row_map_b.keys()))
            for r in sorted(common_rows):
                if row_map_a[r]["normalized_value"] == row_map_b[r]["normalized_value"]:
                    same_row_matches.append({
                        "row": r,
                        "value": row_map_a[r]["display_value"],
                        "first_cell": row_map_a[r]["address"],
                        "second_cell": row_map_b[r]["address"],
                    })

            matched_cells_list = list(matched_cells_dict.values())
            all_unique_duplicate_keys = set(intersection_keys).union(
                set(d["normalized_value"] for d in duplicates_in_first_range)
            ).union(
                set(d["normalized_value"] for d in duplicates_in_second_range)
            )

            return {
                "operation": "COMPARE_DUPLICATES_COMPREHENSIVE",
                "sheet": resolved_sheet,
                "ranges": ranges,
                "comparison_mode": norm_mode,
                "ignore_blank": ignore_blank,
                "duplicate_count": len(all_unique_duplicate_keys),
                "cross_range_count": len(cross_range_duplicates),
                "within_first_range_count": len(duplicates_in_first_range),
                "within_second_range_count": len(duplicates_in_second_range),
                "same_row_match_count": len(same_row_matches),
                "total_matched_cells": len(matched_cells_list),
                "duplicate_groups": cross_range_duplicates or duplicates_in_first_range or duplicates_in_second_range,
                "cross_range_duplicates": cross_range_duplicates,
                "duplicates_in_first_range": duplicates_in_first_range,
                "duplicates_in_second_range": duplicates_in_second_range,
                "same_row_matches": same_row_matches,
                "matched_cells": matched_cells_list,
                "first_range": range_a["raw"],
                "second_range": range_b["raw"],
                "execution": {
                    "range_a_total_cells": all_cells_a_count,
                    "range_a_non_empty": len(non_empty_a),
                    "range_b_total_cells": all_cells_b_count,
                    "range_b_non_empty": len(non_empty_b),
                },
            }

        # Case 2: Single range internal duplicates (e.g. H6:H137)
        target_range = parsed_ranges[0]
        cells = cls.read_range_cells(file_path, resolved_sheet, target_range)
        all_cells_count = len(cells)
        non_empty_cells = [c for c in cells if not c["is_blank"]] if ignore_blank else cells

        val_map: Dict[str, List[Dict[str, Any]]] = {}
        for c in non_empty_cells:
            val_map.setdefault(c["normalized_value"], []).append(c)

        dup_groups = []
        matched_cells = []
        for key, group_cells in sorted(val_map.items()):
            if len(group_cells) > 1:
                sample_display = group_cells[0]["display_value"]
                for c in group_cells:
                    matched_cells.append({
                        "address": c["address"],
                        "row": c["row"],
                        "column": c["column"],
                        "value": c["value"],
                        "display_value": c["display_value"],
                        "reason": "duplicate_within_range",
                        "group_value": sample_display,
                    })
                dup_groups.append({
                    "value": sample_display,
                    "normalized_value": key,
                    "count": len(group_cells),
                    "cells": [c["address"] for c in group_cells],
                })

        return {
            "operation": "FIND_DUPLICATES_WITHIN_RANGE",
            "sheet": resolved_sheet,
            "ranges": ranges,
            "comparison_mode": norm_mode,
            "ignore_blank": ignore_blank,
            "duplicate_count": len(dup_groups),
            "cross_range_count": 0,
            "within_first_range_count": len(dup_groups),
            "within_second_range_count": 0,
            "total_matched_cells": len(matched_cells),
            "duplicate_groups": dup_groups,
            "duplicates_in_first_range": dup_groups,
            "duplicates_in_second_range": [],
            "matched_cells": matched_cells,
            "first_range": target_range["raw"],
            "execution": {
                "range_total_cells": all_cells_count,
                "range_non_empty": len(non_empty_cells),
            },
        }

    @classmethod
    def find_missing_cells(
        cls,
        file_path: str,
        sheet_name: str,
        ranges: List[str],
    ) -> Dict[str, Any]:
        """Finds all blank/missing cells in specified ranges."""
        matched_cells: List[Dict[str, Any]] = []
        parsed_ranges = [parse_excel_range(r) for r in ranges if parse_excel_range(r).get("valid")]

        for r_spec in parsed_ranges:
            cells = cls.read_range_cells(file_path, sheet_name, r_spec)
            for c in cells:
                if c["is_blank"]:
                    matched_cells.append({
                        "address": c["address"],
                        "row": c["row"],
                        "column": c["column"],
                        "value": None,
                        "display_value": "(Trống)",
                        "reason": "missing_value",
                    })

        return {
            "operation": "FIND_MISSING",
            "sheet": sheet_name,
            "ranges": ranges,
            "missing_count": len(matched_cells),
            "matched_cells": matched_cells,
        }

    @classmethod
    def apply_highlights_to_workbook(
        cls,
        file_path: str,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str = "FFFF00",  # Default yellow
        output_path: Optional[str] = None,
    ) -> str:
        """
        Applies permanent highlight fill into the physical XLSX file.
        Returns the output file path.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File không tồn tại: {file_path}")

        if not output_path:
            out_dir = path.parent
            stem = path.stem
            output_path = str(out_dir / f"{stem}_highlighted.xlsx")

        wb = openpyxl.load_workbook(file_path)
        target_sheet = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[target_sheet]

        clean_hex = color_hex.replace("#", "").upper()
        if len(clean_hex) == 6:
            clean_hex = "FF" + clean_hex

        fill = PatternFill(start_color=clean_hex, end_color=clean_hex, fill_type="solid")

        for addr in cell_addresses:
            try:
                cell = ws[addr]
                cell.fill = fill
            except Exception:
                pass

        wb.save(output_path)
        wb.close()
        return output_path


spreadsheet_query_engine = SpreadsheetQueryEngine()
