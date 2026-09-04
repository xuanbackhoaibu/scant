from pathlib import Path
from typing import Any, Dict, List, Optional
import logging
import openpyxl
from openpyxl.styles import PatternFill
from app.services.data.adapters import ExcelAdapter, GoogleSheetsAdapter, SpreadsheetAdapter
from app.services.data.spreadsheet_query_engine import col_index_to_letter, parse_excel_range, spreadsheet_query_engine

logger = logging.getLogger(__name__)


class SpreadsheetActionEngine:
    """
    Executes modifications and visual formatting actions on spreadsheets.
    Supports physical XLSX files and Google Sheets with undo history.
    """

    # In-memory undo history by session / workbook
    _undo_stacks: Dict[str, List[Dict[str, Any]]] = {}

    @classmethod
    def highlight_cells(
        cls,
        file_path: str,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str = "#FEF08A",
        session_id: Optional[str] = None,
        source_type: str = "excel",
        google_url: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Applies highlight color to specific cell coordinates.
        Preserves original fonts, borders, alignments, and formulas.
        """
        sid = session_id or "default"
        clean_hex = color_hex.replace("#", "").upper()
        if len(clean_hex) == 6:
            clean_hex = "FF" + clean_hex

        # Record undo snapshot
        cls._record_undo_snapshot(sid, file_path, sheet_name, cell_addresses, "HIGHLIGHT_CELLS")

        out_path = spreadsheet_query_engine.apply_highlights_to_workbook(
            file_path=file_path,
            sheet_name=sheet_name,
            cell_addresses=cell_addresses,
            color_hex=clean_hex,
        )

        return {
            "ok": True,
            "action": "HIGHLIGHT_CELLS",
            "sheet": sheet_name,
            "highlighted_count": len(cell_addresses),
            "cells": cell_addresses,
            "color": color_hex,
            "modified_file_path": out_path,
            "modified_file_name": Path(out_path).name,
        }

    @classmethod
    def highlight_rows(
        cls,
        file_path: str,
        sheet_name: str,
        row_numbers: List[int],
        color_hex: str = "#FEF08A",
        session_id: Optional[str] = None,
        source_type: str = "excel",
    ) -> Dict[str, Any]:
        """
        Highlights entire rows across used columns.
        """
        schema = spreadsheet_query_engine.get_sheet_schema(file_path, sheet_name)
        cols = schema.get("columns", [])
        cell_addresses = []
        for r in row_numbers:
            for col in cols:
                cell_addresses.append(f"{col['letter']}{r}")

        return cls.highlight_cells(
            file_path=file_path,
            sheet_name=sheet_name,
            cell_addresses=cell_addresses,
            color_hex=color_hex,
            session_id=session_id,
            source_type=source_type,
        )

    @classmethod
    def clear_highlights(
        cls,
        file_path: str,
        sheet_name: str,
        session_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Clears all highlight fills on the specified sheet.
        """
        wb = openpyxl.load_workbook(file_path)
        try:
            target = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
            ws = wb[target]
            no_fill = PatternFill(fill_type=None)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.fill_type:
                        cell.fill = no_fill
            wb.save(file_path)
            return {"ok": True, "action": "CLEAR_HIGHLIGHTS", "sheet": target}
        finally:
            wb.close()

    @classmethod
    def undo_last_action(cls, session_id: str = "default") -> Dict[str, Any]:
        """
        Reverts the last formatting action applied in the session.
        """
        stack = cls._undo_stacks.get(session_id, [])
        if not stack:
            return {"ok": False, "message": "Không có hành động nào để hoàn tác."}

        last_action = stack.pop()
        file_path = last_action.get("file_path")
        sheet_name = last_action.get("sheet_name")
        previous_fills = last_action.get("previous_fills", {})

        if file_path and Path(file_path).exists():
            wb = openpyxl.load_workbook(file_path)
            try:
                target = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
                ws = wb[target]
                for addr, fill_val in previous_fills.items():
                    try:
                        if fill_val:
                            ws[addr].fill = PatternFill(start_color=fill_val, end_color=fill_val, fill_type="solid")
                        else:
                            ws[addr].fill = PatternFill(fill_type=None)
                    except Exception:
                        pass
                wb.save(file_path)
                return {
                    "ok": True,
                    "action": "UNDO",
                    "restored_count": len(previous_fills),
                    "sheet": target,
                }
            finally:
                wb.close()

        return {"ok": False, "message": "Tệp gốc không tồn tại."}

    @classmethod
    def _record_undo_snapshot(
        cls,
        session_id: str,
        file_path: str,
        sheet_name: str,
        cell_addresses: List[str],
        action_name: str,
    ):
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            target = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
            ws = wb[target]
            previous_fills = {}
            for addr in cell_addresses:
                try:
                    cell = ws[addr]
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        previous_fills[addr] = str(cell.fill.start_color.rgb)
                    else:
                        previous_fills[addr] = None
                except Exception:
                    pass
            wb.close()

            cls._undo_stacks.setdefault(session_id, []).append({
                "action": action_name,
                "file_path": file_path,
                "sheet_name": target,
                "previous_fills": previous_fills,
            })
        except Exception as err:
            logger.debug("Failed to record undo snapshot: %s", err)


spreadsheet_action_engine = SpreadsheetActionEngine()
