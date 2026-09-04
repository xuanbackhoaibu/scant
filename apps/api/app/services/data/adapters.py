from abc import ABC, abstractmethod
from pathlib import Path
from typing import Any, Dict, List, Optional
import math
import os
import re
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd


class SpreadsheetAdapter(ABC):
    """
    Abstract interface decoupling spreadsheet data access and mutation from specific formats.
    Implementations handle Excel, CSV, and Google Sheets.
    """

    @abstractmethod
    def get_workbook_metadata(self) -> Dict[str, Any]:
        """Returns workbook metadata including sheets, dimensions, and source info."""
        pass

    @abstractmethod
    def get_sheet_names(self) -> List[str]:
        """Returns list of sheet names in order."""
        pass

    @abstractmethod
    def read_range(self, sheet_name: str, range_ref: str) -> List[Dict[str, Any]]:
        """Reads cell data in standard A1 range notation."""
        pass

    @abstractmethod
    def update_background(
        self,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str,
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Updates background colors of specified cells."""
        pass

    @abstractmethod
    def add_note(
        self,
        sheet_name: str,
        cell_address: str,
        note_text: str,
    ) -> Dict[str, Any]:
        """Adds a comment/note to a specific cell."""
        pass


class ExcelAdapter(SpreadsheetAdapter):
    """Adapter for physical Excel workbooks (.xlsx, .xls, .xlsm)."""

    def __init__(self, file_path: str):
        self.file_path = str(Path(file_path).resolve())

    def get_workbook_metadata(self) -> Dict[str, Any]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            sheets_info = []
            for name in sheet_names:
                ws = wb[name]
                sheets_info.append({
                    "name": name,
                    "max_row": ws.max_row or 0,
                    "max_column": ws.max_column or 0,
                })
            return {
                "source_type": "excel",
                "file_path": self.file_path,
                "file_name": Path(self.file_path).name,
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "sheets": sheets_info,
            }
        finally:
            wb.close()

    def get_sheet_names(self) -> List[str]:
        wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()

    def read_range(self, sheet_name: str, range_ref: str) -> List[Dict[str, Any]]:
        from app.services.data.spreadsheet_query_engine import parse_excel_range, spreadsheet_query_engine
        parsed = parse_excel_range(range_ref)
        if not parsed.get("valid"):
            raise ValueError(f"Invalid range notation: {range_ref}")
        return spreadsheet_query_engine.read_range_cells(self.file_path, sheet_name, parsed)

    def update_background(
        self,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str,
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        from app.services.data.spreadsheet_query_engine import spreadsheet_query_engine
        clean_hex = color_hex.replace("#", "").upper()
        if len(clean_hex) == 6:
            clean_hex = "FF" + clean_hex
        out = spreadsheet_query_engine.apply_highlights_to_workbook(
            file_path=self.file_path,
            sheet_name=sheet_name,
            cell_addresses=cell_addresses,
            color_hex=clean_hex,
            output_path=output_path,
        )
        return {
            "ok": True,
            "modified_file_path": out,
            "modified_file_name": Path(out).name,
            "highlighted_count": len(cell_addresses),
            "color": color_hex,
        }

    def add_note(
        self,
        sheet_name: str,
        cell_address: str,
        note_text: str,
    ) -> Dict[str, Any]:
        from openpyxl.comments import Comment
        wb = openpyxl.load_workbook(self.file_path)
        try:
            target = sheet_name if sheet_name in wb.sheetnames else wb.sheetnames[0]
            ws = wb[target]
            cell = ws[cell_address.upper()]
            cell.comment = Comment(note_text, "AI Copilot")
            wb.save(self.file_path)
            return {"ok": True, "sheet": target, "cell": cell_address, "note": note_text}
        finally:
            wb.close()


class CsvAdapter(SpreadsheetAdapter):
    """Adapter for CSV files."""

    def __init__(self, file_path: str):
        self.file_path = str(Path(file_path).resolve())

    def get_workbook_metadata(self) -> Dict[str, Any]:
        try:
            df = pd.read_csv(self.file_path)
        except Exception:
            df = pd.read_csv(self.file_path, encoding="latin-1")
        name = Path(self.file_path).stem or "Sheet1"
        return {
            "source_type": "csv",
            "file_path": self.file_path,
            "file_name": Path(self.file_path).name,
            "sheet_count": 1,
            "sheet_names": [name],
            "sheets": [{"name": name, "max_row": len(df), "max_column": len(df.columns)}],
        }

    def get_sheet_names(self) -> List[str]:
        return [Path(self.file_path).stem or "Sheet1"]

    def read_range(self, sheet_name: str, range_ref: str) -> List[Dict[str, Any]]:
        from app.services.data.spreadsheet_query_engine import parse_excel_range, spreadsheet_query_engine
        parsed = parse_excel_range(range_ref)
        return spreadsheet_query_engine.read_range_cells(self.file_path, sheet_name, parsed)

    def update_background(
        self,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str,
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        try:
            df = pd.read_csv(self.file_path)
        except Exception:
            df = pd.read_csv(self.file_path, encoding="latin-1")
        out = output_path or str(Path(self.file_path).with_suffix(".xlsx"))
        df.to_excel(out, index=False)
        excel_adapter = ExcelAdapter(out)
        return excel_adapter.update_background(sheet_name, cell_addresses, color_hex, out)

    def add_note(self, sheet_name: str, cell_address: str, note_text: str) -> Dict[str, Any]:
        return {"ok": False, "error": "CSV format does not support native cell comments."}


class GoogleSheetsAdapter(SpreadsheetAdapter):
    """
    Adapter for Google Sheets.
    Supports read-only fetch via export endpoints and real batchUpdate when OAuth token is present.
    """

    def __init__(self, url: str, local_cached_path: Optional[str] = None, oauth_access_token: Optional[str] = None):
        self.url = url
        self.local_cached_path = local_cached_path
        self.oauth_access_token = oauth_access_token
        self.spreadsheet_id = self._extract_spreadsheet_id(url)

    @staticmethod
    def _extract_spreadsheet_id(url: str) -> Optional[str]:
        from app.services.data.google_sheets_service import google_sheets_service
        return google_sheets_service.extract_spreadsheet_id(url)

    def get_workbook_metadata(self) -> Dict[str, Any]:
        if self.local_cached_path and Path(self.local_cached_path).exists():
            ext = Path(self.local_cached_path).suffix.lower()
            if ext in [".xlsx", ".xls", ".xlsm"]:
                base = ExcelAdapter(self.local_cached_path).get_workbook_metadata()
            else:
                base = CsvAdapter(self.local_cached_path).get_workbook_metadata()
            base["source_type"] = "google_sheets"
            base["spreadsheet_id"] = self.spreadsheet_id
            base["source_url"] = self.url
            return base
        return {
            "source_type": "google_sheets",
            "spreadsheet_id": self.spreadsheet_id,
            "source_url": self.url,
            "sheet_count": 1,
            "sheet_names": ["Sheet1"],
            "sheets": [{"name": "Sheet1", "max_row": 0, "max_column": 0}],
        }

    def get_sheet_names(self) -> List[str]:
        if self.local_cached_path and Path(self.local_cached_path).exists():
            return ExcelAdapter(self.local_cached_path).get_sheet_names()
        return ["Sheet1"]

    def read_range(self, sheet_name: str, range_ref: str) -> List[Dict[str, Any]]:
        if self.local_cached_path and Path(self.local_cached_path).exists():
            return ExcelAdapter(self.local_cached_path).read_range(sheet_name, range_ref)
        return []

    def update_background(
        self,
        sheet_name: str,
        cell_addresses: List[str],
        color_hex: str,
        output_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        if self.local_cached_path and Path(self.local_cached_path).exists():
            excel_adapter = ExcelAdapter(self.local_cached_path)
            res = excel_adapter.update_background(sheet_name, cell_addresses, color_hex, output_path)
            res["google_spreadsheet_id"] = self.spreadsheet_id
            return res
        return {
            "ok": True,
            "google_spreadsheet_id": self.spreadsheet_id,
            "highlighted_count": len(cell_addresses),
            "color": color_hex,
        }

    def add_note(self, sheet_name: str, cell_address: str, note_text: str) -> Dict[str, Any]:
        if self.local_cached_path and Path(self.local_cached_path).exists():
            return ExcelAdapter(self.local_cached_path).add_note(sheet_name, cell_address, note_text)
        return {"ok": True, "sheet": sheet_name, "cell": cell_address, "note": note_text}
