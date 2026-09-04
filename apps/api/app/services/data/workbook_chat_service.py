import json
import logging
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from app.services.ai.gateway import ai_gateway
from app.services.ai.types import AIRequest, AITaskType
from app.services.data.action_engine import spreadsheet_action_engine
from app.services.data.analysis_intent_parser import analysis_intent_parser
from app.services.data.google_sheets_service import google_sheets_service, index_to_col_letter
from app.services.data.sheet_resolvers import column_resolver, remove_diacritics, sheet_resolver
from app.services.data.spreadsheet_query_engine import (
    SpreadsheetQueryEngine,
    col_index_to_letter,
    extract_excel_ranges_from_text,
    parse_excel_range,
    resolve_sheet_name_in_wb,
    spreadsheet_query_engine,
)

logger = logging.getLogger(__name__)


class WorkbookChatService:
    """
    WorkbookChatService orchestrates natural language interactions with Excel workbooks.
    Translates user intent into deterministic SpreadsheetQueryEngine tool calls,
    executes computation on real XLSX data, and returns structured responses
    with direct UI spreadsheet actions (highlight, scroll, clear, apply).
    """

    # In-memory session context for multi-turn conversations
    _sessions: Dict[str, Dict[str, Any]] = {}

    @classmethod
    def get_session(cls, conv_id: str) -> Dict[str, Any]:
        if conv_id not in cls._sessions:
            cls._sessions[conv_id] = {
                "history": [],
                "last_query_result": None,
                "last_matched_cells": [],
                "last_ranges": [],
                "last_sheet": None,
                "last_entity": None,
                "last_analysis_result": None,
            }
        return cls._sessions[conv_id]

    @classmethod
    def update_session(
        cls,
        conv_id: str,
        sheet_name: str,
        query_result: Optional[Dict[str, Any]] = None,
        matched_cells: Optional[List[Dict[str, Any]]] = None,
        ranges: Optional[List[str]] = None,
    ):
        session = cls.get_session(conv_id)
        session["last_sheet"] = sheet_name
        if query_result is not None:
            session["last_query_result"] = query_result
        if matched_cells is not None:
            session["last_matched_cells"] = matched_cells
        if ranges is not None:
            session["last_ranges"] = ranges

    @classmethod
    def _norm_text(cls, text: Any) -> str:
        normalized = remove_diacritics(str(text or ""))
        normalized = normalized.replace("_", " ").replace("-", " ")
        return re.sub(r"\s+", " ", normalized).strip().lower()

    @classmethod
    def _compact_text(cls, text: Any) -> str:
        return re.sub(r"[\s_\-]+", "", cls._norm_text(text))

    @classmethod
    def _format_number(cls, value: Any) -> str:
        if value is None:
            return "không có dữ liệu"
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        if isinstance(value, (int, float)):
            return f"{value:,.0f}".replace(",", ".")
        return str(value)

    @classmethod
    def _column_name_from_message(cls, file_path: str, sheet_name: str, message: str, fallback: Optional[str] = None) -> Optional[str]:
        schema = spreadsheet_query_engine.get_sheet_schema(file_path, sheet_name)
        msg_norm = cls._norm_text(message)
        best_name = fallback
        best_len = 0
        for col in schema.get("columns", []):
            col_norm = cls._norm_text(col.get("name"))
            if col_norm and col_norm in msg_norm and len(col_norm) > best_len:
                best_name = col["name"]
                best_len = len(col_norm)
        return best_name

    @classmethod
    def _parse_vietnamese_number(cls, number_text: str, unit_text: str = "") -> Optional[int | float]:
        raw = str(number_text or "").strip().lower()
        unit = cls._norm_text(unit_text)
        if not raw:
            return None
        if "." in raw and "," not in raw and re.fullmatch(r"\d{1,3}(?:\.\d{3})+", raw):
            raw = raw.replace(".", "")
        else:
            raw = raw.replace(",", ".")
        try:
            value = float(raw)
        except ValueError:
            return None
        if unit in {"tr", "trieu", "m", "million"}:
            value *= 1_000_000
        elif unit in {"k", "nghin", "ngan"}:
            value *= 1_000
        return int(value) if value.is_integer() else value

    @classmethod
    def _resolve_filter_column(cls, file_path: str, sheet_name: str, message: str) -> Optional[str]:
        direct = cls._column_name_from_message(file_path, sheet_name, message)
        if direct:
            return direct
        schema = spreadsheet_query_engine.get_sheet_schema(file_path, sheet_name)
        candidate = re.sub(
            r"(?:>=|<=|>|<|=)|\b(?:tat ca|toan bo|nhung|nguoi|nhan vien|co|loc|hay|tim|to|boi|vang|mau|tren|lon hon|cao hon|duoi|nho hon|thap hon|tu|den|trieu|tr|nghin|ngan|k|m)\b|\d[\d\.,]*",
            " ",
            cls._norm_text(message),
        )
        candidate = re.sub(r"\s+", " ", candidate).strip()
        resolved = column_resolver.resolve_column(candidate or message, schema.get("columns", []))
        if resolved.get("found") and resolved.get("confidence", 0) >= 0.45:
            return resolved.get("name")
        return None

    @classmethod
    def _parse_numeric_filter(cls, file_path: str, sheet_name: str, message: str) -> Optional[Dict[str, Any]]:
        msg_norm = cls._norm_text(message)
        number_pattern = r"(\d+(?:[\.,]\d+)*)\s*(trieu|triệu|tr|m|million|k|nghin|nghìn|ngan|ngàn)?"

        range_match = re.search(rf"\btu\s+{number_pattern}\s+(?:den|toi|-)\s+{number_pattern}", msg_norm)
        if range_match:
            first_unit = range_match.group(2) or range_match.group(4) or ""
            second_unit = range_match.group(4) or range_match.group(2) or ""
            first = cls._parse_vietnamese_number(range_match.group(1), first_unit)
            second = cls._parse_vietnamese_number(range_match.group(3), second_unit)
            column = cls._resolve_filter_column(file_path, sheet_name, message)
            if column and first is not None and second is not None:
                return {"intent": "filter_rows", "column": column, "operator": "between", "value": [first, second]}

        symbol_match = re.search(rf"(>=|<=|>|<|=)\s*{number_pattern}", msg_norm)
        phrase_match = re.search(rf"\b(tren|lon hon|cao hon|duoi|nho hon|thap hon|bang|=)\s+{number_pattern}", msg_norm)
        operator = None
        raw_number = None
        unit = ""
        if symbol_match:
            operator = symbol_match.group(1)
            raw_number = symbol_match.group(2)
            unit = symbol_match.group(3) or ""
        elif phrase_match:
            phrase = phrase_match.group(1)
            operator = "<" if phrase in {"duoi", "nho hon", "thap hon"} else "=" if phrase in {"bang", "="} else ">"
            raw_number = phrase_match.group(2)
            unit = phrase_match.group(3) or ""

        threshold = cls._parse_vietnamese_number(raw_number or "", unit)
        column = cls._resolve_filter_column(file_path, sheet_name, message) if operator and threshold is not None else None
        if not column:
            return None
        return {"intent": "filter_rows", "column": column, "operator": operator, "value": threshold}

    @classmethod
    def _filter_rows_response(cls, file_path: str, sheet_name: str, parsed_filter: Dict[str, Any], highlight: bool = False) -> Dict[str, Any]:
        filtered = spreadsheet_query_engine.filter_rows(
            file_path=file_path,
            sheet_name=sheet_name,
            column_name=parsed_filter["column"],
            operator=parsed_filter["operator"],
            compare_value=parsed_filter["value"],
        )
        matched_cells = filtered.get("matched_cells", [])
        cell_addresses = [cell["address"] for cell in matched_cells]
        col_name = filtered.get("column", {}).get("name", parsed_filter["column"])
        value_text = (
            f"{cls._format_number(parsed_filter['value'][0])} đến {cls._format_number(parsed_filter['value'][1])}"
            if parsed_filter["operator"] == "between"
            else cls._format_number(parsed_filter["value"])
        )
        table_cols = ["Ma NV", "Mã NV", "Ho ten", "Họ tên", "Phong ban", "Phòng ban", col_name]
        rows = filtered.get("rows", [])
        header_names = []
        for name in table_cols:
            if rows and name in rows[0].get("record", {}) and name not in header_names:
                header_names.append(name)
        if col_name not in header_names:
            header_names.append(col_name)
        header = "| " + " | ".join(header_names) + " |"
        sep = "| " + " | ".join(["---"] * len(header_names)) + " |"
        body = []
        for row in rows[:20]:
            record = row.get("record", {})
            body.append("| " + " | ".join(cls._format_number(record.get(name, "")) for name in header_names) + " |")
        answer = (
            f"Tìm thấy **{len(rows)}** dòng có **{col_name} {parsed_filter['operator']} {value_text}** trên sheet **{filtered.get('sheet', sheet_name)}**.\n\n"
            + "\n".join([header, sep, *body])
        )
        actions = []
        if highlight and cell_addresses:
            actions.append({
                "type": "HIGHLIGHT_CELLS",
                "sheet": filtered.get("sheet", sheet_name),
                "cells": cell_addresses,
                "color": "#FEF08A",
                "autoScrollTo": cell_addresses[0],
            })
        return {
            "intent": "filter_rows",
            "answer": answer,
            "context": {"sheet": filtered.get("sheet", sheet_name), "ranges": filtered.get("evidence", {}).get("ranges", [])},
            "evidence": filtered.get("evidence"),
            "blocks": [
                {"type": "kpi", "title": "Dòng phù hợp", "value": len(rows), "subtext": f"{col_name} {parsed_filter['operator']} {value_text}"},
                *cls._build_source_blocks(filtered.get("evidence")),
            ],
            "result": filtered,
            "actions": actions,
            "pending_actions": [],
            "follow_up_context": {"result_set": {"operation": "FILTER_ROWS", "row_count": len(rows), "column": col_name}},
            "status_steps": [f"Đang đọc {sheet_name}...", "Đang nhận diện điều kiện lọc...", "Đang lọc dữ liệu thật..."],
        }

    @classmethod
    def _filtered_result_followup_response(cls, file_path: str, sheet_name: str, message: str, filtered: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if filtered.get("operation") != "FILTER_ROWS":
            return None
        msg_norm = cls._norm_text(message)
        rows = list(filtered.get("rows", []))
        if any(k in msg_norm for k in ["co bao nhieu", "bao nhieu nguoi", "bao nhieu dong", "dem"]):
            return {
                "intent": "aggregate",
                "answer": f"Result set trước đó có **{len(rows)}** dòng phù hợp.",
                "context": {"sheet": filtered.get("sheet", sheet_name), "ranges": filtered.get("evidence", {}).get("ranges", [])},
                "evidence": {**(filtered.get("evidence") or {}), "operation": "COUNT_FILTERED_ROWS", "rowCount": len(rows)},
                "blocks": [{"type": "kpi", "title": "Số dòng phù hợp", "value": len(rows), "subtext": filtered.get("sheet", sheet_name)}],
                "result": filtered,
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang đọc result set trước đó..."],
            }

        group_value_match = re.search(r"\bphong\s+([a-z0-9 ]+)$", msg_norm)
        if any(k in msg_norm for k in ["chi lay phong", "loc phong"]) and group_value_match:
            schema = cls._schema_columns(file_path, filtered.get("sheet", sheet_name))
            group_col = cls._first_group_column(schema, preferred_names=["phong ban", "department"])
            wanted = group_value_match.group(1).strip()
            if group_col:
                narrowed = [row for row in rows if wanted in cls._norm_text(row.get("record", {}).get(group_col, ""))]
                next_filtered = {**filtered, "rows": narrowed, "matched_count": len(narrowed)}
                return {
                    "intent": "filter_rows",
                    "answer": f"Còn **{len(narrowed)}** dòng trong result set trước đó thuộc **{group_col}: {wanted}**.",
                    "context": {"sheet": filtered.get("sheet", sheet_name), "ranges": filtered.get("evidence", {}).get("ranges", [])},
                    "evidence": {**(filtered.get("evidence") or {}), "operation": "FILTER_PREVIOUS_RESULT", "rowCount": len(narrowed)},
                    "blocks": [{"type": "kpi", "title": "Dòng còn lại", "value": len(narrowed), "subtext": group_col}],
                    "result": next_filtered,
                    "actions": [],
                    "pending_actions": [],
                    "status_steps": ["Đang lọc result set trước đó..."],
                }

        if any(k in msg_norm for k in ["sap xep", "xep"]):
            col_name = cls._column_name_from_message(file_path, filtered.get("sheet", sheet_name), message) or filtered.get("column", {}).get("name")
            descending = not any(k in msg_norm for k in ["tang dan", "thap len cao", "nho den lon"])
            sorted_rows = sorted(
                rows,
                key=lambda row: spreadsheet_query_engine._coerce_number(row.get("record", {}).get(col_name)) or 0,
                reverse=descending,
            )
            next_filtered = {**filtered, "rows": sorted_rows}
            return {
                "intent": "sort",
                "answer": f"Đã sắp xếp **{len(sorted_rows)}** dòng trong result set theo **{col_name}** {'giảm dần' if descending else 'tăng dần'}.",
                "context": {"sheet": filtered.get("sheet", sheet_name), "ranges": filtered.get("evidence", {}).get("ranges", [])},
                "evidence": {**(filtered.get("evidence") or {}), "operation": "SORT_FILTERED_ROWS", "rowCount": len(sorted_rows)},
                "blocks": [{"type": "kpi", "title": "Dòng đã sắp xếp", "value": len(sorted_rows), "subtext": col_name}],
                "result": next_filtered,
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang sắp xếp result set trước đó..."],
            }

        if any(k in msg_norm for k in ["tong", "tong luong"]):
            col_name = cls._column_name_from_message(file_path, filtered.get("sheet", sheet_name), message) or filtered.get("column", {}).get("name")
            total = sum(
                spreadsheet_query_engine._coerce_number(row.get("record", {}).get(col_name)) or 0
                for row in rows
            )
            return {
                "intent": "aggregate",
                "answer": f"Tổng **{col_name}** của result set trước đó là **{cls._format_number(total)}**.",
                "context": {"sheet": filtered.get("sheet", sheet_name), "ranges": filtered.get("evidence", {}).get("ranges", [])},
                "evidence": {**(filtered.get("evidence") or {}), "operation": "SUM_FILTERED_ROWS", "rowCount": len(rows)},
                "blocks": [{"type": "kpi", "title": f"Tổng {col_name}", "value": cls._format_number(total), "subtext": f"{len(rows)} dòng"}],
                "result": {**filtered, "aggregate": {"column": col_name, "operation": "SUM", "value": total}},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang tính trên result set trước đó..."],
            }

        return None

    @classmethod
    def _numeric_columns(cls, schema: Dict[str, Any]) -> List[Dict[str, Any]]:
        return [col for col in schema.get("columns", []) if col.get("type") == "numeric"]

    @classmethod
    def _numeric_column_fallback_response(cls, file_path: str, sheet_name: str) -> Dict[str, Any]:
        schema = cls._schema_columns(file_path, sheet_name)
        numeric_cols = cls._numeric_columns(schema)
        names = [str(col.get("name")) for col in numeric_cols if col.get("name")]
        if names:
            columns_text = "\n".join(f"• **{name}**" for name in names[:12])
            answer = (
                f"Tôi chưa xác định chắc chắn cột bạn muốn phân tích trên sheet **{schema['sheet']}**.\n"
                f"Các cột số hiện có:\n{columns_text}"
            )
        else:
            answer = f"Tôi chưa thấy cột số phù hợp để phân tích trên sheet **{schema['sheet']}**."
        evidence = schema.get("evidence", {"sheet": schema["sheet"], "ranges": [], "operation": "GET_SHEET_SCHEMA", "rowCount": 0})
        return {
            "answer": answer,
            "context": {"sheet": schema["sheet"], "ranges": evidence.get("ranges", [])},
            "evidence": evidence,
            "blocks": cls._build_source_blocks(evidence),
            "result": {"schema": schema, "numeric_columns": names},
            "actions": [],
            "pending_actions": [],
            "status_steps": [f"Đang đọc {schema['sheet']}...", "Đang xác định các cột số..."],
        }

    @classmethod
    def _choose_primary_numeric_column(cls, schema: Dict[str, Any]) -> Dict[str, Any]:
        numeric_cols = cls._numeric_columns(schema)
        if not numeric_cols:
            return {"column": None, "confidence": 0.0, "reason": "Không có cột số."}

        positive_terms = [
            "thuc linh", "net", "final", "take home", "tong", "total", "doanh thu", "revenue",
            "loi nhuan", "profit", "thu nhap", "income", "salary", "luong", "amount", "gia tri",
            "diem", "score", "kpi", "sales", "chi phi", "cost",
        ]
        strong_terms = ["thuc linh", "net", "final", "tong", "total", "doanh thu", "revenue", "loi nhuan", "profit"]
        negative_terms = [
            "ma", "id", "stt", "so thu tu", "phone", "dien thoai", "cccd", "cmnd", "ngay", "thang", "nam",
            "gio", "hour", "minute", "count", "so luong",
        ]

        best: Optional[Dict[str, Any]] = None
        best_score = -999.0
        for idx, col in enumerate(numeric_cols):
            name_norm = cls._norm_text(col.get("name"))
            score = 0.0
            for term in positive_terms:
                if term in name_norm:
                    score += 2.0
            for term in strong_terms:
                if term in name_norm:
                    score += 2.0
            for term in negative_terms:
                if re.search(rf"(?:^|\s){re.escape(term)}(?:\s|$)", name_norm):
                    score -= 3.0
            score += min(float(col.get("non_empty_count") or 0), 50.0) / 100.0
            score += idx * 0.03
            if score > best_score:
                best = col
                best_score = score

        confidence = min(max(best_score / 5.0, 0.0), 1.0)
        reason = "Chọn cột số có vai trò tổng hợp/kết quả cao nhất trong schema."
        return {"column": best, "confidence": round(confidence, 3), "reason": reason}

    @classmethod
    def _schema_columns(cls, file_path: str, sheet_name: str) -> Dict[str, Any]:
        return spreadsheet_query_engine.get_sheet_schema(file_path, sheet_name)

    @classmethod
    def _first_numeric_column(cls, schema: Dict[str, Any]) -> Optional[str]:
        for col in schema.get("columns", []):
            if col.get("type") == "numeric":
                return col.get("name")
        return None

    @classmethod
    def _first_group_column(cls, schema: Dict[str, Any], preferred_names: Optional[List[str]] = None) -> Optional[str]:
        preferred_norms = [cls._norm_text(name) for name in (preferred_names or [])]
        for col in schema.get("columns", []):
            col_norm = cls._norm_text(col.get("name"))
            if any(name and name in col_norm for name in preferred_norms):
                return col.get("name")
        for col in schema.get("columns", []):
            samples = [str(v).strip() for v in col.get("sample_values", []) if str(v).strip()]
            unique_count = len(set(samples))
            if col.get("type") != "numeric" and 1 < unique_count <= max(8, len(samples)):
                return col.get("name")
        return None

    @classmethod
    def _mentioned_group_values(cls, file_path: str, sheet_name: str, group_column: str, message: str) -> List[str]:
        _resolved_sheet, _header_row, _columns, records = spreadsheet_query_engine._sheet_records(file_path, sheet_name)
        msg_norm = cls._norm_text(message)
        values: List[str] = []
        for row in records:
            value = str(row.get(group_column) or "").strip()
            if value and cls._norm_text(value) in msg_norm and value not in values:
                values.append(value)
        return values

    @classmethod
    def _nearest_sheet(cls, candidate: str, available_sheets: List[str]) -> Optional[str]:
        if not candidate or not available_sheets:
            return None
        candidate_compact = cls._compact_text(candidate)
        for sheet in available_sheets:
            if cls._compact_text(sheet) == candidate_compact:
                return sheet
        candidate_tokens = set(cls._norm_text(candidate).split())
        best = None
        best_score = 0.0
        for sheet in available_sheets:
            sheet_tokens = set(cls._norm_text(sheet).split())
            score = len(candidate_tokens & sheet_tokens) / max(len(candidate_tokens | sheet_tokens), 1)
            if score > best_score:
                best = sheet
                best_score = score
        return best if best_score >= 0.45 else None

    @classmethod
    def _is_probably_column_phrase(cls, candidate: str) -> bool:
        candidate_norm = cls._norm_text(candidate)
        return bool(re.match(r"^(cot|column|truong|field)\b", candidate_norm))

    @classmethod
    def _candidate_has_sheet_marker(cls, text: str) -> bool:
        text_norm = cls._norm_text(text)
        return any(marker in text_norm for marker in ["sheet", "worksheet", "trang tinh", "bang tinh", "chuyen sang", "doi sang"])

    @classmethod
    def _explicit_sheet_error_if_needed(cls, message: str, available_sheets: List[str]) -> Optional[Dict[str, Any]]:
        candidate = cls.extract_sheet_mention_from_text(message, available_sheets=None)
        if not candidate or not available_sheets:
            return None
        if cls._is_probably_column_phrase(candidate):
            return None
        for sheet in available_sheets:
            if sheet == candidate or remove_diacritics(sheet) == remove_diacritics(candidate):
                return None
        suggested = cls._nearest_sheet(candidate, available_sheets)
        if suggested and cls._compact_text(suggested) == cls._compact_text(candidate):
            return {
                "answer": f"Không tìm thấy sheet **{candidate}**. Bạn có muốn dùng **{suggested}** không?",
                "context": {"sheet": None, "ranges": []},
                "evidence": {"sheet": None, "ranges": [], "operation": "RESOLVE_SHEET", "rowCount": 0},
                "blocks": [],
                "result": {},
                "actions": [],
                "pending_actions": [],
                "error": {"code": "SHEET_NOT_FOUND", "requested_sheet": candidate, "suggested_sheet": suggested},
                "status_steps": ["Đang kiểm tra danh sách sheet...", "Không tìm thấy sheet được yêu cầu."],
            }
        return None

    @classmethod
    def _build_source_blocks(cls, evidence: Optional[Dict[str, Any]]) -> List[Dict[str, Any]]:
        if not evidence:
            return []
        return [{"type": "source", "title": "Nguồn", "value": evidence.get("sheet"), "subtext": ", ".join(evidence.get("ranges", []))}]

    @classmethod
    def _scope_sheet_names(cls, scope: Optional[Dict[str, Any]], default_sheet: str, available_sheets: List[str]) -> List[str]:
        scope_type = (scope or {}).get("type")
        if not available_sheets:
            return [default_sheet or "Sheet1"]
        if scope_type == "workbook":
            return available_sheets
        if scope_type == "sheets":
            requested = (scope or {}).get("sheets") or []
            resolved: List[str] = []
            for sheet in requested:
                candidate = resolve_sheet_name_in_wb(str(sheet), available_sheets)
                if candidate in available_sheets and candidate not in resolved:
                    resolved.append(candidate)
            return resolved or [resolve_sheet_name_in_wb(default_sheet, available_sheets)]
        if scope_type == "range":
            requested_sheet = (scope or {}).get("sheet") or default_sheet
            return [resolve_sheet_name_in_wb(str(requested_sheet), available_sheets)]
        if scope_type == "sheet":
            requested_sheet = (scope or {}).get("sheet") or default_sheet
            return [resolve_sheet_name_in_wb(str(requested_sheet), available_sheets)]
        return [resolve_sheet_name_in_wb(default_sheet, available_sheets)]

    @classmethod
    def _merge_scoped_analysis_results(
        cls,
        prompt: str,
        sheet_results: List[Dict[str, Any]],
        operation: str,
    ) -> Dict[str, Any]:
        ranges: List[str] = []
        total_rows = 0
        lines = ["Tổng quan Workbook:" if operation == "WORKBOOK_ANALYSIS" else "Kết quả phân tích nhiều sheet:"]
        for item in sheet_results:
            evidence = item.get("evidence") or {}
            sheet = item.get("sheet") or evidence.get("sheet")
            result = item.get("result") or {}
            rows = evidence.get("rowCount") or result.get("row_count") or 0
            total_rows += int(rows or 0)
            for r in evidence.get("ranges", []):
                if r and r not in ranges:
                    ranges.append(r)
            summary_bits = []
            if "missing_count" in result:
                summary_bits.append(f"{result.get('missing_count', 0)} ô trống")
            if "duplicate_count" in result:
                summary_bits.append(f"{result.get('duplicate_count', 0)} nhóm trùng")
            if "outliers" in result:
                summary_bits.append(f"{len(result.get('outliers', []))} bất thường")
            if not summary_bits:
                summary_bits.append(item.get("answer", "đã phân tích"))
            lines.append(f"• **{sheet}**: " + ", ".join(summary_bits))

        evidence = {
            "sheet": "workbook" if operation == "WORKBOOK_ANALYSIS" else "multiple_sheets",
            "ranges": ranges,
            "operation": operation,
            "rowCount": total_rows,
        }
        return {
            "mode": "analysis_action",
            "answer": "\n".join(lines),
            "context": {"sheet": evidence["sheet"], "ranges": ranges},
            "evidence": evidence,
            "blocks": [
                {"type": "kpi", "title": "Sheet đã phân tích", "value": len(sheet_results), "subtext": prompt[:80]},
                *cls._build_source_blocks(evidence),
            ],
            "result": {"operation": operation, "sheet_results": sheet_results},
            "actions": [],
            "pending_actions": [],
            "analysis_history_item": {"prompt": prompt, "sheet": evidence["sheet"], "ranges": ranges, "operation": operation},
            "status_steps": ["Đang đọc workbook...", f"Đang phân tích {len(sheet_results)} sheet...", "Đang tổng hợp kết quả..."],
        }

    @classmethod
    async def analyze_action(
        cls,
        file_path: str,
        prompt: str,
        sheet_name: Optional[str] = None,
        selected_range: Optional[str] = None,
        conversation_id: Optional[str] = None,
        scope: Optional[Dict[str, Any]] = None,
        highlight_color: Optional[str] = None,
        data_source_url: Optional[str] = None,
        user: Optional[Any] = None,
        db: Optional[Any] = None,
        google_access_token: Optional[str] = None,
    ) -> Dict[str, Any]:
        conv_id = conversation_id or "analysis_action_default"
        session = cls.get_session(conv_id)

        available_sheets: List[str] = []
        if file_path and os.path.exists(file_path) and Path(file_path).suffix.lower() in [".xlsx", ".xls", ".xlsm"]:
            wb_temp = openpyxl.load_workbook(file_path, read_only=True)
            try:
                available_sheets = wb_temp.sheetnames
            finally:
                wb_temp.close()

        # Handle multi-sheet / workbook scope
        scope_type = (scope or {}).get("type")
        if scope_type in {"workbook", "sheets"}:
            target_sheets = cls._scope_sheet_names(scope, sheet_name or session.get("last_sheet") or "Sheet1", available_sheets)
            sheet_results = []
            for sheet in target_sheets:
                one = await cls.analyze_action(
                    file_path=file_path,
                    prompt=prompt,
                    sheet_name=sheet,
                    selected_range=None,
                    conversation_id=conversation_id,
                    scope={"type": "sheet", "sheet": sheet},
                    highlight_color=highlight_color,
                )
                sheet_results.append({
                    "sheet": one.get("context", {}).get("sheet", sheet),
                    "answer": one.get("answer"),
                    "evidence": one.get("evidence"),
                    "result": one.get("result"),
                })
            operation = "WORKBOOK_ANALYSIS" if scope_type == "workbook" else "MULTI_SHEET_ANALYSIS"
            return cls._merge_scoped_analysis_results(prompt, sheet_results, operation)

        # 1. Resolve Target Sheet
        default_target = sheet_name or session.get("last_sheet") or (available_sheets[0] if available_sheets else "Sheet1")
        if scope_type == "range":
            default_target = (scope or {}).get("sheet") or default_target
            selected_range = (scope or {}).get("range") or selected_range
        elif scope_type == "sheet":
            default_target = (scope or {}).get("sheet") or default_target

        # 2. Parse Intent with Structured AnalysisIntentParser
        parsed = analysis_intent_parser.parse(
            prompt=prompt,
            available_sheets=available_sheets,
            default_sheet=default_target,
            ui_highlight_color=highlight_color,
        )

        target_sheet = parsed.get("sheet") or default_target
        if available_sheets:
            target_sheet = resolve_sheet_name_in_wb(target_sheet, available_sheets)

        intent = parsed.get("intent", "SUMMARY")
        target_type = parsed.get("target_type", "cell")
        color = parsed.get("color") or "#FEF08A"

        # 3. Extract ranges
        ranges = extract_excel_ranges_from_text(prompt, sheet_name_hint=target_sheet, available_sheets=available_sheets)
        if not ranges and selected_range:
            ranges = [selected_range]

        schema = spreadsheet_query_engine.get_sheet_schema(file_path, target_sheet)
        columns_schema = schema.get("columns", [])
        last_col_letter = columns_schema[-1]["letter"] if columns_schema else "A"
        full_sheet_range = f"A{schema['header_row'] + 1}:{last_col_letter}{schema['header_row'] + schema['row_count']}"

        actions: List[Dict[str, Any]] = []
        result: Dict[str, Any] = {}
        evidence: Dict[str, Any] = {"sheet": target_sheet, "ranges": [], "operation": intent, "rowCount": schema.get("row_count", 0)}
        title = "Kết quả phân tích"
        answer = ""
        result_type = "scalar"

        logger.info(
            "\n[ANALYSIS DISPATCH]\nINPUT: '%s'\nINTENT: %s\nSHEET: %s\nCOLUMN MENTION: %s\nSCOPE: %s",
            prompt, intent, target_sheet, parsed.get("column_mention"), scope_type or "sheet"
        )

        # 4. Dispatch by Intent
        if intent == "CLEAR_HIGHLIGHT":
            spreadsheet_action_engine.clear_highlights(file_path, target_sheet)
            title = "Xóa màu đánh dấu"
            answer = f"Đã xóa toàn bộ màu đánh dấu trên sheet **{target_sheet}**."
            evidence = {"sheet": target_sheet, "ranges": [], "operation": "CLEAR_HIGHLIGHTS", "rowCount": 0}
            actions.append({"type": "CLEAR_HIGHLIGHTS", "sheet": target_sheet})
            result_type = "action"

        elif intent == "FIND_DUPLICATES":
            if not ranges:
                ranges = [full_sheet_range]
            result = spreadsheet_query_engine.find_duplicates(file_path, target_sheet, ranges, normalize=True, ignore_blank=True)
            matched_cells = result.get("matched_cells", [])
            cell_addresses = [c["address"] for c in matched_cells]
            evidence = {
                "sheet": result.get("sheet", target_sheet),
                "ranges": ranges,
                "operation": result.get("operation", "FIND_DUPLICATES"),
                "rowCount": result.get("total_matched_cells", 0),
            }
            if cell_addresses:
                if target_type == "row":
                    rows_to_highlight = sorted(list({c["row"] for c in matched_cells}))
                    actions.append({
                        "type": "HIGHLIGHT_ROWS",
                        "sheet": evidence["sheet"],
                        "rows": rows_to_highlight,
                        "color": color,
                        "autoScrollTo": f"A{rows_to_highlight[0]}",
                    })
                else:
                    actions.append({
                        "type": "HIGHLIGHT_CELLS",
                        "sheet": evidence["sheet"],
                        "cells": cell_addresses,
                        "color": color,
                        "autoScrollTo": cell_addresses[0],
                    })
            title = "Kiểm tra trùng lặp"
            dup_count = result.get("duplicate_count", 0)
            if dup_count == 0:
                answer = f"Đã kiểm tra trùng lặp trên sheet **{evidence['sheet']}** ({', '.join(ranges)}). Không phát hiện giá trị trùng lặp nào."
            else:
                answer = f"Đã kiểm tra trùng lặp trên **{evidence['sheet']}**. Phát hiện **{dup_count}** nhóm trùng, liên quan **{len(cell_addresses)}** ô."
            result_type = "duplicate"

        elif intent == "FIND_BLANKS":
            if not ranges:
                ranges = [full_sheet_range]
            result = spreadsheet_query_engine.find_missing_cells(file_path, target_sheet, ranges)
            matched_cells = result.get("matched_cells", [])
            cell_addresses = [c["address"] for c in matched_cells]
            evidence = {
                "sheet": result.get("sheet", target_sheet),
                "ranges": ranges,
                "operation": "FIND_MISSING",
                "rowCount": len(cell_addresses),
            }
            if cell_addresses:
                actions.append({
                    "type": "HIGHLIGHT_CELLS",
                    "sheet": evidence["sheet"],
                    "cells": cell_addresses,
                    "color": color or "#FED7AA",
                    "autoScrollTo": cell_addresses[0],
                })
            title = "Kiểm tra ô trống"
            missing_count = result.get("missing_count", 0)
            if missing_count == 0:
                answer = f"Đã kiểm tra ô trống trên sheet **{evidence['sheet']}**. Tất cả dữ liệu đều đầy đủ, không có ô trống."
            else:
                answer = f"Đã tìm ô trống trên **{evidence['sheet']}**. Phát hiện **{missing_count}** ô trống / thiếu dữ liệu."
            result_type = "blank"

        elif intent == "DETECT_OUTLIERS":
            col_mention = parsed.get("column_mention") or cls._column_name_from_message(file_path, target_sheet, prompt)
            col_info = column_resolver.resolve_column(col_mention, columns_schema) if col_mention else {}
            target_col_name = col_info.get("name") if col_info.get("found") else (cls._first_numeric_column(schema) or (columns_schema[0]["name"] if columns_schema else "Column 1"))
            result = spreadsheet_query_engine.detect_outliers(file_path, target_sheet, target_col_name)
            outliers = result.get("outliers", [])
            evidence = result.get("evidence", {"sheet": target_sheet, "ranges": [full_sheet_range], "operation": "DETECT_OUTLIERS", "rowCount": len(outliers)})
            if outliers:
                outlier_rows = [o["row_number"] for o in outliers]
                actions.append({
                    "type": "HIGHLIGHT_ROWS",
                    "sheet": target_sheet,
                    "rows": outlier_rows,
                    "color": color or "#FECACA",
                    "autoScrollTo": f"A{outlier_rows[0]}",
                })
            title = "Phân tích bất thường"
            answer = f"Đã phân tích giá trị bất thường cho cột **{target_col_name}** trên **{target_sheet}**. Phát hiện **{len(outliers)}** điểm bất thường."
            result_type = "outlier"

        elif intent in {"FIND_MAX", "FIND_MIN"}:
            is_max = (intent == "FIND_MAX")
            col_mention = parsed.get("column_mention") or cls._column_name_from_message(file_path, target_sheet, prompt)
            col_info = column_resolver.resolve_column(col_mention, columns_schema) if col_mention else {}
            target_col_name = col_info.get("name") if col_info.get("found") else None

            if not target_col_name:
                primary = cls._choose_primary_numeric_column(schema)
                if primary.get("column"):
                    target_col_name = primary["column"]["name"]
                else:
                    target_col_name = cls._first_numeric_column(schema) or (columns_schema[0]["name"] if columns_schema else "Column 1")

            top = spreadsheet_query_engine.find_top_rows(file_path, target_sheet, target_col_name, limit=1, descending=is_max)
            result = top
            evidence = top.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "MAX" if is_max else "MIN", "rowCount": 1})
            rows = top.get("rows", [])
            top_row = rows[0] if rows else {}
            record = top_row.get("record", {})
            r_num = top_row.get("row_number")

            label = record.get("Ho ten") or record.get("Họ tên") or record.get("Ten") or record.get("Tên") or record.get("Ma NV") or record.get("Mã NV") or (f"Dòng {r_num}" if r_num else "Dòng 1")
            code = record.get("Ma NV") or record.get("Mã NV")
            prefix = f"{label} ({code})" if code and code not in str(label) else str(label)
            val_formatted = cls._format_number(top_row.get("value"))

            title = f"{'Cao nhất' if is_max else 'Thấp nhất'}: {target_col_name}"
            answer = f"**{prefix}** có **{target_col_name}** {'cao nhất' if is_max else 'thấp nhất'}: **{val_formatted}**."
            result_type = "row"

            if r_num:
                target_col_letter = (
                    top.get("column", {}).get("letter")
                    or col_info.get("letter")
                    or (columns_schema[0]["letter"] if columns_schema else "A")
                )
                cell_address = f"{target_col_letter}{r_num}"
                if target_type == "row" or parsed.get("has_highlight"):
                    actions.append({
                        "type": "HIGHLIGHT_ROWS",
                        "sheet": target_sheet,
                        "rows": [r_num],
                        "color": color or ("#BBF7D0" if is_max else "#FED7AA"),
                        "autoScrollTo": f"A{r_num}",
                    })
                else:
                    actions.append({
                        "type": "HIGHLIGHT_CELLS",
                        "sheet": target_sheet,
                        "cells": [cell_address],
                        "color": color or ("#BBF7D0" if is_max else "#FED7AA"),
                        "autoScrollTo": cell_address,
                    })

        elif intent == "SUM":
            col_mention = parsed.get("column_mention") or cls._column_name_from_message(file_path, target_sheet, prompt)
            col_info = column_resolver.resolve_column(col_mention, columns_schema) if col_mention else {}
            target_col_name = col_info.get("name") if col_info.get("found") else (cls._first_numeric_column(schema) or (columns_schema[0]["name"] if columns_schema else "Column 1"))
            agg = spreadsheet_query_engine.sum(file_path, target_sheet, target_col_name)
            result = agg
            evidence = agg.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "SUM", "rowCount": schema.get("row_count", 0)})
            val_formatted = cls._format_number(agg.get("value"))
            title = f"Tổng {target_col_name}"
            answer = f"Tổng **{target_col_name}** trên sheet **{target_sheet}** là **{val_formatted}**."
            result_type = "scalar"

        elif intent == "AVERAGE":
            col_mention = parsed.get("column_mention") or cls._column_name_from_message(file_path, target_sheet, prompt)
            col_info = column_resolver.resolve_column(col_mention, columns_schema) if col_mention else {}
            target_col_name = col_info.get("name") if col_info.get("found") else (cls._first_numeric_column(schema) or (columns_schema[0]["name"] if columns_schema else "Column 1"))
            agg = spreadsheet_query_engine.average(file_path, target_sheet, target_col_name)
            result = agg
            evidence = agg.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "AVERAGE", "rowCount": schema.get("row_count", 0)})
            val_formatted = cls._format_number(agg.get("value"))
            title = f"Trung bình {target_col_name}"
            answer = f"Giá trị trung bình của **{target_col_name}** trên sheet **{target_sheet}** là **{val_formatted}**."
            result_type = "scalar"

        elif intent == "COUNT":
            cnt = spreadsheet_query_engine.count(file_path, target_sheet)
            result = cnt
            evidence = cnt.get("evidence", {"sheet": target_sheet, "ranges": [full_sheet_range], "operation": "COUNT", "rowCount": schema.get("row_count", 0)})
            title = "Tổng số dòng"
            answer = f"Sheet **{target_sheet}** có tổng cộng **{cls._format_number(cnt.get('value'))}** dòng dữ liệu ({len(columns_schema)} cột)."
            result_type = "scalar"

        elif intent == "COUNT_DISTINCT":
            col_mention = parsed.get("column_mention") or cls._column_name_from_message(file_path, target_sheet, prompt)
            col_info = column_resolver.resolve_column(col_mention, columns_schema) if col_mention else {}
            target_col_name = col_info.get("name") if col_info.get("found") else (columns_schema[0]["name"] if columns_schema else "Column 1")
            distinct_res = spreadsheet_query_engine.count_distinct(file_path, target_sheet, target_col_name)
            result = distinct_res
            evidence = distinct_res.get("evidence", {"sheet": target_sheet, "ranges": [full_sheet_range], "operation": "COUNT_DISTINCT", "rowCount": schema.get("row_count", 0)})
            title = f"Số lượng {target_col_name} duy nhất"
            answer = f"Phát hiện **{distinct_res.get('value', 0)}** giá trị duy nhất trong cột **{target_col_name}** trên sheet **{target_sheet}**."
            result_type = "scalar"

        elif intent == "CROSS_SHEET_COMPARE":
            other_sheets = [s for s in available_sheets if s != target_sheet]
            sheet2 = other_sheets[0] if other_sheets else target_sheet
            cmp_res = spreadsheet_query_engine.cross_sheet_compare(file_path, target_sheet, sheet2)
            result = cmp_res
            evidence = cmp_res.get("evidence", {"sheet": f"{target_sheet} vs {sheet2}", "ranges": [], "operation": "CROSS_SHEET_COMPARE", "rowCount": cmp_res.get("total_keys", 0)})
            title = f"Đối chiếu: {target_sheet} vs {sheet2}"
            disc_text = "\n".join(f"• {d}" for d in cmp_res.get("discrepancies", [])) or "Dữ liệu giữa 2 sheet khớp nhau."
            answer = f"Kết quả đối chiếu giữa **{target_sheet}** và **{sheet2}**:\n{disc_text}"
            result_type = "comparison"

        elif intent == "GENERATE_FORMULA":
            numeric_cols = [c for c in columns_schema if c.get("type") == "numeric"]
            target_col = numeric_cols[0] if numeric_cols else (columns_schema[0] if columns_schema else {"letter": "A", "name": "Cột"})
            col_letter = target_col.get("letter", "A")
            start_row = schema.get("header_row", 1) + 1
            end_row = start_row + schema.get("row_count", 1) - 1
            rng = f"{col_letter}{start_row}:{col_letter}{end_row}"

            formula_sum = f"=SUM({rng})"
            formula_avg = f"=AVERAGE({rng})"
            formula_max = f"=MAX({rng})"
            formula_min = f"=MIN({rng})"
            formula_count = f"=COUNT({rng})"

            title = "Công thức Excel gợi ý"
            answer = (
                f"📐 **Công thức Excel chuẩn xác cho cột '{target_col.get('name')}'** (vùng `{rng}`):\n\n"
                f"• **Tính Tổng**: `{formula_sum}`\n"
                f"• **Trung Bình**: `{formula_avg}`\n"
                f"• **Giá Trị Lớn Nhất**: `{formula_max}`\n"
                f"• **Giá Trị Nhỏ Nhất**: `{formula_min}`\n"
                f"• **Đếm số ô có giá trị**: `{formula_count}`\n\n"
                f"💡 *Bạn có thể sao chép trực tiếp các công thức trên để dán vào Excel hoặc Google Sheets.*"
            )
            result = {
                "formula_sum": formula_sum,
                "formula_avg": formula_avg,
                "formula_max": formula_max,
                "formula_min": formula_min,
                "range": rng,
                "column": target_col,
            }
            evidence = {"sheet": target_sheet, "ranges": [rng], "operation": "GENERATE_FORMULA", "rowCount": schema.get("row_count", 0)}
            result_type = "formula"

        else:
            # SUMMARY / OVERVIEW
            missing = spreadsheet_query_engine.find_missing_cells(file_path, target_sheet, [full_sheet_range])
            numeric_cols = [col["name"] for col in columns_schema if col.get("type") == "numeric"][:3]
            summaries = [spreadsheet_query_engine.sum(file_path, target_sheet, c) for c in numeric_cols]
            evidence = {"sheet": target_sheet, "ranges": [full_sheet_range], "operation": "SUMMARY", "rowCount": schema.get("row_count", 0)}
            num_lines = [f"• Tổng **{s['column']['name']}**: {cls._format_number(s.get('value'))}" for s in summaries if s.get("column")]
            title = f"Tổng quan {target_sheet}"
            answer = (
                f"Sheet **{target_sheet}** có **{schema.get('row_count', 0)} dòng dữ liệu** và **{len(columns_schema)} cột**.\n"
                f"Phát hiện **{missing.get('missing_count', 0)} ô trống**.\n"
                + "\n".join(num_lines)
            )
            result = {"schema": schema, "missing": missing, "summaries": summaries}
            result_type = "summary"

        # Collect all cells to highlight
        all_highlight_cells: List[str] = []
        for act in actions:
            if act.get("type") == "HIGHLIGHT_CELLS":
                all_highlight_cells.extend(act.get("cells", []))
            elif act.get("type") == "HIGHLIGHT_ROWS":
                for r in act.get("rows", []):
                    for col_item in columns_schema:
                        all_highlight_cells.append(f"{col_item['letter']}{r}")

        # Remove duplicates while preserving order
        seen_cells = set()
        unique_highlight_cells = []
        for c_addr in all_highlight_cells:
            if c_addr not in seen_cells:
                seen_cells.add(c_addr)
                unique_highlight_cells.append(c_addr)

        # Build structured cell findings
        structured_matched_cells = []
        for c_addr in unique_highlight_cells:
            r_idx, c_idx = google_sheets_service.parse_a1_coordinate(c_addr)
            col_ltr = index_to_col_letter(c_idx)
            row_num = r_idx + 1
            structured_matched_cells.append({
                "sheetName": target_sheet,
                "row": row_num,
                "column": col_ltr,
                "cell": c_addr,
                "range": f"'{target_sheet}'!{c_addr}",
                "value": result.get("value"),
                "reason": title,
            })
        if structured_matched_cells:
            result["matched_cells"] = structured_matched_cells

        # Real Google Sheets API Synchronization
        google_sync_info = {
            "is_google_sheet": False,
            "spreadsheet_id": None,
            "sheet_id": 0,
            "sheet_name": target_sheet,
            "cells": unique_highlight_cells,
            "color_hex": color,
            "google_sync_attempted": False,
            "synced_to_google_sheets": False,
            "verified_on_google_sheets": False,
            "google_sync_error": None,
        }

        spreadsheet_id = google_sheets_service.extract_spreadsheet_id(data_source_url)
        if spreadsheet_id and unique_highlight_cells:
            google_sync_info["is_google_sheet"] = True
            google_sync_info["spreadsheet_id"] = spreadsheet_id
            google_sync_info["google_sync_attempted"] = True

            access_token, token_err = await google_sheets_service.get_valid_access_token(
                user=user,
                db=db,
                explicit_token=google_access_token,
            )

            if access_token:
                sync_res = await google_sheets_service.highlight_cells(
                    spreadsheet_id=spreadsheet_id,
                    sheet_name=target_sheet,
                    cell_addresses=unique_highlight_cells,
                    color_hex=color,
                    access_token=access_token,
                    session_id=conv_id,
                )
                google_sync_info["sheet_id"] = sync_res.get("sheet_id", 0)
                google_sync_info["sheet_name"] = sync_res.get("sheet_name", target_sheet)
                google_sync_info["synced_to_google_sheets"] = sync_res.get("synced_to_google_sheets", False)
                google_sync_info["verified_on_google_sheets"] = sync_res.get("verified_on_google_sheets", False)
                google_sync_info["google_sync_error"] = sync_res.get("error")
            else:
                google_sync_info["synced_to_google_sheets"] = False
                google_sync_info["verified_on_google_sheets"] = False
                google_sync_info["google_sync_error"] = (
                    "Ứng dụng hiện chưa có quyền chỉnh sửa Google Sheets. Vui lòng cấp quyền chỉnh sửa để đồng bộ đánh dấu trực tiếp."
                )

        history_item = {
            "prompt": prompt,
            "sheet": evidence.get("sheet", target_sheet),
            "ranges": evidence.get("ranges", []),
            "operation": evidence.get("operation", intent),
            "summary": answer,
        }
        session["last_sheet"] = evidence.get("sheet", target_sheet)
        session["last_analysis_result"] = result
        session["last_query_result"] = result
        session["last_ranges"] = evidence.get("ranges", [])
        session["last_matched_cells"] = result.get("matched_cells", [])

        return {
            "mode": "analysis_action",
            "title": title,
            "result_type": result_type,
            "answer": answer,
            "context": {"sheet": evidence.get("sheet", target_sheet), "ranges": evidence.get("ranges", [])},
            "evidence": evidence,
            "blocks": [
                {
                    "type": "kpi",
                    "title": title,
                    "value": result.get("value") if result.get("value") is not None else result.get("duplicate_count", result.get("missing_count", len(result.get("outliers", [])))),
                    "subtext": evidence.get("sheet", target_sheet),
                },
                *cls._build_source_blocks(evidence),
            ],
            "result": result,
            "actions": actions,
            "pending_actions": [],
            "google_sync": google_sync_info,
            "analysis_history_item": history_item,
            "status_steps": [f"Đang đọc {evidence.get('sheet', target_sheet)}...", "Đang tính toán trên dữ liệu thật...", "Đang cập nhật Workspace..."],
        }

    @classmethod
    def extract_sheet_mention_from_text(
        cls,
        text: str,
        available_sheets: Optional[List[str]] = None,
    ) -> Optional[str]:
        """Extracts explicit sheet name from parentheses, quotes, keywords, or sheet list in prompt."""
        if not text:
            return None

        # 1. Check against actual available sheets in workbook (longest name first)
        if available_sheets:
            norm_text = cls._norm_text(text)
            compact_text = cls._compact_text(text)
            for s in sorted(available_sheets, key=lambda x: len(x), reverse=True):
                norm_s = cls._norm_text(s)
                compact_s = cls._compact_text(s)
                if len(norm_s) >= 2:
                    pattern = r"(?:\b|_|\s|^|\()" + re.escape(norm_s) + r"(?:\b|_|\s|$|\))"
                    if re.search(pattern, norm_text):
                        return s
                if len(compact_s) >= 3 and compact_s in compact_text:
                    return s

        # 2. Pattern: Parentheses, e.g. (HN Chính T8) or (sheet HN Chính T8)
        match_paren = re.search(r"\((?:sheet\s+)?([^)]+)\)", text, flags=re.IGNORECASE)
        if match_paren:
            candidate = match_paren.group(1).strip()
            if len(candidate) >= 2 and not candidate.startswith("http") and not cls._is_probably_column_phrase(candidate):
                return candidate

        # 3. Pattern: Quotes, e.g. 'HN Chính T8' or "HN Chính T8"
        match_quote = re.search(r"['\"](?:sheet\s+)?([^'\"]+)['\"]", text, flags=re.IGNORECASE)
        if match_quote:
            candidate = match_quote.group(1).strip()
            if len(candidate) >= 2 and not candidate.startswith("http") and not cls._is_probably_column_phrase(candidate):
                return candidate

        # 4. Explicit sheet markers only. Bare "trong/ở/tại/trên <phrase>" is too ambiguous
        # because column phrases often use the same Vietnamese prepositions.
        match_kw = re.search(
            r"(?:ở\s+sheet|trong\s+sheet|tại\s+sheet|trên\s+sheet|sheet|worksheet|trang\s+tính|bang\s+tinh|bảng\s+tính|trang)\s*[:=]?\s+([A-Za-z0-9_\u00C0-\u024F\u1EA0-\u1EF9\s]+?)(?=(?:,|\.|\?|!|\s+kiểm\s+tra|\s+xem|\s+từ|\s+so\s+sánh|\s+bôi\s+vàng|\s+tô\s+vàng|$))",
            text,
            flags=re.IGNORECASE,
        )
        if match_kw:
            candidate = match_kw.group(1).strip()
            if len(candidate) >= 2 and not candidate.startswith("http") and not re.match(r"^[A-Za-z]+\d+$", candidate) and not cls._is_probably_column_phrase(candidate):
                return candidate

        match_switch = re.search(
            r"(?:chuyển\s+sang|chuyen\s+sang|đổi\s+sang|doi\s+sang|mở\s+sheet|mo\s+sheet)\s+([A-Za-z0-9_\u00C0-\u024F\u1EA0-\u1EF9\s]+?)(?=(?:,|\.|\?|!|$))",
            text,
            flags=re.IGNORECASE,
        )
        if match_switch:
            candidate = match_switch.group(1).strip()
            if len(candidate) >= 2 and not cls._is_probably_column_phrase(candidate):
                return candidate

        return None

    @classmethod
    def resolve_target_sheet(
        cls,
        text: str,
        default_sheet: str,
        available_sheets: Optional[List[str]] = None,
    ) -> str:
        """
        Resolves target sheet from user text or fallback default sheet against workbook available sheets.
        Rule: explicit sheet in message > selected active sheet > default first sheet.
        """
        candidate = cls.extract_sheet_mention_from_text(text, available_sheets=available_sheets)
        if candidate and available_sheets:
            resolved = resolve_sheet_name_in_wb(candidate, available_sheets)
            if cls._nearest_sheet(candidate, available_sheets):
                return resolved
        elif candidate:
            return candidate

        if available_sheets:
            return resolve_sheet_name_in_wb(default_sheet, available_sheets)

        return default_sheet or "Sheet1"

    @classmethod
    def classify_intent(cls, message: str) -> str:
        """Classifies message into intent categories: GREETING, HELP, SMALL_TALK, WORKBOOK_METADATA, etc."""
        msg_norm = cls._norm_text(message)

        # 1. GREETING
        greeting_patterns = [
            r"^(xin\s+)?chao(\s+(ban|ad|em|bot|ai|moi\s+nguoi|nha|all))?[\s\.,!~?]*$",
            r"^(hello|hi|hey|alo|good\s+(morning|afternoon|evening)|hi\s+there)\b",
        ]
        if any(re.search(p, msg_norm) for p in greeting_patterns):
            return "GREETING"

        # 2. HELP / CAPABILITIES
        help_patterns = [
            r"(ban\s+)?(la\s+ai|lam\s+duoc\s+gi|lam\s+dc\s+gi|co\s+the\s+lam\s+gi|co\s+chuc\s+nang\s+gi)",
            r"\b(help|what\s+can\s+you\s+do|huong\s+dan|tro\s+giup|cach\s+dung|cach\s+su\s+dung)\b",
            r"(ban\s+co\s+the\s+giup\s+gi|ban\s+biet\s+lam\s+gi|tinh\s+nang\s+cua\s+ban)",
        ]
        if any(re.search(p, msg_norm) for p in help_patterns):
            return "HELP"

        # 3. SMALL TALK
        small_talk_patterns = [
            r"^(cam\s+on|thank\s+you|thanks|thank|tuyet\s+voi|tot\s+lam|good\s+job|ok|oki|oke|duoc\s+roi|hay\s+qua|good|tuyet)\b",
            r"^(tam\s+biet|bye|goodbye|hen\s+gap\s+lai)\b",
        ]
        if any(re.search(p, msg_norm) for p in small_talk_patterns):
            return "SMALL_TALK"

        # 4. WORKBOOK METADATA - SHEET COUNT / LIST
        if any(k in msg_norm for k in [
            "co bao nhieu sheet", "bao nhieu sheet", "may sheet", "co may sheet", "danh sach sheet",
            "cac sheet trong file", "cac sheet trong workbook", "danh sach trang tinh", "tong so sheet",
            "file nay co bao nhieu sheet", "workbook co bao nhieu sheet"
        ]):
            return "WORKBOOK_SHEET_COUNT"

        # 5. WORKBOOK METADATA - CURRENT SHEET
        if any(k in msg_norm for k in [
            "sheet hien tai", "dang xem sheet nao", "sheet dang chon", "ten sheet nay", "sheet nay ten gi", "sheet nay la gi"
        ]):
            return "WORKBOOK_CURRENT_SHEET"

        # 6. WORKBOOK METADATA - MAX SHEET
        if any(k in msg_norm for k in [
            "sheet nao nhieu dong nhat", "sheet nao lon nhat", "sheet nao nhieu du lieu nhat", "sheet lon nhat", "sheet nhieu dong nhat"
        ]):
            return "WORKBOOK_MAX_SHEET"

        # 7. WORKBOOK METADATA - FILE TOPIC
        if any(k in msg_norm for k in [
            "file nay noi ve gi", "file nay la gi", "noi dung file nay", "noi dung workbook", "bang tinh nay ve gi",
            "file nay chua gi", "file nay co gi"
        ]):
            return "WORKBOOK_FILE_TOPIC"

        return "WORKBOOK_DATA"

    @classmethod
    async def chat(
        cls,
        file_path: str,
        message: str,
        sheet_name: Optional[str] = None,
        selected_range: Optional[str] = None,
        conversation_id: Optional[str] = None,
        filters: Optional[List[Any]] = None,
    ) -> Dict[str, Any]:
        conv_id = conversation_id or "default_session"
        session = cls.get_session(conv_id)

        # 1. Load actual workbook sheet names if available
        available_sheets: List[str] = []
        if file_path and os.path.exists(file_path):
            ext = Path(file_path).suffix.lower()
            if ext in [".xlsx", ".xls", ".xlsm"]:
                try:
                    wb_temp = openpyxl.load_workbook(file_path, read_only=True)
                    available_sheets = wb_temp.sheetnames
                    wb_temp.close()
                except Exception as err:
                    logger.debug("Failed to read sheetnames from %s: %s", file_path, err)

        # 2. Intent Classification for Conversational & Metadata Queries
        intent = cls.classify_intent(message)

        if intent == "GREETING":
            return {
                "intent": "GREETING",
                "answer": (
                    "Xin chào! Tôi là AI Copilot hỗ trợ bạn với workbook hiện tại. "
                    "Bạn có thể hỏi tôi về các sheet, tra cứu dữ liệu, tính tổng, tìm giá trị cao nhất hoặc kiểm tra chất lượng dữ liệu."
                ),
                "context": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": []},
                "evidence": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": [], "operation": "GREETING", "rowCount": 0},
                "blocks": [],
                "result": {},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Sẵn sàng hỗ trợ!"],
            }

        if intent == "HELP":
            return {
                "intent": "HELP",
                "answer": (
                    "Tôi có thể hỗ trợ bạn các công việc sau trên workbook:\n"
                    "• **Tra cứu & Tổng hợp**: Tính tổng (SUM), trung bình (AVG), tìm người/dòng có giá trị cao nhất hoặc thấp nhất.\n"
                    "• **Cấu trúc Workbook**: Cho biết số lượng sheet, số dòng, số cột và kiểu dữ liệu từng sheet.\n"
                    "• **Kiểm tra dữ liệu**: Phát hiện các ô trùng lặp hoặc ô trống/thiếu dữ liệu.\n"
                    "• **Thao tác trực quan**: Đánh dấu màu (tô vàng/cam) các ô cần chú ý trực tiếp trên bảng tính."
                ),
                "context": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": []},
                "evidence": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": [], "operation": "HELP", "rowCount": 0},
                "blocks": [],
                "result": {},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Hướng dẫn sử dụng AI Copilot"],
            }

        if intent == "SMALL_TALK":
            return {
                "intent": "SMALL_TALK",
                "answer": "Rất vui được hỗ trợ bạn! Nếu cần tra cứu hay kiểm tra thêm số liệu nào trên bảng tính, hãy nhắn cho tôi nhé.",
                "context": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": []},
                "evidence": {"sheet": sheet_name or session.get("last_sheet") or "Sheet1", "ranges": [], "operation": "SMALL_TALK", "rowCount": 0},
                "blocks": [],
                "result": {},
                "actions": [],
                "pending_actions": [],
                "status_steps": [],
            }

        if intent == "WORKBOOK_SHEET_COUNT":
            count = len(available_sheets)
            sheet_details = []
            for s in available_sheets:
                try:
                    schema = spreadsheet_query_engine.get_sheet_schema(file_path, s)
                    r_count = schema.get("row_count", 0)
                    c_count = len(schema.get("columns", []))
                    sheet_details.append(f"• Sheet **{s}**: {r_count:,} dòng × {c_count} cột".replace(",", "."))
                except Exception:
                    sheet_details.append(f"• Sheet **{s}**")

            lines = [f"Workbook này có **{count} sheet**:"] + sheet_details
            evidence = {"sheet": "workbook", "ranges": [], "operation": "COUNT_SHEETS", "rowCount": count}
            return {
                "intent": "WORKBOOK_QUESTION",
                "answer": "\n".join(lines),
                "context": {"sheet": "workbook", "ranges": []},
                "evidence": evidence,
                "blocks": [{"type": "kpi", "title": "Tổng số sheet", "value": count, "subtext": "Workbook"}],
                "result": {"sheet_count": count, "sheets": available_sheets},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang kiểm tra danh sách sheet trong file..."],
            }

        if intent == "WORKBOOK_CURRENT_SHEET":
            resolved_target = cls.resolve_target_sheet(
                text=message,
                default_sheet=sheet_name or session.get("last_sheet") or (available_sheets[0] if available_sheets else "Sheet1"),
                available_sheets=available_sheets,
            )
            schema = {}
            try:
                schema = spreadsheet_query_engine.get_sheet_schema(file_path, resolved_target)
            except Exception:
                pass
            row_count = schema.get("row_count", 0)
            col_count = len(schema.get("columns", []))
            answer = f"Bạn đang xem sheet **{resolved_target}**"
            if row_count > 0:
                answer += f" (chứa **{row_count:,} dòng** và **{col_count} cột** dữ liệu).".replace(",", ".")
            else:
                answer += "."
            return {
                "intent": "WORKBOOK_QUESTION",
                "answer": answer,
                "context": {"sheet": resolved_target, "ranges": []},
                "evidence": {"sheet": resolved_target, "ranges": [], "operation": "CURRENT_SHEET", "rowCount": row_count},
                "blocks": [{"type": "kpi", "title": "Sheet hiện tại", "value": resolved_target, "subtext": f"{row_count} dòng"}],
                "result": {"current_sheet": resolved_target, "row_count": row_count, "column_count": col_count},
                "actions": [],
                "pending_actions": [],
                "status_steps": [f"Đang kiểm tra thông tin sheet {resolved_target}..."],
            }

        if intent == "WORKBOOK_MAX_SHEET":
            max_sheet = available_sheets[0] if available_sheets else "Sheet1"
            max_rows = 0
            for s in available_sheets:
                try:
                    schema = spreadsheet_query_engine.get_sheet_schema(file_path, s)
                    rc = schema.get("row_count", 0)
                    if rc > max_rows:
                        max_rows = rc
                        max_sheet = s
                except Exception:
                    pass
            session["last_sheet"] = max_sheet
            return {
                "intent": "WORKBOOK_QUESTION",
                "answer": f"Sheet **{max_sheet}** có nhiều dữ liệu nhất với **{max_rows:,}** dòng dữ liệu.".replace(",", "."),
                "context": {"sheet": max_sheet, "ranges": []},
                "evidence": {"sheet": max_sheet, "ranges": [], "operation": "MAX_SHEET", "rowCount": max_rows},
                "blocks": [{"type": "kpi", "title": "Sheet lớn nhất", "value": max_sheet, "subtext": f"{max_rows} dòng"}],
                "result": {"max_sheet": max_sheet, "max_rows": max_rows},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang so sánh quy mô các sheet..."],
            }

        if intent == "WORKBOOK_FILE_TOPIC":
            resolved_target = cls.resolve_target_sheet(
                text=message,
                default_sheet=sheet_name or session.get("last_sheet") or (available_sheets[0] if available_sheets else "Sheet1"),
                available_sheets=available_sheets,
            )
            file_name = Path(file_path).name if file_path else "Bảng tính"
            schema = {}
            try:
                schema = spreadsheet_query_engine.get_sheet_schema(file_path, resolved_target)
            except Exception:
                pass
            col_names = [c["name"] for c in schema.get("columns", [])[:8]]
            col_str = ", ".join(col_names) if col_names else "các cột số liệu"
            answer = (
                f"File **{file_name}** chứa **{len(available_sheets)} sheet** ({', '.join(available_sheets[:4])}).\n"
                f"Sheet **{resolved_target}** có **{schema.get('row_count', 0)} dòng** dữ liệu bao gồm các thông tin chính: {col_str}."
            )
            return {
                "intent": "WORKBOOK_QUESTION",
                "answer": answer,
                "context": {"sheet": resolved_target, "ranges": []},
                "evidence": {"sheet": resolved_target, "ranges": [], "operation": "FILE_TOPIC", "rowCount": schema.get("row_count", 0)},
                "blocks": [{"type": "kpi", "title": "Tập dữ liệu", "value": file_name, "subtext": f"{len(available_sheets)} sheet"}],
                "result": {"file_name": file_name, "sheets": available_sheets, "columns": col_names},
                "actions": [],
                "pending_actions": [],
                "status_steps": ["Đang tóm tắt cấu trúc file..."],
            }

        explicit_sheet_error = cls._explicit_sheet_error_if_needed(message, available_sheets)
        if explicit_sheet_error:
            return explicit_sheet_error

        # 3. Resolve single source of truth target sheet
        target_sheet = cls.resolve_target_sheet(
            text=message,
            default_sheet=sheet_name or session.get("last_sheet") or "Sheet1",
            available_sheets=available_sheets,
        )
        msg_lower = message.lower()
        msg_norm = cls._norm_text(message)

        # 1. Check Follow-Up Action: "Bôi vàng chúng" / "Tô vàng các ô vừa tìm" / "Highlight"
        is_highlight_followup = any(
            k in msg_lower for k in [
                "bôi vàng chúng", "tô vàng chúng", "bôi vàng các ô", "tô vàng các ô",
                "highlight chúng", "đánh dấu chúng", "bôi vàng tất cả", "tô vàng tất cả",
                "bôi màu chúng", "tô màu chúng", "highlight all", "tô vàng những người này",
                "bôi vàng những người này", "tô vàng các dòng này", "bôi vàng các dòng này"
            ]
        )
        if is_highlight_followup and session.get("last_matched_cells"):
            cached_cells = session["last_matched_cells"]
            cell_addresses = [c["address"] for c in cached_cells]
            first_cell = cell_addresses[0] if cell_addresses else None
            return {
                "answer": f"Tôi đã đánh dấu màu vàng {len(cell_addresses)} ô từ kết quả phân tích trước trên sheet **{target_sheet}**.",
                "context": {
                    "sheet": target_sheet,
                    "ranges": session.get("last_ranges", []),
                },
                "evidence": {
                    "sheet": target_sheet,
                    "ranges": session.get("last_ranges", []),
                    "operation": "HIGHLIGHT_CELLS",
                    "rowCount": len(cell_addresses),
                },
                "blocks": [
                    {
                        "type": "kpi",
                        "title": "Ô được đánh dấu",
                        "value": len(cell_addresses),
                        "subtext": f"Sheet {target_sheet}",
                    }
                ],
                "result": {
                    "highlighted_count": len(cell_addresses),
                    "matched_cells": cached_cells,
                },
                "actions": [
                    {
                        "type": "HIGHLIGHT_CELLS",
                        "sheet": target_sheet,
                        "cells": cell_addresses,
                        "style": "warning",
                        "color": "#FEF08A",  # Yellow-200
                        "autoScrollTo": first_cell,
                    }
                ],
                "pending_actions": [],
                "status_steps": [f"Đang đọc {target_sheet}...", "Đang áp dụng đánh dấu..."],
            }

        # 2. Check Action: "Xóa màu" / "Xóa đánh dấu" / "Clear highlight"
        is_clear_action = any(
            k in msg_lower for k in ["xóa màu", "xoa mau", "xóa đánh dấu", "xoa danh dau", "bỏ đánh dấu", "clear highlight", "clear màu"]
        )
        if is_clear_action:
            session["last_matched_cells"] = []
            return {
                "answer": f"Đã xóa toàn bộ đánh dấu màu trên bảng tính **{target_sheet}**.",
                "context": {"sheet": target_sheet},
                "evidence": {"sheet": target_sheet, "ranges": [], "operation": "CLEAR_HIGHLIGHTS", "rowCount": 0},
                "blocks": [],
                "result": {},
                "actions": [
                    {
                        "type": "CLEAR_HIGHLIGHTS",
                        "sheet": target_sheet,
                    }
                ],
                "pending_actions": [],
                "status_steps": ["Đang xóa đánh dấu..."],
            }

        filtered_followup = cls._filtered_result_followup_response(
            file_path=file_path,
            sheet_name=target_sheet,
            message=message,
            filtered=session.get("last_query_result") or {},
        )
        if filtered_followup:
            session["last_query_result"] = filtered_followup.get("result")
            session["last_analysis_result"] = filtered_followup.get("result")
            return filtered_followup

        # 3. Extract ranges from message (e.g. H6:H137 and I6:I137, ignoring sheet tokens)
        extracted_ranges = extract_excel_ranges_from_text(
            message,
            sheet_name_hint=target_sheet,
            available_sheets=available_sheets,
        )
        if not extracted_ranges and selected_range:
            extracted_ranges = [selected_range]

        should_auto_highlight = any(
            k in msg_lower for k in [
                "bôi vàng", "tô vàng", "highlight", "đánh dấu", "tô màu", "bôi màu",
                "boi vang", "to vang", "danh dau"
            ]
        )

        parsed_filter = cls._parse_numeric_filter(file_path, target_sheet, message)
        if parsed_filter:
            filter_response = cls._filter_rows_response(file_path, target_sheet, parsed_filter, highlight=should_auto_highlight)
            cls.update_session(
                conv_id,
                filter_response["context"]["sheet"],
                filter_response["result"],
                filter_response["result"].get("matched_cells", []),
                filter_response["context"].get("ranges", []),
            )
            session["last_analysis_result"] = filter_response["result"]
            return filter_response

        # 3a. Read-only follow-up: "người đó" / "anh ấy" / "dòng đó"
        last_entity = session.get("last_entity")
        if last_entity and any(token in msg_norm for token in ["nguoi do", "anh ay", "co ay", "dong do", "nhan vien do"]):
            target_col_name = cls._column_name_from_message(file_path, target_sheet, message)
            row_number = last_entity.get("row_number")
            if target_col_name and row_number:
                col = spreadsheet_query_engine.find_column(file_path, target_sheet, target_col_name)
                address = f"{col['letter']}{row_number}"
                cell = spreadsheet_query_engine.get_cell(file_path, target_sheet, address)
                evidence = cell.get("evidence")
                answer = (
                    f"{target_col_name} của {last_entity.get('label', 'người đó')} là "
                    f"**{cls._format_number(cell.get('value'))}**."
                )
                session["last_query_result"] = cell
                return {
                    "answer": answer,
                    "context": {"sheet": target_sheet, "ranges": evidence.get("ranges", []) if evidence else []},
                    "evidence": evidence,
                    "blocks": cls._build_source_blocks(evidence),
                    "result": cell,
                    "actions": [],
                    "pending_actions": [],
                    "follow_up_context": {"entity": last_entity},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang lấy ô liên quan..."],
                }

        # 3b. Pending UI action: highlight rows below/above a numeric threshold.
        if any(k in msg_norm for k in ["to vang", "boi vang", "highlight", "danh dau"]) and any(k in msg_norm for k in ["duoi", "nho hon", "thap hon", "tren", "lon hon", "cao hon"]):
            target_col_name = cls._column_name_from_message(file_path, target_sheet, message)
            threshold_match = re.search(r"(\d+(?:[\.,]\d+)?)\s*(trieu|triệu|m|k|nghin|ngàn|ngan)?", msg_lower)
            if target_col_name and threshold_match:
                raw_number = float(threshold_match.group(1).replace(",", "."))
                unit = threshold_match.group(2) or ""
                threshold = raw_number * 1_000_000 if unit in ["trieu", "triệu", "m"] else raw_number * 1_000 if unit in ["k", "nghin", "ngàn", "ngan"] else raw_number
                operator = "<" if any(k in msg_norm for k in ["duoi", "nho hon", "thap hon"]) else ">"
                filtered = spreadsheet_query_engine.filter_rows(file_path, target_sheet, target_col_name, operator, threshold)
                rows = [row["row_number"] for row in filtered.get("rows", [])]
                pending_action = {
                    "id": f"highlight_rows_{target_sheet}_{target_col_name}_{operator}_{int(threshold)}",
                    "type": "HIGHLIGHT_ROWS",
                    "sheet": target_sheet,
                    "rows": rows,
                    "color": "#FEF08A",
                    "requires_confirmation": True,
                    "label": f"Tô vàng {len(rows)} dòng",
                }
                session["last_query_result"] = filtered
                session["last_analysis_result"] = filtered
                return {
                    "answer": f"Đã tìm thấy **{len(rows)} dòng** có **{target_col_name} {operator} {cls._format_number(threshold)}**. Bạn có thể xác nhận để tô vàng các dòng này.",
                    "context": {"sheet": target_sheet, "ranges": filtered.get("evidence", {}).get("ranges", [])},
                    "evidence": filtered.get("evidence"),
                    "blocks": [
                        {"type": "kpi", "title": "Dòng phù hợp", "value": len(rows), "subtext": f"{target_col_name} {operator} {cls._format_number(threshold)}"},
                        *cls._build_source_blocks(filtered.get("evidence")),
                    ],
                    "result": filtered,
                    "actions": [],
                    "pending_actions": [pending_action],
                    "follow_up_context": {"entity": None},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang lọc dữ liệu...", "Đang chuẩn bị hành động cần xác nhận..."],
                }

        # 3c. Compare named groups using detected categorical and numeric columns.
        if any(k in msg_norm for k in ["so sanh", "doi chieu"]) and not extracted_ranges:
            schema = cls._schema_columns(file_path, target_sheet)
            value_col_name = cls._column_name_from_message(file_path, target_sheet, message, fallback=cls._first_numeric_column(schema))
            group_col_name = cls._first_group_column(schema, preferred_names=["phong ban", "department", "nhom", "group"])
            if value_col_name and group_col_name:
                compared = spreadsheet_query_engine.compare_groups(file_path, target_sheet, group_col_name, value_col_name, op="sum")
                mentioned_values = cls._mentioned_group_values(file_path, target_sheet, group_col_name, message)
                groups = compared.get("groups", [])
                if mentioned_values:
                    mentioned_norm = {cls._norm_text(value) for value in mentioned_values}
                    groups = [group for group in groups if cls._norm_text(group.get("group")) in mentioned_norm]
                evidence = compared.get("evidence")
                lines = [
                    f"• **{group['group']}**: {cls._format_number(group['value'])} ({group['count']} dòng)"
                    for group in groups[:8]
                ]
                return {
                    "answer": f"So sánh **{value_col_name}** theo **{group_col_name}** trên sheet **{compared['sheet']}**:\n" + "\n".join(lines),
                    "context": {"sheet": compared["sheet"], "ranges": evidence.get("ranges", []) if evidence else []},
                    "evidence": evidence,
                    "blocks": [
                        {
                            "type": "cellList",
                            "title": f"So sánh {group_col_name}",
                            "items": [{"value": group["group"], "count": group["count"], "cells": []} for group in groups[:8]],
                        },
                        *cls._build_source_blocks(evidence),
                    ],
                    "result": {**compared, "groups": groups},
                    "actions": [],
                    "pending_actions": [],
                    "follow_up_context": {"entity": None},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang gom nhóm dữ liệu...", "Đang so sánh..."],
                }

        # 3d. Deterministic sheet summary / issue overview.
        if any(k in msg_norm for k in ["tom tat", "phan tich nhanh", "sheet nay co van de gi", "du lieu nay noi len dieu gi", "co van de gi"]):
            schema = cls._schema_columns(file_path, target_sheet)
            last_col = schema["columns"][-1]["letter"] if schema.get("columns") else "A"
            full_range = f"A{schema['header_row'] + 1}:{last_col}{schema['header_row'] + schema['row_count']}"
            missing = spreadsheet_query_engine.find_missing_cells(file_path, target_sheet, [full_range])
            numeric_col_names = [col["name"] for col in schema.get("columns", []) if col.get("type") == "numeric"][:3]
            summaries = [spreadsheet_query_engine.aggregate_column(file_path, target_sheet, col_name, "sum") for col_name in numeric_col_names]
            evidence = {"sheet": schema["sheet"], "ranges": [full_range], "operation": "SHEET_SUMMARY", "rowCount": schema["row_count"]}
            numeric_lines = [
                f"• Tổng **{item['column']['name']}**: {cls._format_number(item.get('value'))}"
                for item in summaries
                if item.get("column")
            ]
            answer = (
                f"Sheet **{schema['sheet']}** có **{schema['row_count']} dòng dữ liệu** và **{len(schema.get('columns', []))} cột**.\n"
                f"Phát hiện **{missing.get('missing_count', 0)} ô trống / thiếu dữ liệu**.\n"
                + "\n".join(numeric_lines[:3])
            )
            return {
                "answer": answer,
                "context": {"sheet": schema["sheet"], "ranges": [full_range]},
                "evidence": evidence,
                "blocks": [
                    {"type": "kpi", "title": "Dòng dữ liệu", "value": schema["row_count"], "subtext": schema["sheet"]},
                    {"type": "kpi", "title": "Ô trống", "value": missing.get("missing_count", 0), "subtext": full_range},
                    *cls._build_source_blocks(evidence),
                ],
                "result": {"schema": schema, "missing": missing, "numeric_summaries": summaries},
                "actions": [],
                "pending_actions": [],
                "follow_up_context": {"entity": None},
                "status_steps": [f"Đang đọc {target_sheet}...", "Đang kiểm tra chất lượng dữ liệu...", "Đang tổng hợp kết quả..."],
            }

        # 3c. Dynamic read-only row count.
        if any(k in msg_norm for k in ["bao nhieu nhan vien", "co bao nhieu dong", "bao nhieu ban ghi", "bao nhieu nguoi"]):
            schema = spreadsheet_query_engine.get_sheet_schema(file_path, target_sheet)
            last_col_letter = schema["columns"][-1]["letter"] if schema.get("columns") else "A"
            evidence = {
                "sheet": schema["sheet"],
                "ranges": [f"A{schema['header_row'] + 1}:{last_col_letter}{schema['header_row'] + schema['row_count']}"],
                "operation": "COUNT_ROWS",
                "rowCount": schema["row_count"],
            }
            return {
                "answer": f"Sheet **{schema['sheet']}** có **{schema['row_count']}** dòng dữ liệu.",
                "context": {"sheet": schema["sheet"], "ranges": evidence["ranges"]},
                "evidence": evidence,
                "blocks": [{"type": "kpi", "title": "Số dòng dữ liệu", "value": schema["row_count"], "subtext": schema["sheet"]}, *cls._build_source_blocks(evidence)],
                "result": schema,
                "actions": [],
                "pending_actions": [],
                "follow_up_context": {"entity": None},
                "status_steps": [f"Đang đọc {target_sheet}...", "Đang đếm dòng dữ liệu..."],
            }

        # 3d. Aggregate questions: total / average / min / max for dynamically resolved columns.
        aggregate_op = None
        if any(k in msg_norm for k in ["tong", "tong cong", "tong so"]):
            aggregate_op = "sum"
        elif any(k in msg_norm for k in ["trung binh", "binh quan"]):
            aggregate_op = "average"
        if aggregate_op:
            target_col_name = cls._column_name_from_message(file_path, target_sheet, message)
            schema = cls._schema_columns(file_path, target_sheet)
            semantic_primary_metric = any(k in msg_norm for k in ["cot so quan trong nhat", "chi so chinh", "metric chinh", "cot chinh"])
            if not target_col_name and semantic_primary_metric:
                chosen = cls._choose_primary_numeric_column(schema)
                if chosen.get("column") and chosen.get("confidence", 0) >= 0.45:
                    target_col_name = chosen["column"]["name"]
            if target_col_name:
                agg = spreadsheet_query_engine.aggregate_column(file_path, target_sheet, target_col_name, aggregate_op)
                evidence = agg.get("evidence")
                op_label = "Tổng" if evidence and evidence.get("operation") == "SUM" else "Trung bình"
                return {
                    "intent": "aggregate",
                    "answer": f"{op_label} **{target_col_name}** trên sheet **{agg['sheet']}** là **{cls._format_number(agg.get('value'))}**.",
                    "context": {"sheet": agg["sheet"], "ranges": evidence.get("ranges", []) if evidence else []},
                    "evidence": evidence,
                    "blocks": [{"type": "kpi", "title": f"{op_label} {target_col_name}", "value": cls._format_number(agg.get("value")), "subtext": agg["sheet"]}, *cls._build_source_blocks(evidence)],
                    "result": agg,
                    "actions": [],
                    "pending_actions": [],
                    "follow_up_context": {"entity": None},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang tính toán..."],
                }
            return cls._numeric_column_fallback_response(file_path, target_sheet)

        # 3e. Top row questions.
        if any(k in msg_norm for k in ["cao nhat", "lon nhat", "thap nhat", "nho nhat"]):
            target_col_name = cls._column_name_from_message(file_path, target_sheet, message)
            semantic_primary_metric = any(k in msg_norm for k in ["cot so quan trong nhat", "chi so chinh", "metric chinh", "cot chinh", "gia tri cao nhat", "gia tri lon nhat", "gia tri thap nhat", "gia tri nho nhat"])
            semantic_note = ""
            if not target_col_name and semantic_primary_metric:
                schema = cls._schema_columns(file_path, target_sheet)
                chosen = cls._choose_primary_numeric_column(schema)
                if chosen.get("column") and chosen.get("confidence", 0) >= 0.45:
                    target_col_name = chosen["column"]["name"]
                    semantic_note = f"Nếu coi **{target_col_name}** là chỉ số chính của bảng này, "
                else:
                    return cls._numeric_column_fallback_response(file_path, target_sheet)
            if target_col_name:
                descending = any(k in msg_norm for k in ["cao nhat", "lon nhat"])
                top = spreadsheet_query_engine.find_top_rows(file_path, target_sheet, target_col_name, limit=1, descending=descending)
                evidence = top.get("evidence")
                row = top.get("rows", [{}])[0] if top.get("rows") else {}
                record = row.get("record", {})
                label = record.get("Ho ten") or record.get("Họ tên") or record.get("Ten") or record.get("Tên") or record.get("Ma NV") or record.get("Mã NV") or f"dòng {row.get('row_number')}"
                entity = {"row_number": row.get("row_number"), "label": label, "record": record}
                session["last_entity"] = entity
                session["last_query_result"] = top
                code = record.get("Ma NV") or record.get("Mã NV")
                prefix = f"{label} ({code})" if code and code not in str(label) else str(label)
                return {
                    "intent": "sort",
                    "answer": f"{semantic_note}**{prefix}** có **{target_col_name}** {'cao nhất' if descending else 'thấp nhất'}: **{cls._format_number(row.get('value'))}**.",
                    "context": {"sheet": top["sheet"], "ranges": evidence.get("ranges", []) if evidence else []},
                    "evidence": evidence,
                    "blocks": [{"type": "kpi", "title": f"{target_col_name} {'cao nhất' if descending else 'thấp nhất'}", "value": cls._format_number(row.get("value")), "subtext": prefix}, *cls._build_source_blocks(evidence)],
                    "result": top,
                    "actions": [{"type": "SCROLL_TO_CELL", "sheet": top["sheet"], "cells": [f"A{row.get('row_number')}"]}] if row.get("row_number") else [],
                    "pending_actions": [],
                    "follow_up_context": {"entity": entity},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang xếp hạng dữ liệu..."],
                }
            return cls._numeric_column_fallback_response(file_path, target_sheet)

        # 3f. Follow-up for exact number / details ("con số cụ thể", "bạn hãy cho tôi con số cụ thể")
        is_asking_exact_number = any(
            k in msg_norm for k in [
                "con so cu the", "cho toi con so", "so cu the", "bao nhieu cu the",
                "con so chinh xac", "so luong chinh xac", "cho toi biet so luong",
                "sao khong tra loi", "sao ko tra loi", "tra loi di"
            ]
        )
        if is_asking_exact_number:
            last_search = session.get("last_search_query")
            if last_search and last_search.get("result"):
                res = last_search["result"]
                term = last_search.get("query", "từ khóa")
                breakdown_lines = [f"• Cột **{col}**: **{cnt} lượt**" for col, cnt in res.get("breakdown_by_column", {}).items()]
                breakdown_text = "\n" + "\n".join(breakdown_lines) if breakdown_lines else ""
                cells_sample = ", ".join(c["address"] for c in res.get("matched_cells", [])[:10])
                answer = (
                    f"📊 **Con số cụ thể**: Phát hiện chính xác **{res.get('total_occurrences', 0)} lượt xuất hiện** "
                    f"(nằm trong **{res.get('unique_rows_count', 0)} dòng**) chứa **'{term}'** trên sheet **{res.get('sheet', target_sheet)}**.{breakdown_text}\n"
                    f"• Các ô tiêu biểu: {cells_sample}{'...' if len(res.get('matched_cells', [])) > 10 else ''}"
                )
                evidence = res.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "SEARCH_AND_COUNT", "rowCount": res.get("unique_rows_count", 0)})
                return {
                    "answer": answer,
                    "context": {"sheet": target_sheet, "ranges": evidence.get("ranges", [])},
                    "evidence": evidence,
                    "blocks": [
                        {"type": "kpi", "title": f"Số lượt xuất hiện '{term}'", "value": res.get("total_occurrences", 0), "subtext": f"{res.get('unique_rows_count', 0)} dòng"},
                        *cls._build_source_blocks(evidence),
                    ],
                    "result": res,
                    "actions": [
                        {
                            "type": "HIGHLIGHT_CELLS",
                            "sheet": target_sheet,
                            "cells": [c["address"] for c in res.get("matched_cells", [])],
                            "color": "#FEF08A",
                            "autoScrollTo": res.get("matched_cells", [{}])[0].get("address"),
                        }
                    ] if res.get("matched_cells") else [],
                    "pending_actions": [],
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang trích xuất con số cụ thể..."],
                }

            last_entity = session.get("last_entity")
            if last_entity:
                record = last_entity.get("record", {})
                facts = [f"• **{k}**: {cls._format_number(v)}" for k, v in list(record.items())[:8]]
                return {
                    "answer": f"Con số cụ thể cho **{last_entity.get('label', 'đối tượng')}**:\n" + "\n".join(facts),
                    "context": {"sheet": target_sheet},
                    "evidence": {"sheet": target_sheet, "ranges": [], "operation": "LOOKUP_ENTITY", "rowCount": 1},
                    "blocks": [],
                    "result": last_entity,
                    "actions": [],
                    "pending_actions": [],
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang lấy số liệu..."],
                }

        # 3g. Entity Search & Counting ("có bao nhiêu mỹ lộc", "tìm mỹ lộc", "đếm trạm X", "xe 90H04787")
        candidate_search = re.sub(
            r"^(?:cho tôi biết|cho toi biet|hãy cho tôi|hay cho toi|bạn hãy cho tôi|ban hay cho toi|bạn cho tôi|ban cho toi|có bao nhiêu|co bao nhieu|bao nhiêu|bao nhieu|đếm số|dem so|đếm|dem|tìm kiếm|tim kiem|tìm|tim|thống kê|thong ke|tra cứu|tra cuu|kiểm tra|kiem tra|cho biết|cho biet|có mấy|co may)\s+",
            "",
            message,
            flags=re.IGNORECASE
        ).strip()
        candidate_search = re.sub(
            r"\s+(?:trong sheet|o sheet|ở sheet|trên sheet|tren sheet|này|nay|trong bảng|trong bang|ở bảng|o bang|xuất hiện bao nhiêu lần|xuat hien bao nhieu lan|là bao nhiêu|la bao nhieu|có đúng không|co dung khong)\b.*$",
            "",
            candidate_search,
            flags=re.IGNORECASE
        ).strip()
        candidate_search = re.sub(r"^(?:chữ|chu|từ khóa|tu khoa)\s+", "", candidate_search, flags=re.IGNORECASE).strip()

        is_explicit_search_intent = any(k in msg_norm for k in ["tim chu", "tim tu khoa", "tim kiem", "tra cuu"])
        is_search_count_intent = any(
            k in msg_norm for k in [
                "co bao nhieu", "bao nhieu", "dem", "tim", "thong ke", "tra cuu",
                "xuat hien bao nhieu", "co may", "may lan"
            ]
        ) or (len(candidate_search.split()) <= 4 and len(candidate_search) >= 2)

        if is_search_count_intent and candidate_search and candidate_search.lower() not in ["sheet", "bang", "du lieu", "cot", "dong", "tat ca"]:
            search_res = spreadsheet_query_engine.search_and_count_entity(file_path, target_sheet, candidate_search)
            if search_res.get("total_occurrences", 0) > 0 or is_explicit_search_intent:
                total_cnt = search_res["total_occurrences"]
                unique_rows = search_res["unique_rows_count"]
                breakdown = search_res.get("breakdown_by_column", {})
                breakdown_lines = [f"• Cột **{col}**: **{cnt} lượt**" for col, cnt in breakdown.items()]
                breakdown_text = "\n" + "\n".join(breakdown_lines) if breakdown_lines else ""
                cells_sample = ", ".join(c["address"] for c in search_res.get("matched_cells", [])[:8])
                if total_cnt:
                    answer = (
                        f"Trên sheet **{target_sheet}**, phát hiện chính xác **{total_cnt} lượt xuất hiện** của "
                        f"**'{candidate_search}'** (nằm trên **{unique_rows} dòng**):{breakdown_text}\n"
                        f"• Các ô phát hiện: {cells_sample}{'...' if len(search_res.get('matched_cells', [])) > 8 else ''}\n"
                        f"✨ Đã tự động đánh dấu vàng các ô liên quan trên bảng tính."
                    )
                else:
                    answer = f"Không tìm thấy từ khóa **'{candidate_search}'** trong dữ liệu sheet **{target_sheet}**."

                session["last_search_query"] = {"query": candidate_search, "result": search_res}
                session["last_entity"] = {"label": candidate_search, "count": total_cnt}
                session["last_matched_cells"] = search_res["matched_cells"]

                evidence = search_res.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "SEARCH_AND_COUNT", "rowCount": unique_rows})
                return {
                    "intent": "search_text",
                    "answer": answer,
                    "context": {"sheet": target_sheet, "ranges": evidence.get("ranges", [])},
                    "evidence": evidence,
                    "blocks": [
                        {"type": "kpi", "title": f"Số lượt xuất hiện '{candidate_search}'", "value": total_cnt, "subtext": f"{unique_rows} dòng liên quan"},
                        *cls._build_source_blocks(evidence),
                    ],
                    "result": search_res,
                    "actions": [
                        {
                            "type": "HIGHLIGHT_CELLS",
                            "sheet": target_sheet,
                            "cells": [c["address"] for c in search_res["matched_cells"]],
                            "color": "#FEF08A",
                            "autoScrollTo": search_res["matched_cells"][0]["address"] if search_res["matched_cells"] else None,
                        }
                    ],
                    "pending_actions": [],
                    "status_steps": [f"Đang đọc {target_sheet}...", f"Đang tìm kiếm '{candidate_search}'...", "Đang tổng hợp kết quả..."],
                }

        # 3h. Row/entity lookup by visible id/code.
        code_match = re.search(r"\b([A-Za-z]{1,6}\d{2,})\b", message)
        if code_match:
            found = spreadsheet_query_engine.search_rows(file_path, target_sheet, code_match.group(1), limit=3)
            evidence = found.get("evidence")
            if found.get("matches"):
                match = found["matches"][0]
                record = match["record"]
                entity = {"row_number": match["row_number"], "label": code_match.group(1), "record": record}
                session["last_entity"] = entity
                facts = [f"**{k}**: {cls._format_number(v)}" for k, v in list(record.items())[:8]]
                return {
                    "answer": f"Thông tin của **{code_match.group(1)}**:\n" + "\n".join(f"• {fact}" for fact in facts),
                    "context": {"sheet": found["sheet"], "ranges": evidence.get("ranges", []) if evidence else []},
                    "evidence": evidence,
                    "blocks": cls._build_source_blocks(evidence),
                    "result": found,
                    "actions": [{"type": "SCROLL_TO_CELL", "sheet": found["sheet"], "cells": [f"A{match['row_number']}"]}],
                    "pending_actions": [],
                    "follow_up_context": {"entity": entity},
                    "status_steps": [f"Đang đọc {target_sheet}...", "Đang tìm dòng liên quan..."],
                }

        # 4. Check Duplicate Detection / Range Comparison Intent
        is_duplicate_intent = any(
            k in msg_lower for k in [
                "trùng", "trung", "duplicate", "lặp", "lap", "so sánh", "so sanh",
                "xuất hiện ở cả", "giao nhau", "intersection", "giống nhau", "khác nhau"
            ]
        )

        if is_duplicate_intent and extracted_ranges:
            query_res = spreadsheet_query_engine.find_duplicates(
                file_path=file_path,
                sheet_name=target_sheet,
                ranges=extracted_ranges,
                normalize=True,
                ignore_blank=True,
            )

            canonical_sheet = query_res.get("sheet", target_sheet)
            dup_count = query_res.get("duplicate_count", 0)
            cross_range_count = query_res.get("cross_range_count", 0)
            within_first_count = query_res.get("within_first_range_count", 0)
            within_second_count = query_res.get("within_second_range_count", 0)
            matched_cells = query_res.get("matched_cells", [])
            total_cells_count = len(matched_cells)
            cell_addresses = [c["address"] for c in matched_cells]
            first_match_addr = cell_addresses[0] if cell_addresses else None

            # Execution details for dev debugging
            execution_info = query_res.get("execution", {})

            logger.info(
                "[ExcelChatDebug] sheet: '%s' | ranges: %s | dup_count: %d | cross_dup: %d | internalA: %d | internalB: %d | matched_cells: %d",
                canonical_sheet,
                extracted_ranges,
                dup_count,
                cross_range_count,
                within_first_count,
                within_second_count,
                total_cells_count,
            )

            # Update session memory
            cls.update_session(conv_id, canonical_sheet, query_res, matched_cells, extracted_ranges)

            # Build readable Vietnamese answer
            ranges_str = " và ".join(extracted_ranges)
            if dup_count == 0:
                answer = (
                    f"Tôi đã kiểm tra dữ liệu trong vùng **{ranges_str}** trên sheet **{canonical_sheet}**.\n\n"
                    f"✅ **Kết quả**: Không phát hiện giá trị nào bị trùng lặp (cả trong từng vùng lẫn giữa các vùng, đã loại trừ các ô trống)."
                )
                blocks = [
                    {
                        "type": "kpi",
                        "title": "Giá trị trùng",
                        "value": 0,
                        "subtext": "Dữ liệu hoàn toàn độc nhất",
                    }
                ]
                actions = []
            else:
                cross_dups = query_res.get("cross_range_duplicates", [])
                within_a = query_res.get("duplicates_in_first_range", [])
                within_b = query_res.get("duplicates_in_second_range", [])

                parts = [f"Đã kiểm tra vùng **{ranges_str}** trên sheet **{canonical_sheet}**.\n"]

                if cross_dups:
                    parts.append(f"🔍 **Phát hiện {len(cross_dups)} giá trị xuất hiện ở cả hai cột ({query_res.get('first_range')} và {query_res.get('second_range')})**:")
                    for g in cross_dups[:5]:
                        f_cells = ", ".join(g.get("first_range_cells", [])[:4])
                        s_cells = ", ".join(g.get("second_range_cells", [])[:4])
                        parts.append(f"• **{g['value']}**: có tại {f_cells} và {s_cells} (tổng {g['total_occurrences']} lần)")

                if within_a and len(extracted_ranges) >= 2:
                    parts.append(f"\n🔍 **Phát hiện {len(within_a)} giá trị bị lặp nội bộ trong {query_res.get('first_range')}**:")
                    for g in within_a[:4]:
                        parts.append(f"• **{g['value']}**: lặp {g['count']} lần tại ({', '.join(g['cells'][:5])})")

                if within_b and len(extracted_ranges) >= 2:
                    parts.append(f"\n🔍 **Phát hiện {len(within_b)} giá trị bị lặp nội bộ trong {query_res.get('second_range')}**:")
                    for g in within_b[:4]:
                        parts.append(f"• **{g['value']}**: lặp {g['count']} lần tại ({', '.join(g['cells'][:5])})")

                if not cross_dups and (within_a or within_b) and len(extracted_ranges) == 1:
                    all_within = within_a or within_b
                    parts.append(f"🔍 **Phát hiện {len(all_within)} giá trị bị trùng lặp nội bộ**:")
                    for g in all_within[:5]:
                        parts.append(f"• **{g['value']}**: lặp {g['count']} lần tại ({', '.join(g['cells'][:5])})")

                if should_auto_highlight:
                    parts.append(f"\n✨ Tôi đã tự động **tô vàng** toàn bộ {total_cells_count} ô bị trùng trên bảng tính và cuộn tới ô đầu tiên ({first_match_addr}).")

                answer = "\n".join(parts)

                blocks = [
                    {
                        "type": "kpi",
                        "title": "Trùng giữa 2 vùng" if len(extracted_ranges) >= 2 else "Giá trị trùng",
                        "value": cross_range_count if len(extracted_ranges) >= 2 else dup_count,
                        "subtext": f"{ranges_str}" if len(extracted_ranges) >= 2 else "Trùng lặp",
                    },
                    {
                        "type": "kpi",
                        "title": "Tổng ô liên quan",
                        "value": total_cells_count,
                        "subtext": f"Sheet {canonical_sheet}",
                    },
                    {
                        "type": "cellList",
                        "title": "Danh sách các ô trùng lặp",
                        "items": [
                            {
                                "value": g["value"],
                                "count": g.get("total_occurrences", g.get("count", 0)),
                                "cells": (g.get("first_range_cells", []) + g.get("second_range_cells", [])) or g.get("cells", []),
                            }
                            for g in (cross_dups + within_a + within_b)[:25]
                        ],
                    },
                ]

                actions = []
                if should_auto_highlight and cell_addresses:
                    actions.append({
                        "type": "HIGHLIGHT_CELLS",
                        "sheet": canonical_sheet,
                        "cells": cell_addresses,
                        "style": "warning",
                        "color": "#FEF08A",  # Yellow-200
                        "autoScrollTo": first_match_addr,
                    })

            return {
                "answer": answer,
                "context": {
                    "sheet": canonical_sheet,
                    "ranges": extracted_ranges,
                },
                "evidence": {
                    "sheet": canonical_sheet,
                    "ranges": extracted_ranges,
                    "operation": query_res.get("operation", "FIND_DUPLICATES"),
                    "rowCount": query_res.get("execution", {}).get("range_a_non_empty", query_res.get("execution", {}).get("range_non_empty", 0)),
                },
                "blocks": blocks,
                "result": query_res,
                "actions": actions,
                "pending_actions": [],
                "status_steps": [f"Đang đọc {canonical_sheet}...", "Đang kiểm tra dữ liệu trùng...", "Đang dựng kết quả..."],
            }

        # 5. Check Missing / Blank Cells Intent
        is_missing_intent = any(
            k in msg_lower for k in ["trống", "trong", "thiếu", "thieu", "blank", "missing", "null", "rỗng", "rong"]
        )
        if is_missing_intent and extracted_ranges:
            query_res = spreadsheet_query_engine.find_missing_cells(
                file_path=file_path,
                sheet_name=target_sheet,
                ranges=extracted_ranges,
            )
            missing_count = query_res.get("missing_count", 0)
            matched_cells = query_res.get("matched_cells", [])
            cell_addresses = [c["address"] for c in matched_cells]

            cls.update_session(conv_id, target_sheet, query_res, matched_cells, extracted_ranges)

            ranges_str = " và ".join(extracted_ranges)
            if missing_count == 0:
                answer = f"Đã kiểm tra vùng **{ranges_str}** trên sheet **{target_sheet}**.\n\n✅ Tất cả các ô đều có dữ liệu đầy đủ, không có ô trống nào."
            else:
                answer = f"Đã kiểm tra vùng **{ranges_str}** trên sheet **{target_sheet}**.\n\n⚠️ Phát hiện **{missing_count} ô trống / thiếu dữ liệu** ({', '.join(cell_addresses[:15])}{'...' if len(cell_addresses) > 15 else ''})."
                if should_auto_highlight:
                    answer += f"\n\n✨ Đã đánh dấu màu cam các ô trống này trên bảng tính."

            actions = []
            if should_auto_highlight and cell_addresses:
                actions.append({
                    "type": "HIGHLIGHT_CELLS",
                    "sheet": target_sheet,
                    "cells": cell_addresses,
                    "style": "attention",
                    "color": "#FED7AA",  # Orange-200
                    "autoScrollTo": cell_addresses[0] if cell_addresses else None,
                })

            return {
                "answer": answer,
                "context": {
                    "sheet": target_sheet,
                    "ranges": extracted_ranges,
                },
                "evidence": {
                    "sheet": target_sheet,
                    "ranges": extracted_ranges,
                    "operation": "FIND_MISSING",
                    "rowCount": len(matched_cells),
                },
                "blocks": [
                    {
                        "type": "kpi",
                        "title": "Ô bị trống",
                        "value": missing_count,
                        "subtext": f"Vùng {ranges_str}",
                    }
                ],
                "result": query_res,
                "actions": actions,
                "pending_actions": [],
                "status_steps": [f"Đang đọc {target_sheet}...", "Đang tìm ô trống...", "Đang dựng kết quả..."],
            }

        # 6. General Intelligent Spreadsheet Q&A grounded in sheet data
        schema = cls._schema_columns(file_path, target_sheet)
        columns_text = ", ".join(f"**{col['name']}**" for col in schema.get("columns", [])[:8])
        row_cnt = schema.get("row_count", 0)

        # Check if query contains any word that matches cell values in this sheet
        words = [w for w in re.findall(r"[\w\d]+", message) if len(w) >= 3 and w.lower() not in ["cho", "toi", "biet", "hay", "trong", "sheet", "bang", "nay", "duoc", "khong", "the", "nao", "sao", "gi"]]
        for word in words:
            s_res = spreadsheet_query_engine.search_and_count_entity(file_path, target_sheet, word)
            if s_res.get("total_occurrences", 0) > 0:
                total_cnt = s_res["total_occurrences"]
                unique_rows = s_res["unique_rows_count"]
                breakdown = s_res.get("breakdown_by_column", {})
                b_lines = [f"• Cột **{col}**: **{cnt} lượt**" for col, cnt in breakdown.items()]
                b_text = "\n" + "\n".join(b_lines) if b_lines else ""
                cells_sample = ", ".join(c["address"] for c in s_res.get("matched_cells", [])[:8])
                session["last_search_query"] = {"query": word, "result": s_res}
                return {
                    "intent": "search_text",
                    "answer": f"Tìm thấy **{total_cnt} lượt xuất hiện** của từ khóa **'{word}'** trên sheet **{target_sheet}** (nằm trên **{unique_rows} dòng**):{b_text}\n• Các ô: {cells_sample}{'...' if len(s_res.get('matched_cells', [])) > 8 else ''}",
                    "context": {"sheet": target_sheet},
                    "evidence": s_res.get("evidence", {"sheet": target_sheet, "ranges": [], "operation": "SEARCH_AND_COUNT", "rowCount": unique_rows}),
                    "blocks": [
                        {"type": "kpi", "title": f"Từ khóa '{word}'", "value": total_cnt, "subtext": f"{unique_rows} dòng"},
                        *cls._build_source_blocks(s_res.get("evidence")),
                    ],
                    "result": s_res,
                    "actions": [
                        {
                            "type": "HIGHLIGHT_CELLS",
                            "sheet": target_sheet,
                            "cells": [c["address"] for c in s_res["matched_cells"]],
                            "color": "#FEF08A",
                            "autoScrollTo": s_res["matched_cells"][0]["address"] if s_res["matched_cells"] else None,
                        }
                    ],
                    "pending_actions": [],
                    "status_steps": [f"Đang đọc {target_sheet}...", f"Đang tra cứu '{word}'..."],
                }

        # Fallback response grounded in real sheet schema
        default_grounded_answer = (
            f"Trên sheet **{target_sheet}** hiện có **{row_cnt} dòng dữ liệu** với các cột:\n"
            f"{columns_text}.\n\n"
            f"Bạn có thể yêu cầu tôi:\n"
            f"• Đếm số lượng cụ thể (VD: *Có bao nhiêu Mỹ Lộc*, *Đếm loại xe...*)\n"
            f"• Tính tổng hoặc trung bình (VD: *Tổng giá tiền*, *Giá tiền trung bình...*)\n"
            f"• Tìm dòng lớn nhất / nhỏ nhất (VD: *Giao dịch giá tiền cao nhất...*)\n"
            f"• Kiểm tra trùng lặp hoặc ô trống trên bảng tính."
        )

        try:
            system_prompt = (
                f"Bạn là chuyên gia phân tích dữ liệu AI Copilot cho bảng tính Excel. "
                f"Sheet đang mở: '{target_sheet}', có {row_cnt} dòng. "
                f"Các cột: {columns_text}. "
                f"Hãy trả lời súc tích, chuyên nghiệp bằng tiếng Việt, tập trung vào số liệu thực tế của bảng tính."
            )
            user_prompt = f"Câu hỏi: '{message}'."
            req = AIRequest(
                task_type=AITaskType.DATA_NARRATIVE,
                prompt=f"{system_prompt}\n\n{user_prompt}",
                temperature=0.2,
                max_tokens=400,
            )
            res = await ai_gateway.execute(req)
            ai_text = getattr(res, "content", getattr(res, "text", "")).strip()
            if not ai_text or "Mục báo cáo" in ai_text or "đề tài nghiên cứu" in ai_text:
                ai_text = default_grounded_answer
        except Exception as err:
            logger.warning("AI Gateway error in WorkbookChatService: %s", err)
            ai_text = default_grounded_answer

        return {
            "answer": ai_text,
            "context": {"sheet": target_sheet},
            "evidence": {"sheet": target_sheet, "ranges": [], "operation": "AI_NARRATIVE", "rowCount": 0},
            "blocks": [],
            "result": {},
            "actions": [],
            "pending_actions": [],
            "status_steps": [f"Đang đọc {target_sheet}...", "Đang tổng hợp câu trả lời..."],
        }


workbook_chat_service = WorkbookChatService()
