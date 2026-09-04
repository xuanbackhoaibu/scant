from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import datetime
import math
import os
import re
import unicodedata
import openpyxl
import pandas as pd
from app.services.data.spreadsheet_query_engine import col_index_to_letter, remove_diacritics, resolve_sheet_name_in_wb


class WorkbookScanner:
    """
    High-speed, deterministic full-workbook scanner and structural analyzer.
    Extracts complete WorkbookContext across all sheets without querying external LLMs.
    """

    PREVIEW_ROW_LIMIT = 100

    @classmethod
    def _is_blank(cls, val: Any) -> bool:
        if val is None:
            return True
        if isinstance(val, float) and math.isnan(val):
            return True
        s = str(val).strip()
        return len(s) == 0 or s.lower() in ["nan", "null", "none"]

    @classmethod
    def _detect_data_type(cls, values: List[Any]) -> str:
        if not values:
            return "text"
        numeric_count = 0
        date_count = 0
        bool_count = 0
        formula_count = 0
        non_empty = [v for v in values if not cls._is_blank(v)]
        if not non_empty:
            return "text"

        for v in non_empty:
            if isinstance(v, str) and v.startswith("="):
                formula_count += 1
                continue
            if isinstance(v, bool):
                bool_count += 1
                continue
            if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
                numeric_count += 1
                continue
            if isinstance(v, (datetime.date, datetime.datetime, pd.Timestamp)):
                date_count += 1
                continue
            # Check string representation for numeric
            text = str(v).strip()
            # Currency or formatted number
            clean_num = re.sub(r"[^\d,\.\-]", "", text)
            if clean_num:
                if "," in clean_num and "." in clean_num:
                    clean_num = clean_num.replace(".", "").replace(",", ".")
                elif "," in clean_num:
                    clean_num = clean_num.replace(",", ".")
                try:
                    float(clean_num)
                    numeric_count += 1
                    continue
                except ValueError:
                    pass

        total = len(non_empty)
        if formula_count / total >= 0.5:
            return "formula"
        if bool_count / total >= 0.8:
            return "boolean"
        if date_count / total >= 0.7:
            return "date"
        if numeric_count / total >= 0.7:
            return "number"
        if (numeric_count + date_count + bool_count) / total >= 0.3 and (numeric_count + date_count + bool_count) / total < 0.7:
            return "mixed"
        return "text"

    @classmethod
    def _detect_header_row(cls, ws, max_scan: int = 20) -> int:
        best_row = 1
        best_score = -1
        max_rows = min(ws.max_row or 1, max_scan)
        max_cols = ws.max_column or 1

        for r in range(1, max_rows + 1):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_cols + 1)]
            non_empty_count = sum(1 for v in row_vals if not cls._is_blank(v))
            if non_empty_count == 0:
                continue
            # Header rows typically have more text strings than pure numbers
            text_count = sum(1 for v in row_vals if isinstance(v, str) and not cls._is_blank(v))
            score = non_empty_count * 2 + text_count
            if score > best_score:
                best_score = score
                best_row = r
        return best_row

    @classmethod
    def scan_workbook(
        cls,
        file_path: str,
        source_type: str = "excel",
        source_url: Optional[str] = None,
        preferred_active_sheet: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Scans all sheets of the given workbook, extracting unified typed WorkbookContext.
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Tệp không tồn tại: {file_path}")

        file_name = path.name
        ext = path.suffix.lower()
        workbook_id = f"wb_{path.stem}_{int(path.stat().st_mtime)}"

        # If CSV
        if ext == ".csv":
            return cls._scan_csv(file_path, workbook_id, file_name, source_type, source_url)

        # Excel XLSX/XLS/XLSM
        wb = openpyxl.load_workbook(file_path, data_only=True)
        try:
            sheet_names = list(wb.sheetnames)
            resolved_active = resolve_sheet_name_in_wb(preferred_active_sheet, sheet_names) if sheet_names else "Sheet1"
            sheets_context: List[Dict[str, Any]] = []
            total_rows_all = 0
            max_cols_all = 0

            for s_idx, sheet_name in enumerate(sheet_names):
                ws = wb[sheet_name]
                header_row = cls._detect_header_row(ws)
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                total_rows_all += max(0, max_row - header_row)
                if max_col > max_cols_all:
                    max_cols_all = max_col

                # Extract headers
                headers: List[str] = []
                seen_headers: Dict[str, int] = {}
                for c in range(1, max_col + 1):
                    val = ws.cell(row=header_row, column=c).value
                    h_name = str(val).strip() if not cls._is_blank(val) else f"Column_{col_index_to_letter(c)}"
                    if h_name in seen_headers:
                        seen_headers[h_name] += 1
                        h_name = f"{h_name}_{seen_headers[h_name]}"
                    else:
                        seen_headers[h_name] = 1
                    headers.append(h_name)

                # Scan columns
                columns_meta: List[Dict[str, Any]] = []
                for c_idx in range(1, max_col + 1):
                    letter = col_index_to_letter(c_idx)
                    col_name = headers[c_idx - 1]
                    cell_values = []
                    for r in range(header_row + 1, max_row + 1):
                        cell_values.append(ws.cell(row=r, column=c_idx).value)

                    non_empty = [v for v in cell_values if not cls._is_blank(v)]
                    empty_count = len(cell_values) - len(non_empty)
                    data_type = cls._detect_data_type(non_empty)
                    unique_vals = list({str(v).strip() for v in non_empty})
                    unique_count = len(unique_vals)

                    # Numeric statistics
                    min_val = None
                    max_val = None
                    sum_val = None
                    avg_val = None
                    if data_type == "number" and non_empty:
                        nums = []
                        for v in non_empty:
                            if isinstance(v, (int, float)) and not (isinstance(v, float) and math.isnan(v)):
                                nums.append(float(v))
                            else:
                                clean = re.sub(r"[^\d,\.\-]", "", str(v))
                                if clean:
                                    try:
                                        nums.append(float(clean.replace(",", ".")))
                                    except ValueError:
                                        pass
                        if nums:
                            min_val = round(min(nums), 4)
                            max_val = round(max(nums), 4)
                            sum_val = round(sum(nums), 4)
                            avg_val = round(sum(nums) / len(nums), 4)

                    columns_meta.append({
                        "index": c_idx,
                        "letter": letter,
                        "name": col_name,
                        "dataType": data_type,
                        "nonEmptyCount": len(non_empty),
                        "emptyCount": empty_count,
                        "uniqueCount": unique_count,
                        "sampleValues": [str(v) for v in non_empty[:5]],
                        "min": min_val,
                        "max": max_val,
                        "sum": sum_val,
                        "avg": avg_val,
                    })

                # Build preview records (up to PREVIEW_ROW_LIMIT)
                preview_records = []
                preview_end_row = min(max_row, header_row + cls.PREVIEW_ROW_LIMIT)
                for r in range(header_row + 1, preview_end_row + 1):
                    rec = {"_row_number": r}
                    has_val = False
                    for c_idx in range(1, max_col + 1):
                        v = ws.cell(row=r, column=c_idx).value
                        if not cls._is_blank(v):
                            has_val = True
                        rec[headers[c_idx - 1]] = str(v) if v is not None else ""
                    if has_val:
                        preview_records.append(rec)

                # Used range notation
                last_col_letter = col_index_to_letter(max_col) if max_col > 0 else "A"
                used_range = f"A{header_row}:{last_col_letter}{max_row}" if max_row > 0 else "A1"

                # Merged cells
                merged_ranges = [str(m) for m in ws.merged_cells.ranges]

                sheets_context.append({
                    "id": s_idx,
                    "name": sheet_name,
                    "rowCount": max(0, max_row - header_row),
                    "totalSheetRows": max_row,
                    "columnCount": max_col,
                    "usedRange": used_range,
                    "headerRow": header_row,
                    "headers": headers,
                    "columns": columns_meta,
                    "previewRecords": preview_records,
                    "mergedCells": merged_ranges,
                    "duplicateProfile": {
                        "hasDuplicates": any(c["nonEmptyCount"] > c["uniqueCount"] for c in columns_meta),
                    },
                })

            return {
                "workbookId": workbook_id,
                "sourceType": source_type,
                "fileName": file_name,
                "filePath": file_path,
                "sourceUrl": source_url,
                "sheetCount": len(sheet_names),
                "sheets": sheets_context,
                "activeSheet": resolved_active,
                "totalRowsAllSheets": total_rows_all,
                "maxColsAllSheets": max_cols_all,
                "status": "ready",
            }
        finally:
            wb.close()

    @classmethod
    def _scan_csv(
        cls,
        file_path: str,
        workbook_id: str,
        file_name: str,
        source_type: str,
        source_url: Optional[str],
    ) -> Dict[str, Any]:
        try:
            df = pd.read_csv(file_path)
        except Exception:
            df = pd.read_csv(file_path, encoding="latin-1")

        sheet_name = Path(file_name).stem or "Sheet1"
        columns_meta: List[Dict[str, Any]] = []
        headers = [str(c) for c in df.columns]

        for c_idx, col in enumerate(df.columns, start=1):
            letter = col_index_to_letter(c_idx)
            series = df[col]
            non_empty = series.dropna().tolist()
            empty_count = len(series) - len(non_empty)
            data_type = cls._detect_data_type(non_empty)
            unique_count = len(set(non_empty))

            min_val = None
            max_val = None
            sum_val = None
            avg_val = None
            if pd.api.types.is_numeric_dtype(series):
                valid_nums = series.dropna()
                if not valid_nums.empty:
                    min_val = round(float(valid_nums.min()), 4)
                    max_val = round(float(valid_nums.max()), 4)
                    sum_val = round(float(valid_nums.sum()), 4)
                    avg_val = round(float(valid_nums.mean()), 4)

            columns_meta.append({
                "index": c_idx,
                "letter": letter,
                "name": str(col),
                "dataType": data_type,
                "nonEmptyCount": len(non_empty),
                "emptyCount": empty_count,
                "uniqueCount": unique_count,
                "sampleValues": [str(v) for v in non_empty[:5]],
                "min": min_val,
                "max": max_val,
                "sum": sum_val,
                "avg": avg_val,
            })

        preview_records = []
        for r_idx, row in enumerate(df.head(cls.PREVIEW_ROW_LIMIT).to_dict(orient="records"), start=2):
            rec = {"_row_number": r_idx}
            for k, v in row.items():
                rec[str(k)] = str(v) if pd.notna(v) else ""
            preview_records.append(rec)

        last_col = col_index_to_letter(len(df.columns))
        used_range = f"A1:{last_col}{len(df) + 1}"

        sheet_ctx = {
            "id": 0,
            "name": sheet_name,
            "rowCount": len(df),
            "totalSheetRows": len(df) + 1,
            "columnCount": len(df.columns),
            "usedRange": used_range,
            "headerRow": 1,
            "headers": headers,
            "columns": columns_meta,
            "previewRecords": preview_records,
            "mergedCells": [],
            "duplicateProfile": {
                "hasDuplicates": any(c["nonEmptyCount"] > c["uniqueCount"] for c in columns_meta),
            },
        }

        return {
            "workbookId": workbook_id,
            "sourceType": source_type,
            "fileName": file_name,
            "filePath": file_path,
            "sourceUrl": source_url,
            "sheetCount": 1,
            "sheets": [sheet_ctx],
            "activeSheet": sheet_name,
            "totalRowsAllSheets": len(df),
            "maxColsAllSheets": len(df.columns),
            "status": "ready",
        }


workbook_scanner = WorkbookScanner()
