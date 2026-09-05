import html
import logging
import os
import re
import unicodedata
from typing import Any, Dict, List, Optional
import httpx
from sqlalchemy.orm import Session

from app.models.entities import Evidence, Source, UploadedFile, Dataset

logger = logging.getLogger(__name__)


class EvidenceService:
    """
    Evidence Extraction and Verification Service.
    Extracts verifiable chunks from:
    1. Web Articles & URLs (paragraph chunks, character offsets)
    2. PDF Documents (page numbers, paragraph indices)
    3. Word DOCX Documents (headings, section titles, paragraphs)
    4. Excel / CSV Datasets (sheet name, cell range, calculations: COUNT, SUM, AVG, MIN, MAX)
    5. Manual Evidence Quotes with locators
    """

    @staticmethod
    def normalize_text(text: str) -> str:
        if not text:
            return ""
        # Unicode normalize and collapse excess whitespace
        normalized = unicodedata.normalize("NFKC", text)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        return normalized

    @classmethod
    def extract_evidence_from_text(
        cls,
        text: str,
        evidence_type: str = "WEB_TEXT",
        source_url: Optional[str] = None,
        max_chunks: int = 50,
    ) -> List[Dict[str, Any]]:
        """Splits raw text into paragraph evidence chunks with character offsets."""
        chunks = []
        if not text:
            return chunks

        paragraphs = [p.strip() for p in re.split(r"\n\s*\n+", text) if p.strip()]
        current_offset = 0

        for idx, para in enumerate(paragraphs[:max_chunks]):
            start_off = text.find(para, current_offset)
            if start_off == -1:
                start_off = current_offset
            end_off = start_off + len(para)
            current_offset = end_off

            # Keep chunks with meaningful length (at least 20 chars)
            if len(para) >= 20:
                chunks.append({
                    "evidence_type": evidence_type,
                    "quote": para,
                    "normalized_text": cls.normalize_text(para),
                    "paragraph_index": idx + 1,
                    "start_offset": start_off,
                    "end_offset": end_off,
                    "source_url": source_url,
                    "metadata_json": {"char_length": len(para)},
                })

        return chunks

    @classmethod
    async def extract_evidence_from_url(
        cls,
        url: str,
        timeout: float = 10.0,
        max_chunks: int = 50,
    ) -> List[Dict[str, Any]]:
        """Fetches web page, extracts sanitized text paragraphs as evidence chunks."""
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        }
        async with httpx.AsyncClient(headers=headers, follow_redirects=True, timeout=timeout) as client:
            resp = await client.get(url)
            if resp.status_code >= 400:
                return []
            raw_html = resp.text

        # Strip scripts and styles
        cleaned_html = re.sub(r"<(script|style|nav|footer|header|aside)[^>]*>.*?</\1>", " ", raw_html, flags=re.DOTALL | re.IGNORECASE)
        # Convert break tags to newlines
        cleaned_html = re.sub(r"<(p|br|div|li|h[1-6])[^>]*>", "\n\n", cleaned_html, flags=re.IGNORECASE)
        # Strip all other HTML tags
        text = re.sub(r"<[^>]+>", " ", cleaned_html)
        text = html.unescape(text)

        return cls.extract_evidence_from_text(
            text=text,
            evidence_type="WEB_TEXT",
            source_url=url,
            max_chunks=max_chunks,
        )

    @classmethod
    def extract_evidence_from_pdf(
        cls,
        file_path: str,
        max_chunks: int = 100,
    ) -> List[Dict[str, Any]]:
        """Extracts PDF text chunks by page and block index using pymupdf."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"PDF file not found: {file_path}")

        import pymupdf as fitz

        chunks = []
        doc = fitz.open(file_path)
        chunk_count = 0

        for page_idx, page in enumerate(doc):
            blocks = page.get_text("blocks")  # (x0, y0, x1, y1, text, block_no, block_type)
            for block in blocks:
                if len(block) >= 5:
                    block_text = block[4].strip()
                    if len(block_text) >= 25:
                        chunk_count += 1
                        chunks.append({
                            "evidence_type": "PDF_TEXT",
                            "quote": block_text,
                            "normalized_text": cls.normalize_text(block_text),
                            "page_number": page_idx + 1,
                            "paragraph_index": block[5] if len(block) > 5 else chunk_count,
                            "metadata_json": {
                                "bbox": [round(b, 1) for b in block[:4]],
                                "total_pages": len(doc),
                            },
                        })
                        if chunk_count >= max_chunks:
                            break
            if chunk_count >= max_chunks:
                break

        doc.close()
        return chunks

    @classmethod
    def extract_evidence_from_docx(
        cls,
        file_path: str,
        max_chunks: int = 100,
    ) -> List[Dict[str, Any]]:
        """Extracts Word DOCX paragraphs with active section titles."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Docx file not found: {file_path}")

        import docx

        chunks = []
        doc = docx.Document(file_path)
        current_heading = "Mở đầu"
        para_idx = 0

        for p in doc.paragraphs:
            text = p.text.strip()
            if not text:
                continue

            style_name = p.style.name if p.style else ""
            if "Heading" in style_name or "heading" in style_name.lower():
                current_heading = text
                continue

            if len(text) >= 20:
                para_idx += 1
                chunks.append({
                    "evidence_type": "DOCX_TEXT",
                    "quote": text,
                    "normalized_text": cls.normalize_text(text),
                    "section_title": current_heading,
                    "paragraph_index": para_idx,
                    "metadata_json": {"style": style_name},
                })
                if len(chunks) >= max_chunks:
                    break

        return chunks

    @classmethod
    def calculate_excel_evidence(
        cls,
        file_path: str,
        sheet_name: Optional[str] = None,
        cell_range: Optional[str] = None,
        operation: str = "COUNT",
    ) -> Dict[str, Any]:
        """
        Extracts and verifies an exact Excel/CSV range with mathematical computation.
        Supported operations: COUNT, SUM, AVG, MIN, MAX, CELL_VALUE.
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        operation_upper = operation.upper()

        if ext == ".csv":
            import pandas as pd
            df = pd.read_csv(file_path)
            actual_sheet = sheet_name or "CSV_DATA"
            
            # If cell_range is specified as column name or row slice
            if cell_range and ":" in cell_range:
                m = re.match(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", cell_range)
                if m:
                    c1_str, r1_str, c2_str, r2_str = m.groups()
                    c1 = sum((ord(char.upper()) - 64) * (26 ** i) for i, char in enumerate(reversed(c1_str))) - 1
                    c2 = sum((ord(char.upper()) - 64) * (26 ** i) for i, char in enumerate(reversed(c2_str))) - 1
                    r1 = max(0, int(r1_str) - 1)
                    r2 = int(r2_str)
                    sliced = df.iloc[r1:r2, c1:c2 + 1]
                    values = sliced.values.flatten().tolist()
                else:
                    values = df.values.flatten().tolist()
            elif cell_range and cell_range in df.columns:
                values = df[cell_range].dropna().tolist()
            else:
                values = df.values.flatten().tolist()

        else:
            # Excel XLSX via openpyxl
            import openpyxl
            from openpyxl.utils import range_boundaries

            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                actual_sheet = sheet_name
            else:
                ws = wb.active
                actual_sheet = ws.title

            values = []
            if cell_range:
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                    for row in ws.iter_rows(
                        min_row=min_row,
                        max_row=max_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    ):
                        for cell_val in row:
                            if cell_val is not None:
                                values.append(cell_val)
                except Exception as e:
                    logger.warning(f"Could not parse range {cell_range}: {e}")
                    for row in ws.iter_rows(values_only=True):
                        for c in row:
                            if c is not None:
                                values.append(c)
            else:
                for row in ws.iter_rows(values_only=True):
                    for c in row:
                        if c is not None:
                            values.append(c)
            wb.close()

        # Compute operation
        numeric_values = []
        for v in values:
            if isinstance(v, (int, float)):
                numeric_values.append(float(v))
            elif isinstance(v, str):
                cleaned = v.replace(",", "").replace(".", "").strip()
                if cleaned.isdigit():
                    try:
                        numeric_values.append(float(v.replace(",", "")))
                    except ValueError:
                        pass

        calc_result_str = ""
        if operation_upper == "COUNT":
            count = len(values)
            calc_result_str = f"{count} giá trị"
        elif operation_upper == "SUM":
            total = sum(numeric_values)
            calc_result_str = f"{total:,.2f}".rstrip("0").rstrip(".") if numeric_values else "0"
        elif operation_upper == "AVG":
            avg = sum(numeric_values) / len(numeric_values) if numeric_values else 0
            calc_result_str = f"{avg:,.2f}".rstrip("0").rstrip(".")
        elif operation_upper == "MIN":
            min_val = min(numeric_values) if numeric_values else 0
            calc_result_str = f"{min_val:,.2f}".rstrip("0").rstrip(".")
        elif operation_upper == "MAX":
            max_val = max(numeric_values) if numeric_values else 0
            calc_result_str = f"{max_val:,.2f}".rstrip("0").rstrip(".")
        elif operation_upper == "CELL_VALUE":
            calc_result_str = str(values[0]) if values else "N/A"
        else:
            calc_result_str = f"{len(values)} giá trị (không xác định phép tính)"

        range_display = cell_range or "Toàn bộ bảng"
        quote = (
            f"[Bảng tính: '{actual_sheet}', Vùng: '{range_display}', "
            f"Phép tính: {operation_upper}] => Kết quả: {calc_result_str}"
        )

        return {
            "evidence_type": "EXCEL_RANGE",
            "sheet_name": actual_sheet,
            "cell_range": range_display,
            "operation": operation_upper,
            "calculation_result": calc_result_str,
            "quote": quote,
            "normalized_text": cls.normalize_text(quote),
            "metadata_json": {
                "total_items": len(values),
                "numeric_items": len(numeric_values),
                "sample_values": [str(v) for v in values[:5]],
            },
        }

    @classmethod
    def create_evidence(
        cls,
        db: Session,
        project_id: str,
        source_id: str,
        evidence_data: Dict[str, Any],
    ) -> Evidence:
        """Persists a single Evidence record into the database."""
        quote = evidence_data.get("quote", "").strip()
        normalized = evidence_data.get("normalized_text") or cls.normalize_text(quote)

        evidence = Evidence(
            project_id=project_id,
            source_id=source_id,
            evidence_type=evidence_data.get("evidence_type", "WEB_TEXT"),
            quote=quote,
            normalized_text=normalized,
            page_number=evidence_data.get("page_number"),
            section_title=evidence_data.get("section_title"),
            paragraph_index=evidence_data.get("paragraph_index"),
            start_offset=evidence_data.get("start_offset"),
            end_offset=evidence_data.get("end_offset"),
            sheet_name=evidence_data.get("sheet_name"),
            cell_range=evidence_data.get("cell_range"),
            operation=evidence_data.get("operation"),
            calculation_result=evidence_data.get("calculation_result"),
            source_url=evidence_data.get("source_url"),
            metadata_json=evidence_data.get("metadata_json", {}),
        )
        db.add(evidence)
        db.commit()
        db.refresh(evidence)
        return evidence

    @classmethod
    async def auto_extract_and_save_for_source(
        cls,
        db: Session,
        source: Source,
        max_chunks: int = 30,
    ) -> List[Evidence]:
        """Automatically parses a source and extracts initial evidence chunks."""
        extracted_chunks: List[Dict[str, Any]] = []

        # 1. If source points to an uploaded file
        if source.file_id:
            uploaded_file = db.query(UploadedFile).filter(UploadedFile.id == source.file_id).first()
            if uploaded_file and os.path.exists(uploaded_file.file_path):
                f_type = uploaded_file.file_type.lower()
                if f_type in ["pdf", ".pdf"]:
                    extracted_chunks = cls.extract_evidence_from_pdf(uploaded_file.file_path, max_chunks=max_chunks)
                elif f_type in ["docx", ".docx"]:
                    extracted_chunks = cls.extract_evidence_from_docx(uploaded_file.file_path, max_chunks=max_chunks)
                elif f_type in ["xlsx", "xls", "csv", ".xlsx", ".csv"]:
                    res = cls.calculate_excel_evidence(uploaded_file.file_path, operation="COUNT")
                    extracted_chunks = [res]

        # 2. If source has abstract or content_extracted
        if not extracted_chunks:
            text_to_chunk = source.content_extracted or source.abstract or source.summary
            if text_to_chunk and len(text_to_chunk.strip()) > 30:
                extracted_chunks = cls.extract_evidence_from_text(
                    text=text_to_chunk,
                    evidence_type="WEB_TEXT" if source.url else "DOCX_TEXT",
                    source_url=source.url,
                    max_chunks=max_chunks,
                )

        # 3. If source has a reachable URL and no text extracted yet
        if not extracted_chunks and source.url and source.url.startswith("http"):
            try:
                extracted_chunks = await cls.extract_evidence_from_url(source.url, max_chunks=max_chunks)
            except Exception as e:
                logger.warning(f"Could not extract chunks from {source.url}: {e}")

        # 4. Fallback: create single evidence chunk from title and subtitle
        if not extracted_chunks and source.title:
            quote = f"{source.title}. {source.subtitle or ''}".strip()
            extracted_chunks = [{
                "evidence_type": "WEB_TEXT",
                "quote": quote,
                "normalized_text": cls.normalize_text(quote),
                "paragraph_index": 1,
                "source_url": source.url,
                "metadata_json": {"is_title_fallback": True},
            }]

        saved_evidences = []
        for chunk in extracted_chunks:
            ev = cls.create_evidence(
                db=db,
                project_id=source.project_id,
                source_id=source.id,
                evidence_data=chunk,
            )
            saved_evidences.append(ev)

        return saved_evidences

    @classmethod
    def get_evidences_by_source(cls, db: Session, source_id: str) -> List[Evidence]:
        return db.query(Evidence).filter(Evidence.source_id == source_id).order_by(Evidence.page_number, Evidence.paragraph_index, Evidence.created_at).all()

    @classmethod
    def get_evidence_by_id(cls, db: Session, evidence_id: str) -> Optional[Evidence]:
        return db.query(Evidence).filter(Evidence.id == evidence_id).first()

    @classmethod
    async def create_evidence_async(
        cls,
        db: Any,
        project_id: str,
        source_id: str,
        evidence_data: Dict[str, Any],
    ) -> Evidence:
        quote = evidence_data.get("quote", "").strip()
        normalized = evidence_data.get("normalized_text") or cls.normalize_text(quote)

        evidence = Evidence(
            project_id=project_id,
            source_id=source_id,
            evidence_type=evidence_data.get("evidence_type", "WEB_TEXT"),
            quote=quote,
            normalized_text=normalized,
            page_number=evidence_data.get("page_number"),
            section_title=evidence_data.get("section_title"),
            paragraph_index=evidence_data.get("paragraph_index"),
            start_offset=evidence_data.get("start_offset"),
            end_offset=evidence_data.get("end_offset"),
            sheet_name=evidence_data.get("sheet_name"),
            cell_range=evidence_data.get("cell_range"),
            operation=evidence_data.get("operation"),
            calculation_result=evidence_data.get("calculation_result"),
            source_url=evidence_data.get("source_url"),
            metadata_json=evidence_data.get("metadata_json", {}),
        )
        db.add(evidence)
        await db.commit()
        await db.refresh(evidence)
        return evidence

    @classmethod
    async def auto_extract_and_save_async(
        cls,
        db: Any,
        source: Source,
        max_chunks: int = 30,
    ) -> List[Evidence]:
        """Automatically parses a source and extracts initial evidence chunks (AsyncSession)."""
        extracted_chunks: List[Dict[str, Any]] = []

        # 1. If source points to an uploaded file
        if source.file_id:
            from sqlalchemy import select
            stmt = select(UploadedFile).where(UploadedFile.id == source.file_id)
            res = await db.execute(stmt)
            uploaded_file = res.scalars().first()
            if uploaded_file and os.path.exists(uploaded_file.file_path):
                f_type = uploaded_file.file_type.lower()
                if f_type in ["pdf", ".pdf"]:
                    extracted_chunks = cls.extract_evidence_from_pdf(uploaded_file.file_path, max_chunks=max_chunks)
                elif f_type in ["docx", ".docx"]:
                    extracted_chunks = cls.extract_evidence_from_docx(uploaded_file.file_path, max_chunks=max_chunks)
                elif f_type in ["xlsx", "xls", "csv", ".xlsx", ".csv"]:
                    res_calc = cls.calculate_excel_evidence(uploaded_file.file_path, operation="COUNT")
                    extracted_chunks = [res_calc]

        # 2. If source has abstract or content_extracted
        if not extracted_chunks:
            text_to_chunk = source.content_extracted or source.abstract or source.summary
            if text_to_chunk and len(text_to_chunk.strip()) > 30:
                extracted_chunks = cls.extract_evidence_from_text(
                    text=text_to_chunk,
                    evidence_type="WEB_TEXT" if source.url else "DOCX_TEXT",
                    source_url=source.url,
                    max_chunks=max_chunks,
                )

        # 3. If source has a reachable URL and no text extracted yet
        if not extracted_chunks and source.url and source.url.startswith("http"):
            try:
                extracted_chunks = await cls.extract_evidence_from_url(source.url, max_chunks=max_chunks)
            except Exception as e:
                logger.warning(f"Could not extract chunks from {source.url}: {e}")

        # 4. Fallback: create single evidence chunk from title and subtitle
        if not extracted_chunks and source.title:
            quote = f"{source.title}. {source.subtitle or ''}".strip()
            extracted_chunks = [{
                "evidence_type": "WEB_TEXT",
                "quote": quote,
                "normalized_text": cls.normalize_text(quote),
                "paragraph_index": 1,
                "source_url": source.url,
                "metadata_json": {"is_title_fallback": True},
            }]

        saved_evidences = []
        for chunk in extracted_chunks:
            ev = await cls.create_evidence_async(
                db=db,
                project_id=source.project_id,
                source_id=source.id,
                evidence_data=chunk,
            )
            saved_evidences.append(ev)

        return saved_evidences


evidence_service = EvidenceService()
